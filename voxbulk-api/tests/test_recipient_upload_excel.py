"""Excel/CSV recipient upload parsing."""

from __future__ import annotations

import io

import pytest

from app.services.platform_catalog_service import ServiceOrderService


def test_parse_xlsx_with_mobile_number_header():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Full Name", "Mobile Number", "Language"])
    ws.append(["Sarah Ahmed", "+447700900123", "en"])
    buf = io.BytesIO()
    wb.save(buf)
    rows = ServiceOrderService.parse_recipient_file(buf.getvalue(), "contacts.xlsx")
    assert len(rows) == 1
    assert rows[0]["name"] == "Sarah Ahmed"
    assert rows[0]["phone"] == "+447700900123"


def test_parse_xlsx_phone_stored_as_integer():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "phone"])
    ws.append(["James Lee", 447700900456])
    buf = io.BytesIO()
    wb.save(buf)
    rows = ServiceOrderService.parse_recipient_file(buf.getvalue(), "contacts.xlsx")
    assert rows[0]["phone"] == "447700900456"


def test_parse_xlsx_detected_by_magic_bytes_without_extension():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "phone"])
    ws.append(["Alex", "+441234567890"])
    buf = io.BytesIO()
    wb.save(buf)
    rows = ServiceOrderService.parse_recipient_file(buf.getvalue(), "upload")
    assert len(rows) == 1


def test_parse_legacy_xls_rejected_with_clear_message():
    with pytest.raises(ValueError, match="Legacy .xls"):
        ServiceOrderService.parse_recipient_file(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64, "old.xls")
