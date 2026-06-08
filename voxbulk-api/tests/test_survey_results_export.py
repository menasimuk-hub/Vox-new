"""Tests for survey results PDF/CSV/Excel exports."""

from __future__ import annotations

import json

from app.services.survey_results_export_service import (
    build_survey_results_csv,
    build_survey_results_export_html,
    build_survey_results_xlsx,
    question_option_buckets,
)
from app.services.survey_results_service import build_survey_results_pdf


def _sample_payload() -> dict:
    return {
        "order": {
            "title": "Patient experience",
            "organisation_name": "Demo Clinic",
            "channel": "whatsapp",
            "goal": "Satisfaction",
        },
        "summary": {
            "total_recipients": 10,
            "completed_count": 8,
            "response_rate_pct": 80,
            "average_satisfaction_5": 4.0,
            "recommend_pct": 75,
            "nps_score": 62,
            "nps_label": "Good",
            "nps_promoters_pct": 50,
            "nps_passives_pct": 25,
            "nps_detractors_pct": 25,
            "average_call_duration_label": "WhatsApp survey",
            "sentiment_counts": {"positive": 5, "neutral": 2, "negative": 1},
        },
        "aggregates": [
            {
                "question": "How would you rate your visit?",
                "total": 8,
                "visualization": "sentiment_breakdown",
                "breakdown": [
                    {"key": "positive", "label": "Positive", "count": 3, "pct": 38},
                    {"key": "neutral", "label": "Neutral", "count": 3, "pct": 38},
                    {"key": "negative", "label": "Negative", "count": 2, "pct": 25},
                ],
                "responses": [{"answer": "9", "count": 3}, {"answer": "7", "count": 3}, {"answer": "5", "count": 2}],
            },
            {
                "question": "Would you recommend us?",
                "total": 8,
                "responses": [{"answer": "Yes", "count": 6}, {"answer": "No", "count": 2}],
            },
        ],
        "recommendations": [{"title": "Improve waits", "text": "Review queue times.", "impact": "High"}],
        "respondents": [
            {
                "id": "r1",
                "name": "Alex",
                "initials": "AL",
                "status": "completed",
                "final_feedback_yes_no": "Yes",
                "final_additional_feedback": "Great team",
                "wa_answers": [],
            }
        ],
    }


def test_question_option_buckets_rating_and_yes_no():
    payload = _sample_payload()
    rating = question_option_buckets(payload["aggregates"][0])
    assert [row["label"] for row in rating] == ["Excellent", "Expected", "Poor"]
    assert sum(row["count"] for row in rating) == 8

    yes_no = question_option_buckets(payload["aggregates"][1])
    assert yes_no == [
        {"label": "Yes", "count": 6, "pct": 75},
        {"label": "No", "count": 2, "pct": 25},
    ]


def test_build_survey_results_csv_includes_kpis_and_breakdown():
    csv_text = build_survey_results_csv(_sample_payload())
    assert csv_text.startswith("\ufeff")
    assert "Excellent rate %" in csv_text
    assert "Question breakdown" in csv_text
    assert "Excellent" in csv_text
    assert "Yes" in csv_text
    assert "Alex" in csv_text


def test_build_survey_results_csv_anonymous_uses_initials():
    csv_text = build_survey_results_csv(_sample_payload(), anonymous=True)
    assert "AL" in csv_text
    assert "Alex" not in csv_text


def test_build_survey_results_xlsx_has_summary_and_questions():
    data = build_survey_results_xlsx(_sample_payload())
    assert data[:2] == b"PK"

    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    assert "Summary" in wb.sheetnames
    assert "Questions" in wb.sheetnames
    summary = wb["Summary"]
    values = [row[0].value for row in summary.iter_rows(min_row=2, max_col=1)]
    assert "Completed responses" in values


def test_build_survey_results_export_html_contains_kpis():
    html = build_survey_results_export_html(_sample_payload(), logo_uri="")
    assert "Excellent rate" in html
    assert "Question breakdown" in html
    assert "Excellent" in html
    assert "Yes" in html
    assert "62" in html


def test_build_survey_results_pdf_bytes():
    pdf = build_survey_results_pdf(_sample_payload())
    assert pdf.startswith(b"%PDF")
