#!/usr/bin/env python3
"""Generate HubSpot test import Excel (local file only — not committed).

Creates sheets for:
  - Appointment Manager sync (contacts + appointment_date)
  - WA survey manual recipients
  - Won-deal CRM survey automation (contacts + deals)

Usage:
  cd voxbulk-api
  python scripts/generate_hubspot_test_import_xlsx.py

  python scripts/generate_hubspot_test_import_xlsx.py --email zaghlolno@gmail.com
  python scripts/generate_hubspot_test_import_xlsx.py --output C:/Users/me/Downloads/hubspot_test.xlsx
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = Path.home() / "Downloads" / "hubspot_voxbulk_test_import.xlsx"


def plus_email(base: str, tag: str) -> str:
    base = base.strip().lower()
    if "@" not in base:
        raise ValueError(f"Invalid email: {base}")
    local, domain = base.split("@", 1)
    local = re.sub(r"\+.*$", "", local)
    return f"{local}+{tag}@{domain}"


def _appt_rows(email: str, *, start: datetime) -> list[dict]:
    slots = [
        ("Sara", "Patel", "appt1", "+447700900123", 2, 10, 0, "VoxBulk Test Clinic", "Hygiene visit"),
        ("Jamal", "Okafor", "appt2", "+447700900445", 3, 14, 30, "VoxBulk Test Clinic", "Implant follow-up"),
        ("Amelia", "Chen", "appt3", "+447700900332", 4, 9, 0, "VoxBulk Test Clinic", "Check-up"),
        ("Omar", "Hassan", "appt4", "+447700900991", 5, 11, 0, "VoxBulk Test Clinic", "Root canal"),
    ]
    rows = []
    for first, last, tag, phone, day_offset, hour, minute, company, service in slots:
        dt = start + timedelta(days=day_offset)
        dt = dt.replace(hour=hour, minute=minute, second=0, microsecond=0)
        rows.append(
            {
                "firstname": first,
                "lastname": last,
                "email": plus_email(email, tag),
                "phone": phone,
                "appointment_date": dt.strftime("%Y-%m-%d %H:%M"),
                "company": company,
                "service_type": service,
                "hubspot_list": "VoxBulk · Appointment test",
            }
        )
    return rows


def _wa_survey_rows(email: str) -> list[dict]:
    contacts = [
        ("Alex", "Turner", "survey1", "+447700900501"),
        ("Sam", "Patel", "survey2", "+447700900502"),
        ("Jordan", "Lee", "survey3", "+447700900503"),
        ("Riley", "Morgan", "survey4", "+447700900504"),
    ]
    return [
        {
            "firstname": first,
            "lastname": last,
            "email": plus_email(email, tag),
            "phone": phone,
            "lifecyclestage": "customer",
            "company": "VoxBulk Test",
            "hubspot_list": "VoxBulk · WA survey test",
        }
        for first, last, tag, phone in contacts
    ]


def _won_deal_contact_rows(email: str) -> list[dict]:
    return [
        {
            "firstname": "Deal",
            "lastname": "Winner One",
            "email": plus_email(email, "won1"),
            "phone": "+447700901001",
            "hubspot_list": "VoxBulk · Won deal survey test",
        },
        {
            "firstname": "Deal",
            "lastname": "Winner Two",
            "email": plus_email(email, "won2"),
            "phone": "+447700901002",
            "hubspot_list": "VoxBulk · Won deal survey test",
        },
    ]


def _won_deal_rows(email: str) -> list[dict]:
    return [
        {
            "dealname": "VoxBulk test won deal 1",
            "dealstage": "closedwon",
            "pipeline": "default",
            "amount": "5000",
            "closedate": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "associated_contact_email": plus_email(email, "won1"),
            "hubspot_list": "VoxBulk · Won deal survey test",
        },
        {
            "dealname": "VoxBulk test won deal 2",
            "dealstage": "closedwon",
            "pipeline": "default",
            "amount": "12000",
            "closedate": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "associated_contact_email": plus_email(email, "won2"),
            "hubspot_list": "VoxBulk · Won deal survey test",
        },
    ]


def _readme_rows() -> list[dict]:
    return [
        {"step": "1", "action": "HubSpot → create static lists (MANUAL): Appointments + Survey (+ optional Confirmed/Cancelled)"},
        {"step": "2", "action": "Add contacts to lists with phone; appointment list also needs appointment_date property"},
        {"step": "3", "action": "Dashboard → Integrations → connect HubSpot (reconnect once for list scopes)"},
        {"step": "4", "action": "Appointment Manager setup → pick HubSpot appointment list → Launch → Sync CRM"},
        {"step": "5", "action": "New survey → Import from HubSpot list (can be same list as appointments)"},
        {"step": "6", "action": "Optional: set Confirmed/Cancelled lists in appointment setup for write-back"},
        {"step": "7", "action": "Replace +447700… phones with real mobiles that receive WhatsApp / Telnyx test calls"},
    ]


def write_workbook(path: Path, *, email: str, start: datetime) -> None:
    import openpyxl
    from openpyxl.styles import Font

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    sheets: list[tuple[str, list[dict]]] = [
        ("README", _readme_rows()),
        (
            "appointment_contacts",
            _appt_rows(email, start=start),
        ),
        ("wa_survey_contacts", _wa_survey_rows(email)),
        ("won_deal_contacts", _won_deal_contact_rows(email)),
        ("won_deals", _won_deal_rows(email)),
    ]

    first = True
    for title, rows in sheets:
        if first:
            ws = wb.active
            ws.title = title[:31]
            first = False
        else:
            ws = wb.create_sheet(title=title[:31])
        if not rows:
            continue
        columns = list(rows[0].keys())
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(col, "") for col in columns])
        for col in ws.columns:
            width = min(52, max(len(str(cell.value or "")) for cell in col) + 2)
            ws.column_dimensions[col[0].column_letter].width = width

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate HubSpot test import Excel locally")
    parser.add_argument(
        "--email",
        default="you@gmail.com",
        help="Base inbox for plus-address test contacts (e.g. zaghlolno@gmail.com)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--start-date",
        default="",
        help="First appointment anchor date YYYY-MM-DD (default: tomorrow UTC)",
    )
    args = parser.parse_args()

    if args.start_date.strip():
        start = datetime.strptime(args.start_date.strip(), "%Y-%m-%d")
    else:
        start = datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(days=1)

    write_workbook(args.output.expanduser().resolve(), email=args.email.strip(), start=start)
    print(f"Wrote HubSpot test import workbook:")
    print(f"  {args.output.expanduser().resolve()}")
    print("")
    print("Import order in HubSpot:")
    print("  1. appointment_contacts  (needs appointment_date property)")
    print("  2. wa_survey_contacts")
    print("  3. won_deal_contacts")
    print("  4. won_deals  (adjust dealstage/pipeline labels to your portal)")
    print("")
    print("Tip: use --email your@gmail.com and replace phone column with real test numbers.")


if __name__ == "__main__":
    main()
