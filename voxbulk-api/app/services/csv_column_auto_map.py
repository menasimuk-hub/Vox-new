"""Auto-detect CSV/Excel column mapping for Apify campaign imports."""

from __future__ import annotations

import csv
import io
import re
from typing import Any


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(h or "").strip().lower()).strip("_")


# field -> preferred header aliases (exact / contains, checked in order)
_FIELD_ALIASES: list[tuple[str, list[str]]] = [
    ("email", ["email_address", "e_mail", "email", "mail", "work_email", "business_email"]),
    ("first_name", ["first_name", "firstname", "given_name", "first", "fname"]),
    ("last_name", ["last_name", "lastname", "family_name", "surname", "last", "lname"]),
    ("full_name", ["full_name", "fullname_name", "contact_name", "person_name", "name"]),
    ("job_title", ["job_title", "jobtitle", "title", "role", "position", "designation"]),
    (
        "company_name",
        [
            "company_name",
            "company",
            "organisation",
            "organization",
            "org_name",
            "org",
            "stand_name",
            "business_name",
            "employer",
        ],
    ),
    ("event_name", ["event_name", "event", "event_title", "show_name", "expo_name", "exhibition"]),
    ("sector", ["sector", "industry", "vertical", "category"]),
    ("country_code", ["country_code", "country", "nation", "location", "region"]),
    ("promo_code", ["promo_code", "promo", "coupon", "voucher_code", "discount_code"]),
]


def auto_map_headers(headers: list[str]) -> dict[str, str]:
    """Map logical fields → original header labels. Skips ambiguous 'name' if first/last already set."""
    cleaned = [str(h or "").strip() for h in headers if str(h or "").strip()]
    norms = {h: _norm_header(h) for h in cleaned}
    mapping: dict[str, str] = {}

    def claim(field: str, header: str) -> None:
        if field not in mapping:
            mapping[field] = header

    # Pass 1: exact alias match
    for field, aliases in _FIELD_ALIASES:
        for h in cleaned:
            n = norms[h]
            if n in aliases:
                claim(field, h)
                break

    # Pass 2: contains match (longer aliases first)
    for field, aliases in _FIELD_ALIASES:
        if field in mapping:
            continue
        ordered = sorted(aliases, key=len, reverse=True)
        for h in cleaned:
            n = norms[h]
            if any(a == n or (len(a) >= 3 and a in n) for a in ordered):
                # Avoid mapping generic "name" to company when company already mapped
                if field == "full_name" and n in {"company", "company_name", "organisation", "organization"}:
                    continue
                if field == "company_name" and n in {"full_name", "contact_name", "person_name"}:
                    continue
                claim(field, h)
                break

    # If we have first+last, drop full_name to avoid double-use of same column
    if mapping.get("first_name") and mapping.get("last_name"):
        mapping.pop("full_name", None)
    # Don't use the same header for email and something else
    email_h = mapping.get("email")
    if email_h:
        for k, v in list(mapping.items()):
            if k != "email" and v == email_h:
                del mapping[k]
    return mapping


def split_full_name(full: str) -> tuple[str, str]:
    parts = [p for p in str(full or "").strip().split() if p]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def rows_from_mapping(
    reader_rows: list[dict[str, Any]],
    mapping: dict[str, str],
) -> list[dict[str, Any]]:
    """Apply mapping to raw dict rows → normalised contact dicts."""
    email_col = str(mapping.get("email") or "").strip()
    if not email_col:
        return []

    def get(row: dict[str, Any], field: str) -> str:
        col = str(mapping.get(field) or "").strip()
        if not col:
            return ""
        return str(row.get(col) or "").strip()

    out: list[dict[str, Any]] = []
    for row in reader_rows:
        email = str(row.get(email_col) or "").strip().lower()
        if not email or "@" not in email:
            continue
        first = get(row, "first_name")
        last = get(row, "last_name")
        if not first and not last:
            first, last = split_full_name(get(row, "full_name"))
        out.append(
            {
                "email": email,
                "first_name": first[:120],
                "last_name": last[:120],
                "company_name": (get(row, "company_name") or get(row, "company"))[:255],
                "event_name": (get(row, "event_name") or get(row, "event-name"))[:255],
                "job_title": get(row, "job_title")[:255],
                "sector": get(row, "sector").lower()[:64],
                "country_code": (get(row, "country_code") or get(row, "country") or "GB").upper()[:8] or "GB",
                "promo_code": get(row, "promo_code")[:64],
            }
        )
    return out


def parse_tabular_bytes(raw: bytes, filename: str = "") -> tuple[list[str], list[dict[str, str]]]:
    """Parse CSV or Excel (.xlsx) into headers + row dicts."""
    name = str(filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm") or (
        raw[:4] == b"PK\x03\x04" and b"xl/" in raw[:2000]
    ):
        return _parse_xlsx(raw)
    from app.utils.text_decoding import decode_uploaded_text

    text = decode_uploaded_text(raw)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], []
    headers = [str(h or "").strip() for h in reader.fieldnames if str(h or "").strip()]
    rows: list[dict[str, str]] = []
    for row in reader:
        rows.append({str(k or "").strip(): str(v or "").strip() for k, v in row.items() if str(k or "").strip()})
    return headers, rows


def _parse_xlsx(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel support requires openpyxl") from exc
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(c or "").strip() for c in header_row if str(c or "").strip()]
    if not headers:
        return [], []
    # Keep alignment with original columns (including blanks in middle)
    full_headers = [str(c or "").strip() or f"column_{i+1}" for i, c in enumerate(header_row)]
    out: list[dict[str, str]] = []
    for cells in rows_iter:
        d = {}
        empty = True
        for i, h in enumerate(full_headers):
            val = ""
            if cells and i < len(cells) and cells[i] is not None:
                val = str(cells[i]).strip()
            if val:
                empty = False
            d[h] = val
        if not empty:
            out.append(d)
    # Prefer non-placeholder headers for mapping
    usable = [h for h in full_headers if not h.startswith("column_")]
    return usable or full_headers, out
