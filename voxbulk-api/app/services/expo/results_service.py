"""Expo lead results — customer dashboard summary/leads/export, and admin platform overview."""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.expo import ExpoBooth, ExpoExhibition, ExpoLead, ExpoResponse, ExpoSession, ExpoVoiceNoteJob

_LANGUAGE_LABELS: dict[str, str] = {
    "ar": "Arabic",
    "en": "English",
    "tr": "Turkish",
    "fa": "Farsi",
    "ur": "Urdu",
    "fr": "French",
    "de": "German",
    "es": "Spanish",
    "it": "Italian",
    "pt": "Portuguese",
    "ru": "Russian",
    "zh": "Chinese",
    "hi": "Hindi",
    "nl": "Dutch",
    "pl": "Polish",
    "el": "Greek",
    "he": "Hebrew",
    "ja": "Japanese",
    "ko": "Korean",
    "id": "Indonesian",
    "ms": "Malay",
    "th": "Thai",
    "vi": "Vietnamese",
    "ro": "Romanian",
    "uk": "Ukrainian",
    "sv": "Swedish",
    "no": "Norwegian",
    "da": "Danish",
    "fi": "Finnish",
    "cs": "Czech",
    "hu": "Hungarian",
    "bg": "Bulgarian",
    "sk": "Slovak",
    "sr": "Serbian",
    "hr": "Croatian",
    "bn": "Bengali",
    "ta": "Tamil",
    "sw": "Swahili",
    "am": "Amharic",
    "fil": "Filipino",
    "az": "Azerbaijani",
    "ku": "Kurdish",
    "ps": "Pashto",
}


def detected_language_label(code: str | None) -> str | None:
    """Friendly display name for a stored language code (falls back to the raw code)."""
    clean = str(code or "").strip().lower()
    if not clean:
        return None
    base = clean.split("-")[0]
    return _LANGUAGE_LABELS.get(clean) or _LANGUAGE_LABELS.get(base) or clean.upper()


_EMPTY_SUMMARY: dict[str, Any] = {
    "ok": True,
    "scans": 0,
    "scans_today": 0,
    "scans_yesterday": 0,
    "sessions_started": 0,
    "sessions_today": 0,
    "completed_leads": 0,
    "leads_today": 0,
    "leads_yesterday": 0,
    "hot": 0,
    "warm": 0,
    "cold": 0,
    "offers_sent": 0,
    "emails_sent": 0,
    "catalogue_requested": 0,
    "assets_opened": 0,
    "booths_total": 0,
    "booths_live": 0,
    "daily": [],
}


def _london_day_window(*, days_ago: int = 0, now: datetime | None = None) -> tuple[datetime, datetime]:
    """Return naive-UTC [start, end) for a London calendar day (0 = today)."""
    from datetime import time, timedelta
    from zoneinfo import ZoneInfo

    london = ZoneInfo("Europe/London")
    utc = ZoneInfo("UTC")
    stamp = now or datetime.now(utc)
    if stamp.tzinfo is None:
        stamp = stamp.replace(tzinfo=utc)
    local = stamp.astimezone(london) - timedelta(days=max(0, int(days_ago)))
    start_local = datetime.combine(local.date(), time.min, tzinfo=london)
    end_local = start_local + timedelta(days=1)
    start = start_local.astimezone(utc).replace(tzinfo=None)
    end = end_local.astimezone(utc).replace(tzinfo=None)
    return start, end


class ExpoResultsService:
    @staticmethod
    def _booth_ids_for_org(
        db: Session,
        org_id: str,
        *,
        booth_id: str | None = None,
        created_by_user_id: str | None = None,
    ) -> list[str]:
        q = select(ExpoBooth.id).where(ExpoBooth.org_id == org_id)
        if booth_id:
            q = q.where(ExpoBooth.id == booth_id)
        if created_by_user_id:
            q = q.where(ExpoBooth.created_by_user_id == created_by_user_id)
        return [row[0] for row in db.execute(q).all()]

    @staticmethod
    def customer_summary(
        db: Session,
        org_id: str,
        *,
        booth_id: str | None = None,
        created_by_user_id: str | None = None,
    ) -> dict[str, Any]:
        from app.services.expo.booth_service import booth_is_expired

        booth_ids = ExpoResultsService._booth_ids_for_org(
            db, org_id, booth_id=booth_id, created_by_user_id=created_by_user_id
        )
        if not booth_ids:
            return dict(_EMPTY_SUMMARY)

        booths = db.execute(select(ExpoBooth).where(ExpoBooth.id.in_(booth_ids))).scalars().all()
        scans = int(sum(int(b.scan_count or 0) for b in booths))
        booths_live = sum(1 for b in booths if not booth_is_expired(b) and str(b.status or "").lower() == "active")

        sessions_started = int(
            db.execute(
                select(func.count()).select_from(ExpoSession).where(ExpoSession.booth_id.in_(booth_ids))
            ).scalar()
            or 0
        )
        leads = db.execute(select(ExpoLead).where(ExpoLead.booth_id.in_(booth_ids))).scalars().all()

        today_start, today_end = _london_day_window(days_ago=0)
        yday_start, yday_end = _london_day_window(days_ago=1)

        sessions_today = int(
            db.execute(
                select(func.count())
                .select_from(ExpoSession)
                .where(
                    ExpoSession.booth_id.in_(booth_ids),
                    ExpoSession.started_at >= today_start,
                    ExpoSession.started_at < today_end,
                )
            ).scalar()
            or 0
        )
        sessions_yesterday = int(
            db.execute(
                select(func.count())
                .select_from(ExpoSession)
                .where(
                    ExpoSession.booth_id.in_(booth_ids),
                    ExpoSession.started_at >= yday_start,
                    ExpoSession.started_at < yday_end,
                )
            ).scalar()
            or 0
        )
        leads_today = sum(
            1
            for lead in leads
            if lead.created_at and today_start <= lead.created_at < today_end
        )
        leads_yesterday = sum(
            1
            for lead in leads
            if lead.created_at and yday_start <= lead.created_at < yday_end
        )

        # Prefer session starts as "scans today" (each visitor chat open); fall back stays consistent with total.
        scans_today = sessions_today
        scans_yesterday = sessions_yesterday

        daily: list[dict[str, Any]] = []
        for ago in range(6, -1, -1):
            start, end = _london_day_window(days_ago=ago)
            day_sessions = int(
                db.execute(
                    select(func.count())
                    .select_from(ExpoSession)
                    .where(
                        ExpoSession.booth_id.in_(booth_ids),
                        ExpoSession.started_at >= start,
                        ExpoSession.started_at < end,
                    )
                ).scalar()
                or 0
            )
            day_leads = sum(
                1 for lead in leads if lead.created_at and start <= lead.created_at < end
            )
            day_hot = sum(
                1
                for lead in leads
                if lead.lead_score == "hot" and lead.created_at and start <= lead.created_at < end
            )
            from zoneinfo import ZoneInfo

            london = ZoneInfo("Europe/London")
            utc = ZoneInfo("UTC")
            label = start.replace(tzinfo=utc).astimezone(london).strftime("%a")
            daily.append(
                {
                    "day": label,
                    "scans": day_sessions,
                    "leads": day_leads,
                    "hot": day_hot,
                }
            )

        return {
            "ok": True,
            "scans": scans,
            "scans_today": scans_today,
            "scans_yesterday": scans_yesterday,
            "sessions_started": sessions_started,
            "sessions_today": sessions_today,
            "completed_leads": sum(1 for lead in leads if lead.lead_score),
            "leads_today": leads_today,
            "leads_yesterday": leads_yesterday,
            "hot": sum(1 for lead in leads if lead.lead_score == "hot"),
            "warm": sum(1 for lead in leads if lead.lead_score == "warm"),
            "cold": sum(1 for lead in leads if lead.lead_score == "cold"),
            "offers_sent": sum(1 for lead in leads if lead.offer_sent_at is not None),
            "emails_sent": sum(1 for lead in leads if lead.offer_sent_at is not None),
            "catalogue_requested": sum(1 for lead in leads if lead.consent_acknowledged),
            "assets_opened": sum(
                1
                for lead in leads
                if str(getattr(lead, "assets_opened_json", None) or "").strip()
                not in {"", "[]", "null"}
            ),
            "booths_total": len(booths),
            "booths_live": booths_live,
            "daily": daily,
        }

    @staticmethod
    def customer_leads(
        db: Session,
        org_id: str,
        *,
        booth_id: str | None = None,
        lead_id: str | None = None,
        score: str | None = None,
        catalogue_requested: bool | None = None,
        price_list_requested: bool | None = None,
        asset_opened: bool | None = None,
        created_by_user_id: str | None = None,
        limit: int = 200,
    ) -> list[dict[str, Any]]:
        booth_ids = ExpoResultsService._booth_ids_for_org(
            db, org_id, booth_id=booth_id, created_by_user_id=created_by_user_id
        )
        if not booth_ids:
            return []

        q = select(ExpoLead).where(ExpoLead.booth_id.in_(booth_ids), ExpoLead.org_id == org_id)
        if lead_id:
            q = q.where(ExpoLead.id == str(lead_id).strip())
        if score:
            q = q.where(ExpoLead.lead_score == str(score).strip().lower())
        if catalogue_requested is True:
            q = q.where(ExpoLead.consent_acknowledged.is_(True))
        elif catalogue_requested is False:
            q = q.where(ExpoLead.consent_acknowledged.is_(False))
        capped = max(1, min(int(limit or 200), 5000))
        rows = db.execute(q.order_by(ExpoLead.created_at.desc()).limit(capped)).scalars().all()

        # Drop leads from preview sessions (wizard / ?preview=1)
        session_ids = [r.session_id for r in rows if r.session_id]
        preview_session_ids: set[str] = set()
        if session_ids:
            preview_session_ids = {
                s.id
                for s in db.execute(
                    select(ExpoSession).where(
                        ExpoSession.id.in_(session_ids),
                        ExpoSession.is_preview.is_(True),
                    )
                ).scalars().all()
            }
        rows = [r for r in rows if not r.session_id or r.session_id not in preview_session_ids]

        booths = {
            b.id: b for b in db.execute(select(ExpoBooth).where(ExpoBooth.id.in_(booth_ids))).scalars().all()
        }
        items = [ExpoResultsService._lead_to_dict(lead, booths.get(lead.booth_id)) for lead in rows]
        if price_list_requested is True:
            items = [i for i in items if i.get("price_list_requested")]
        elif price_list_requested is False:
            items = [i for i in items if not i.get("price_list_requested")]
        if asset_opened is True:
            items = [i for i in items if int(i.get("assets_opened_count") or 0) > 0]
        elif asset_opened is False:
            items = [i for i in items if int(i.get("assets_opened_count") or 0) == 0]
        return items

    @staticmethod
    def delete_lead(
        db: Session,
        org_id: str,
        *,
        lead_id: str,
        created_by_user_id: str | None = None,
    ) -> None:
        lead = db.execute(
            select(ExpoLead).where(ExpoLead.id == lead_id, ExpoLead.org_id == org_id)
        ).scalar_one_or_none()
        if lead is None:
            raise ValueError("Lead not found")
        booth = db.get(ExpoBooth, lead.booth_id)
        if booth is None or booth.org_id != org_id:
            raise ValueError("Lead not found")
        if created_by_user_id and booth.created_by_user_id != created_by_user_id:
            raise ValueError("Lead not found")
        session_id = lead.session_id
        db.delete(lead)
        if session_id:
            for resp in db.execute(select(ExpoResponse).where(ExpoResponse.session_id == session_id)).scalars().all():
                db.delete(resp)
            for job in db.execute(
                select(ExpoVoiceNoteJob).where(ExpoVoiceNoteJob.session_id == session_id)
            ).scalars().all():
                db.delete(job)
            session = db.get(ExpoSession, session_id)
            if session is not None:
                db.delete(session)
        db.commit()

    @staticmethod
    def lead_detail(
        db: Session,
        org_id: str,
        *,
        lead_id: str,
        created_by_user_id: str | None = None,
    ) -> dict[str, Any]:
        lead = db.execute(
            select(ExpoLead).where(ExpoLead.id == lead_id, ExpoLead.org_id == org_id)
        ).scalar_one_or_none()
        if lead is None:
            raise ValueError("Lead not found")
        booth = db.get(ExpoBooth, lead.booth_id)
        if booth is None or booth.org_id != org_id:
            raise ValueError("Lead not found")
        if created_by_user_id and booth.created_by_user_id != created_by_user_id:
            raise ValueError("Lead not found")
        data = ExpoResultsService._lead_to_dict(lead, booth)
        answers: list[dict[str, Any]] = []
        if lead.session_id:
            rows = (
                db.execute(
                    select(ExpoResponse)
                    .where(ExpoResponse.session_id == lead.session_id)
                    .order_by(ExpoResponse.step_order.asc(), ExpoResponse.created_at.asc())
                )
                .scalars()
                .all()
            )
            from app.services.expo.question_bank import SELECTABLE_QUESTION_BANK, parse_question_config

            labels = {str(q["key"]): str(q.get("label") or q["key"]) for q in SELECTABLE_QUESTION_BANK}
            prompts = {str(q["key"]): str(q.get("prompt") or "") for q in SELECTABLE_QUESTION_BANK}
            labels.update(
                {
                    "contact": "Contact",
                    "name": "Name",
                    "company": "Company",
                    "mobile": "Mobile",
                    "business_card": "Business card",
                    "consent": "Consent",
                }
            )
            for step in parse_question_config(booth.question_config_json):
                key = str(step.get("key") or "")
                if key and step.get("prompt"):
                    prompts[key] = str(step.get("prompt"))
                if key and step.get("label"):
                    labels[key] = str(step.get("label"))
            for r in rows:
                key = str(r.question_key or "")
                answers.append(
                    {
                        "question_key": key,
                        "question_label": labels.get(key, key.replace("_", " ").title()),
                        "question_prompt": prompts.get(key) or labels.get(key) or key,
                        "answer_text": r.answer_text,
                        "original_text": r.original_text,
                        "answer_text_en": r.answer_text_en,
                        "answer_source": r.answer_source,
                        "step_order": r.step_order,
                    }
                )
        data["answers"] = answers
        return data

    @staticmethod
    def resolve_lead_card_path(
        db: Session,
        org_id: str,
        *,
        lead_id: str,
        created_by_user_id: str | None = None,
    ):
        from pathlib import Path

        lead = db.execute(
            select(ExpoLead).where(ExpoLead.id == lead_id, ExpoLead.org_id == org_id)
        ).scalar_one_or_none()
        if lead is None or not lead.business_card_path:
            return None
        booth = db.get(ExpoBooth, lead.booth_id)
        if booth is None or booth.org_id != org_id:
            return None
        if created_by_user_id and booth.created_by_user_id != created_by_user_id:
            return None
        rel = str(lead.business_card_path).strip().replace("\\", "/")
        if not rel or ".." in rel.split("/"):
            return None
        root = Path(__file__).resolve().parents[3]
        abs_path = (root / rel).resolve()
        cards_root = (root / "data" / "expo-cards").resolve()
        try:
            abs_path.relative_to(cards_root)
        except ValueError:
            return None
        return abs_path if abs_path.is_file() else None

    @staticmethod
    def _lead_to_dict(lead: ExpoLead, booth: ExpoBooth | None) -> dict[str, Any]:
        from app.services.expo.offer_delivery_service import normalize_asset_purpose

        try:
            assets_sent = json.loads(lead.assets_sent_json or "[]")
            if not isinstance(assets_sent, list):
                assets_sent = []
        except (json.JSONDecodeError, TypeError):
            assets_sent = []
        try:
            assets_opened = json.loads(getattr(lead, "assets_opened_json", None) or "[]")
            if not isinstance(assets_opened, list):
                assets_opened = []
        except (json.JSONDecodeError, TypeError):
            assets_opened = []
        purposes: set[str] = set()
        for item in assets_sent:
            if isinstance(item, dict):
                purposes.add(normalize_asset_purpose(item.get("purpose")))
        catalogue_requested = bool(lead.consent_acknowledged) or "catalogue" in purposes
        price_list_requested = bool(lead.consent_acknowledged) or "price_list" in purposes
        return {
            "id": lead.id,
            "booth_id": lead.booth_id,
            "booth_name": booth.name if booth else None,
            "booth_code": booth.booth_code if booth else None,
            "detected_language": lead.detected_language,
            "detected_language_label": detected_language_label(lead.detected_language),
            "country_hint": lead.country_hint,
            "name": lead.name,
            "company": lead.company,
            "visitor_phone": lead.visitor_phone,
            "visitor_email": lead.visitor_email,
            "business_card_path": lead.business_card_path,
            "business_card_url": (
                f"/expo/results/leads/{lead.id}/card-image" if lead.business_card_path else None
            ),
            "interest": lead.interest,
            "buying_timeline": lead.buying_timeline,
            "lead_score": lead.lead_score,
            "consent_acknowledged": lead.consent_acknowledged,
            "catalogue_requested": catalogue_requested,
            "price_list_requested": price_list_requested,
            "offer_sent_at": lead.offer_sent_at.isoformat() if lead.offer_sent_at else None,
            "email_sent_at": lead.offer_sent_at.isoformat() if lead.offer_sent_at else None,
            "offer_interested": bool(getattr(lead, "offer_interested", False)),
            "assets_sent": assets_sent,
            "assets_opened": assets_opened,
            "assets_opened_count": len(assets_opened),
            "follow_up_status": lead.follow_up_status,
            "created_at": lead.created_at.isoformat() if lead.created_at else None,
        }

    @staticmethod
    def export_csv(
        db: Session,
        org_id: str,
        *,
        booth_id: str | None = None,
        lead_id: str | None = None,
        created_by_user_id: str | None = None,
    ) -> str:
        leads = ExpoResultsService.customer_leads(
            db,
            org_id,
            booth_id=booth_id,
            lead_id=lead_id,
            created_by_user_id=created_by_user_id,
            limit=5000,
        )
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(
            [
                "Name",
                "Company",
                "Phone",
                "Email",
                "Interest",
                "Timeline",
                "Score",
                "Consent",
                "Offer interested",
                "Booth",
                "Assets sent",
                "Created at",
            ]
        )
        for lead in leads:
            writer.writerow(
                [
                    lead.get("name") or "",
                    lead.get("company") or "",
                    lead.get("visitor_phone") or "",
                    lead.get("visitor_email") or "",
                    lead.get("interest") or "",
                    lead.get("buying_timeline") or "",
                    lead.get("lead_score") or "",
                    "Yes" if lead.get("consent_acknowledged") else "No",
                    "Yes" if lead.get("offer_interested") else "No",
                    lead.get("booth_name") or "",
                    ", ".join(
                        a.get("title") if isinstance(a, dict) else str(a)
                        for a in (lead.get("assets_sent") or [])
                    ),
                    lead.get("created_at") or "",
                ]
            )
        return buf.getvalue()

    @staticmethod
    def export_xlsx(
        db: Session,
        org_id: str,
        *,
        booth_id: str | None = None,
        lead_id: str | None = None,
        created_by_user_id: str | None = None,
    ) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font
        except Exception as e:
            raise RuntimeError("Excel export requires openpyxl on the server.") from e

        leads = ExpoResultsService.customer_leads(
            db,
            org_id,
            booth_id=booth_id,
            lead_id=lead_id,
            created_by_user_id=created_by_user_id,
            limit=5000,
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expo leads"
        headers = [
            "Name",
            "Company",
            "Phone",
            "Email",
            "Interest",
            "Timeline",
            "Score",
            "Consent",
            "Offer interested",
            "Booth",
            "Assets sent",
            "Created at",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for lead in leads:
            phone = str(lead.get("visitor_phone") or "")
            email = str(lead.get("visitor_email") or "")
            if phone.startswith("web-pending-") or phone.startswith("web-card-"):
                phone = ""
            if email.endswith("@expo.local"):
                email = ""
            interest = str(lead.get("interest") or "")
            if interest == "[Translation unavailable]":
                interest = ""
            assets = lead.get("assets_sent") or []
            assets_txt = ", ".join(
                a.get("title") if isinstance(a, dict) else str(a) for a in assets
            )
            ws.append(
                [
                    lead.get("name") or "",
                    lead.get("company") or "",
                    phone,
                    email,
                    interest,
                    lead.get("buying_timeline") or "",
                    lead.get("lead_score") or "",
                    "Yes" if lead.get("consent_acknowledged") else "No",
                    "Yes" if lead.get("offer_interested") else "No",
                    lead.get("booth_name") or "",
                    assets_txt,
                    lead.get("created_at") or "",
                ]
            )
        # Reasonable column widths
        from openpyxl.utils import get_column_letter

        widths = [22, 22, 18, 28, 36, 16, 10, 10, 14, 22, 28, 20]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def admin_overview(db: Session) -> dict[str, Any]:
        booths_total = int(db.execute(select(func.count()).select_from(ExpoBooth)).scalar() or 0)
        exhibitions_total = int(db.execute(select(func.count()).select_from(ExpoExhibition)).scalar() or 0)
        orgs_with_expo = int(
            db.execute(select(func.count(func.distinct(ExpoBooth.org_id)))).scalar() or 0
        )
        scans = int(db.execute(select(func.coalesce(func.sum(ExpoBooth.scan_count), 0))).scalar() or 0)
        leads_total = int(db.execute(select(func.count()).select_from(ExpoLead)).scalar() or 0)
        hot = int(
            db.execute(select(func.count()).select_from(ExpoLead).where(ExpoLead.lead_score == "hot")).scalar()
            or 0
        )
        warm = int(
            db.execute(select(func.count()).select_from(ExpoLead).where(ExpoLead.lead_score == "warm")).scalar()
            or 0
        )
        cold = int(
            db.execute(select(func.count()).select_from(ExpoLead).where(ExpoLead.lead_score == "cold")).scalar()
            or 0
        )
        return {
            "ok": True,
            "booths_total": booths_total,
            "exhibitions_total": exhibitions_total,
            "orgs_with_expo": orgs_with_expo,
            "scans": scans,
            "leads_total": leads_total,
            "hot": hot,
            "warm": warm,
            "cold": cold,
        }
