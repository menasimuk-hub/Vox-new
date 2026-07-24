"""Merge CSV + CV uploads into interview draft recipient lists."""

from __future__ import annotations

import json
import logging
import re
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)

from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.service_order import ServiceOrder, ServiceOrderRecipient
from app.services.interview_cv_parse_service import ParsedCv, name_from_filename, name_similarity, parse_uploaded_cv_files
from app.services.platform_catalog_service import ServiceOrderService

MATCH_THRESHOLD = 0.82

# Longer Arabic header keys first so substring match prefers specific labels.
_ARABIC_HEADER_ALIASES: list[tuple[str, str]] = [
    ("الاسمالكامل", "name"),
    ("اسمالمرشح", "name"),
    ("اسمالشخص", "name"),
    ("رقمالواتساب", "phone"),
    ("رقمالجوال", "phone"),
    ("رقمالهاتف", "phone"),
    ("رقمالموبايل", "phone"),
    ("البريدالإلكتروني", "email"),
    ("البريدالالكتروني", "email"),
    ("البريدالالكترونى", "email"),
    ("الاسم", "name"),
    ("المرشح", "name"),
    ("اسم", "name"),
    ("الجوال", "phone"),
    ("الهاتف", "phone"),
    ("الموبايل", "phone"),
    ("موبايل", "phone"),
    ("جوال", "phone"),
    ("هاتف", "phone"),
    ("واتساب", "phone"),
    ("البريد", "email"),
    ("ايميل", "email"),
    ("إيميل", "email"),
    ("بريد", "email"),
]


def _arabic_header_compact(value: str) -> str:
    """Strip spaces/punctuation so Arabic Excel headers can be matched."""
    text = str(value or "").strip().replace("\u0640", "")  # tatweel
    return re.sub(r"[\s\W_]+", "", text, flags=re.UNICODE)


def _canonical_contact_header(header: str) -> str:
    """Map Excel/CSV header to name|phone|email (English + Arabic)."""
    raw = str(header or "").strip()
    if not raw:
        return ""
    compact = _arabic_header_compact(raw)
    for key, field in _ARABIC_HEADER_ALIASES:
        if compact == key or (key and key in compact):
            return field
    return ServiceOrderService._norm_header(raw)


def _contact_fields_from_row(
    headers: list[str],
    cells: list[str],
) -> tuple[str, str, str]:
    """Resolve name/phone/email from headers, with positional fallback."""
    data: dict[str, str] = {}
    for i, header in enumerate(headers):
        if not header or i >= len(cells):
            continue
        # Prefer first non-empty value when duplicate headers map to the same field.
        if header not in data or not data[header]:
            data[header] = cells[i]
    name, phone, email, _lang = ServiceOrderService._row_field_values(data)
    if name or phone:
        return name, phone, email
    # No recognized headers (e.g. all Arabic wiped previously) — template column order.
    if len(cells) >= 2 and (cells[0].strip() or cells[1].strip()):
        return (
            cells[0].strip(),
            cells[1].strip(),
            cells[2].strip() if len(cells) > 2 else "",
        )
    return name, phone, email


def _norm_phone(value: str | None) -> str:
    return re.sub(r"\D", "", str(value or ""))



def _norm_email(value: str | None) -> str:
    return str(value or "").strip().lower()


def _norm_name(value: str | None) -> str:
    return str(value or "").strip().lower()


def find_duplicate_recipient(
    recipients: list[ServiceOrderRecipient],
    *,
    name: str | None,
    phone: str | None,
    email: str | None,
) -> ServiceOrderRecipient | None:
    """Match a prior row for the same person (not phone-only — avoids merging ZIP batch CVs)."""
    np = _norm_phone(phone)
    ne = _norm_email(email)
    nn = _norm_name(name)
    for r in recipients:
        if nn and len(nn) >= 2 and _norm_name(r.name) == nn:
            return r
        if ne and _norm_email(r.email) == ne:
            return r
        if np and _norm_phone(r.phone) == np:
            if nn and r.name and name_similarity(r.name, name) >= MATCH_THRESHOLD:
                return r
            if ne and _norm_email(r.email) == ne:
                return r
    return None


def _remove_recipient(db: Session, recipient: ServiceOrderRecipient, recipients: list[ServiceOrderRecipient]) -> None:
    from sqlalchemy import inspect

    from app.services.career_cv_storage_service import delete_cv_file

    delete_cv_file(recipient.cv_storage_key)
    state = inspect(recipient)
    if state.persistent:
        db.delete(recipient)
    if recipient in recipients:
        recipients.remove(recipient)


def _loads_json(raw: str | None) -> Any:
    try:
        return json.loads(raw or "")
    except Exception:
        return None


def _dumps_json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False)


def _phone_intake_error(phone: str | None) -> str | None:
    """Return a user-facing phone error, or None when missing/valid."""
    raw = str(phone or "").strip()
    if not raw:
        return "Phone missing — click to add"
    from app.services.recipient_contact_validation import coerce_interview_phone_e164

    _, err = coerce_interview_phone_e164(raw)
    return err


def _coerce_contact_phone(raw: str | None) -> tuple[str | None, list[str]]:
    """Normalize contact phone for storage; return (phone, intake_errors)."""
    phone = str(raw or "").strip() or None
    if not phone:
        return None, ["Phone missing — click to add"]
    from app.services.recipient_contact_validation import coerce_interview_phone_e164

    e164, err = coerce_interview_phone_e164(phone)
    if err:
        return phone, [err]
    return e164 or phone, []


def compute_intake_errors(recipient: ServiceOrderRecipient) -> list[str]:
    errors: list[str] = []
    stored = _loads_json(recipient.intake_errors_json)
    if isinstance(stored, list):
        # Drop stale phone messages — recompute from current phone value.
        for x in stored:
            s = str(x)
            if not s:
                continue
            low = s.lower()
            if "phone missing" in low or "e.164" in low or "valid phone" in low or "phone number" in low:
                continue
            if "phone can only" in low or "phone is too long" in low:
                continue
            errors.append(s)
    if not str(recipient.name or "").strip():
        errors.append("Name missing")
    phone_err = _phone_intake_error(recipient.phone)
    if phone_err:
        errors.append(phone_err)
    quality = str(recipient.cv_quality or "missing")
    if quality == "low_quality":
        errors.append("CV low-quality — generic questions only")
    elif quality == "corrupt":
        errors.append("CV unreadable")
    elif quality == "missing" and str(recipient.intake_source or "") not in {
        "csv",
        "zoho_recruit",
        "merged",
    }:
        errors.append("CV missing")
    # de-dupe preserve order
    seen: set[str] = set()
    out: list[str] = []
    for err in errors:
        key = err.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(err)
    return out


def recipient_intake_dict(
    recipient: ServiceOrderRecipient,
    *,
    position: str = "",
    booking_token: Any | None = None,
    order: ServiceOrder | None = None,
) -> dict[str, Any]:
    base = ServiceOrderService.recipient_to_dict(recipient)
    result_parsed = _loads_json(recipient.result_json) or {}
    if not isinstance(result_parsed, dict):
        result_parsed = {}
    if booking_token is not None:
        result_parsed = dict(result_parsed)
        if getattr(booking_token, "booked_start_at", None) and not result_parsed.get("booking_cancelled_at"):
            result_parsed["booked_start_at"] = booking_token.booked_start_at.isoformat()
            if getattr(booking_token, "booked_end_at", None):
                result_parsed["booked_end_at"] = booking_token.booked_end_at.isoformat()
            if getattr(booking_token, "updated_at", None):
                result_parsed.setdefault("booking_confirmed_at", booking_token.updated_at.isoformat())
        if getattr(booking_token, "wa_sent_at", None):
            result_parsed.setdefault("invite_wa_sent_at", booking_token.wa_sent_at.isoformat())
    cv_parsed = _loads_json(recipient.cv_parsed_json) or {}
    if not isinstance(cv_parsed, dict):
        cv_parsed = {}
    errors = compute_intake_errors(recipient)
    ready = bool(str(recipient.name or "").strip() and str(recipient.phone or "").strip())
    from app.services.interview_ats_service import ats_display_for_recipient

    base.update(
        {
            "cv_quality": recipient.cv_quality or "missing",
            "cv_filename": recipient.cv_filename,
            "intake_source": recipient.intake_source,
            "intake_errors": errors,
            "intake_ready": ready,
            "cv_skills": cv_parsed.get("skills") or [],
            "cv_job_titles": cv_parsed.get("job_titles") or [],
            "has_cv_file": bool(recipient.cv_storage_key or (recipient.cv_text or "").strip()),
        }
    )
    base.update(ats_display_for_recipient(recipient, position=position))
    from app.services.interview_activity_service import InterviewActivityService

    base["activity_status"] = InterviewActivityService.activity_status(
        recipient, parsed=result_parsed, order=order
    )
    if result_parsed.get("exclusion_label") and base["activity_status"] == "auto_excluded":
        base["activity_status_label"] = str(result_parsed.get("exclusion_label"))
    base["booked_start_at"] = result_parsed.get("booked_start_at")
    base["booked_end_at"] = result_parsed.get("booked_end_at")
    phone_raw = str(recipient.phone or "").strip()
    if phone_raw:
        from sqlalchemy.orm import object_session

        from app.services.telnyx_phone_allowlist_service import TelnyxPhoneAllowlistService

        db = object_session(recipient)
        phone_check = (
            TelnyxPhoneAllowlistService.validate_phone_db(db, phone_raw)
            if db is not None
            else TelnyxPhoneAllowlistService.validate_phone(phone_raw)
        )
        base["phone_call_allowed"] = bool(phone_check.get("allowed"))
        base["phone_call_block_reason"] = phone_check.get("reason")
        base["phone_country"] = phone_check.get("country")
        base["phone_line_type"] = phone_check.get("line_type")
        if phone_check.get("normalized"):
            base["phone_normalized"] = phone_check.get("normalized")
    else:
        base["phone_call_allowed"] = False
        base["phone_call_block_reason"] = "Phone number is required"
    tok_str = str(result_parsed.get("booking_token") or "").strip()
    url_str = str(result_parsed.get("booking_url") or "").strip()
    if booking_token is not None:
        tok_str = str(getattr(booking_token, "token", "") or tok_str).strip()
    if tok_str and not url_str:
        from app.services.interview_booking_service import booking_url_for_token

        url_str = booking_url_for_token(tok_str)
    if tok_str:
        base["booking_token"] = tok_str
    if url_str:
        base["booking_url"] = url_str
    from app.services.interview_booking_service import _recipient_outreach_email

    outreach = _recipient_outreach_email(recipient)
    if outreach:
        base["outreach_email"] = outreach
    base["invite_email_failed"] = result_parsed.get("invite_email_failed")
    base["invite_email_sent_at"] = result_parsed.get("invite_email_sent_at")
    return base


def _assert_interview_email_intake(order: ServiceOrder) -> None:
    if order.service_code != "interview":
        raise ValueError("Only interview orders support CV email intake")
    if order.status in {"completed", "cancelled"}:
        raise ValueError("Cannot add CVs to a finished or cancelled campaign")


def _assert_interview_draft(order: ServiceOrder) -> None:
    if order.service_code != "interview":
        raise ValueError("Only interview orders support CV intake")
    if order.payment_status == "approved":
        raise ValueError("Cannot change candidates after payment is approved")
    if order.status in {"running", "completed", "cancelled"}:
        raise ValueError("Cannot change candidates while campaign is active or finished")


def _order_config(order: ServiceOrder) -> dict[str, Any]:
    parsed = _loads_json(order.config_json)
    return parsed if isinstance(parsed, dict) else {}


def _booking_invites_sent(order: ServiceOrder) -> bool:
    cfg = _order_config(order)
    if not cfg.get("booking_invites_sent_at"):
        return False
    dispatch = cfg.get("last_invite_dispatch")
    if isinstance(dispatch, dict) and dispatch.get("ok") is False:
        return False
    return True


def _assert_can_delete_intake_recipient(order: ServiceOrder, recipient: ServiceOrderRecipient) -> None:
    if order.service_code != "interview":
        raise ValueError("Only interview orders support CV intake")
    if order.status in {"running", "completed", "cancelled"}:
        raise ValueError("Cannot remove candidates while the campaign is active or finished")
    if _booking_invites_sent(order):
        raise ValueError("Cannot remove candidates after booking invites have been sent")
    terminal = {"completed", "done", "calling", "in_progress", "ringing"}
    if str(recipient.status or "").strip().lower() in terminal:
        raise ValueError("Cannot remove a candidate who has started or completed their interview")


def get_latest_interview_draft(db: Session, *, org_id: str) -> ServiceOrder | None:
    """Return the most recent in-progress draft that has saved content (do not purge here)."""
    from sqlalchemy import func

    rows = list(
        db.execute(
            select(ServiceOrder)
            .where(
                ServiceOrder.org_id == org_id,
                ServiceOrder.service_code == "interview",
                ServiceOrder.status == "draft",
                ServiceOrder.payment_status == "unpaid",
            )
            .order_by(ServiceOrder.updated_at.desc())
        ).scalars()
    )
    for order in rows:
        count = db.execute(
            select(func.count())
            .select_from(ServiceOrderRecipient)
            .where(ServiceOrderRecipient.order_id == order.id)
        ).scalar_one()
        if not is_empty_interview_draft(order, recipient_count=int(count or 0)):
            return order
    return None


def interview_draft_visible_in_saved_list(order: ServiceOrder, *, recipient_count: int) -> bool:
    """Unpaid interview drafts only appear in Saved after the user saves or adds candidates."""
    if order.service_code != "interview":
        return True
    if order.status != "draft" or order.payment_status != "unpaid":
        return True
    if recipient_count > 0 or int(order.recipient_count or 0) > 0:
        return True
    cfg = _loads_json(order.config_json) or {}
    if cfg.get("draft_saved_by_user"):
        return True
    return False


def _interview_draft_has_meaningful_config(cfg: dict[str, Any]) -> bool:
    meaningful = [
        cfg.get("role"),
        cfg.get("position"),
        cfg.get("criteria"),
        cfg.get("screening_criteria"),
        cfg.get("approved_script"),
        cfg.get("generated_script_draft"),
        cfg.get("agent_id"),
        cfg.get("title"),
    ]
    if any(str(v or "").strip() for v in meaningful):
        return True
    if cfg.get("script_approved") or cfg.get("cv_email_enabled"):
        return True
    if str(cfg.get("scheduled_start") or "").strip() or str(cfg.get("scheduled_end") or "").strip():
        return True
    if str(cfg.get("calling_start") or "").strip() or str(cfg.get("calling_end") or "").strip():
        return True
    delivery = str(cfg.get("delivery") or "").strip().lower()
    if delivery and delivery not in {"ai_call", ""}:
        return True
    duration = cfg.get("expected_duration_minutes")
    if duration is not None and str(duration).strip() not in {"", "0"}:
        return True
    return False


def is_empty_interview_draft(order: ServiceOrder, *, recipient_count: int) -> bool:
    if order.status != "draft" or order.payment_status != "unpaid":
        return False
    if recipient_count > 0:
        return False
    if int(order.recipient_count or 0) > 0:
        return False
    cfg = _loads_json(order.config_json) or {}
    if _interview_draft_has_meaningful_config(cfg):
        return False
    title = str(order.title or "").strip().lower()
    return title in {"", "interview draft", "interview", "untitled interview"}


def purge_empty_interview_drafts(db: Session, *, org_id: str, keep_order_id: str | None = None) -> int:
    from sqlalchemy import func

    rows = list(
        db.execute(
            select(ServiceOrder)
            .where(
                ServiceOrder.org_id == org_id,
                ServiceOrder.service_code == "interview",
                ServiceOrder.status == "draft",
                ServiceOrder.payment_status == "unpaid",
            )
        ).scalars()
    )
    deleted = 0
    for order in rows:
        if keep_order_id and order.id == keep_order_id:
            continue
        count = db.execute(
            select(func.count())
            .select_from(ServiceOrderRecipient)
            .where(ServiceOrderRecipient.order_id == order.id)
        ).scalar_one()
        if is_empty_interview_draft(order, recipient_count=int(count or 0)):
            ServiceOrderService.delete_order(db, order)
            deleted += 1
    return deleted


def ensure_interview_draft_order(
    db: Session,
    *,
    org_id: str,
    user_id: str,
    title: str = "Interview draft",
    role: str = "",
    criteria: str = "",
) -> ServiceOrder:
    purge_empty_interview_drafts(db, org_id=org_id)
    rows = list(
        db.execute(
            select(ServiceOrder)
            .where(
                ServiceOrder.org_id == org_id,
                ServiceOrder.service_code == "interview",
                ServiceOrder.status == "draft",
                ServiceOrder.payment_status == "unpaid",
            )
            .order_by(ServiceOrder.updated_at.desc())
            .limit(1)
        ).scalars()
    )
    if rows:
        order = rows[0]
        config = _loads_json(order.config_json) or {}
        changed = False
        if role and not config.get("role"):
            config["role"] = role
            changed = True
        if criteria and not config.get("criteria"):
            config["criteria"] = criteria
            changed = True
        if changed:
            order.config_json = _dumps_json(config)
            order.updated_at = datetime.utcnow()
            db.add(order)
            db.commit()
            db.refresh(order)
        from app.services.interview_reference_service import ensure_order_reference_id

        return ensure_order_reference_id(db, order)
    config: dict[str, Any] = {}
    if role:
        config["role"] = role
    if criteria:
        config["criteria"] = criteria
    order = ServiceOrderService.create_order(
        db,
        org_id=org_id,
        user_id=user_id,
        service_code="interview",
        title=title.strip() or "Interview draft",
        config=config,
    )
    from app.services.interview_reference_service import ensure_order_reference_id

    return ensure_order_reference_id(db, order)


def create_new_interview_draft(
    db: Session,
    *,
    org_id: str,
    user_id: str,
) -> ServiceOrder:
    """Always create a fresh draft (new reference ID) without touching existing drafts."""
    purge_empty_interview_drafts(db, org_id=org_id)
    order = ServiceOrderService.create_order(
        db,
        org_id=org_id,
        user_id=user_id,
        service_code="interview",
        title="Interview draft",
        config={},
    )
    from app.services.interview_reference_service import ensure_order_reference_id

    order = ensure_order_reference_id(db, order)
    purge_empty_interview_drafts(db, org_id=org_id, keep_order_id=order.id)
    return order


def abandon_empty_interview_draft(db: Session, *, org_id: str, order_id: str) -> bool:
    """Delete a draft interview order when it has no saved content."""
    order = ServiceOrderService.get_order(db, order_id, org_id=org_id)
    if order is None or order.service_code != "interview":
        return False
    from sqlalchemy import func

    count = db.execute(
        select(func.count())
        .select_from(ServiceOrderRecipient)
        .where(ServiceOrderRecipient.order_id == order.id)
    ).scalar_one()
    if not is_empty_interview_draft(order, recipient_count=int(count or 0)):
        return False
    ServiceOrderService.delete_order(db, order)
    return True


def _apply_parsed_cv(recipient: ServiceOrderRecipient, parsed: ParsedCv, *, merge: bool) -> None:
    recipient.cv_filename = parsed.filename
    recipient.cv_text = parsed.text or None
    recipient.cv_quality = parsed.quality
    recipient.cv_parsed_json = _dumps_json(parsed.to_dict())
    recipient.intake_errors_json = _dumps_json(parsed.errors)
    if parsed.name and (not recipient.name or merge):
        recipient.name = parsed.name
    if parsed.phone and (not recipient.phone or merge):
        recipient.phone = parsed.phone
    if parsed.email and (not recipient.email or merge):
        recipient.email = parsed.email
    if recipient.intake_source == "csv":
        recipient.intake_source = "merged"
    elif not recipient.intake_source:
        recipient.intake_source = "cv"


def _find_match(recipients: list[ServiceOrderRecipient], parsed: ParsedCv) -> ServiceOrderRecipient | None:
    if not parsed.name:
        return None
    best: ServiceOrderRecipient | None = None
    best_score = 0.0
    for r in recipients:
        score = name_similarity(r.name, parsed.name)
        if score > best_score:
            best_score = score
            best = r
    if best and best_score >= MATCH_THRESHOLD:
        return best
    return None


CONTACT_EXTENSIONS = (".csv", ".xlsx", ".xls")
CV_EXTENSIONS = (".pdf", ".docx", ".doc", ".zip", ".txt")


def classify_intake_filename(filename: str) -> str:
    lower = str(filename or "").lower().replace("\\", "/")
    base = lower.rsplit("/", 1)[-1]
    for ext in CONTACT_EXTENSIONS:
        if base.endswith(ext):
            return "contact"
    for ext in CV_EXTENSIONS:
        if base.endswith(ext):
            return "cv"
    return "unknown"


def _find_match_for_contact(recipients: list[ServiceOrderRecipient], row: dict[str, str | None]) -> ServiceOrderRecipient | None:
    phone = str(row.get("phone") or "").strip()
    name = str(row.get("name") or "").strip()
    if phone:
        norm_phone = re.sub(r"\D", "", phone)
        for r in recipients:
            if norm_phone and norm_phone == re.sub(r"\D", "", str(r.phone or "")):
                return r
    if name:
        best: ServiceOrderRecipient | None = None
        best_score = 0.0
        for r in recipients:
            score = name_similarity(r.name, name)
            if score > best_score:
                best_score = score
                best = r
        if best and best_score >= MATCH_THRESHOLD:
            return best
    return None


def intake_contacts_merge(db: Session, order: ServiceOrder, rows: list[dict[str, str | None]]) -> ServiceOrder:
    _assert_interview_draft(order)
    recipients = list(db.execute(select(ServiceOrderRecipient).where(ServiceOrderRecipient.order_id == order.id)).scalars())
    for row in rows:
        name = str(row.get("name") or "").strip()
        phone_raw = str(row.get("phone") or "").strip() or None
        email = str(row.get("email") or "").strip() or None
        if not name and not phone_raw:
            continue
        phone, phone_errors = _coerce_contact_phone(phone_raw)
        match = _find_match_for_contact(recipients, row)
        if match:
            if name and (not match.name or match.name == "Unknown"):
                match.name = name
            if phone and not match.phone:
                match.phone = phone
                match.intake_errors_json = _dumps_json(compute_intake_errors(match))
            if email and not match.email:
                match.email = email
            if match.intake_source == "cv":
                match.intake_source = "merged"
            db.add(match)
            continue
        recipient = ServiceOrderRecipient(
            order_id=order.id,
            row_number=len(recipients) + 1,
            name=name or "Unknown",
            phone=phone,
            email=email,
            status="pending",
            cv_quality="missing",
            intake_source="csv",
            intake_errors_json=_dumps_json(phone_errors),
        )
        db.add(recipient)
        recipients.append(recipient)
    db.flush()
    recipients = list(
        db.execute(
            select(ServiceOrderRecipient)
            .where(ServiceOrderRecipient.order_id == order.id)
            .order_by(ServiceOrderRecipient.row_number)
        ).scalars()
    )
    for i, r in enumerate(recipients, start=1):
        r.row_number = i
        db.add(r)
    order.recipient_count = len(recipients)
    order.updated_at = datetime.utcnow()
    db.add(order)
    db.commit()
    db.refresh(order)
    return order


def intake_mixed_files(db: Session, order: ServiceOrder, files: list[tuple[str, bytes]]) -> dict[str, Any]:
    _assert_interview_draft(order)
    if not files:
        raise ValueError("Upload at least one file")

    contact_files: list[tuple[str, bytes]] = []
    cv_files: list[tuple[str, bytes]] = []
    rejected: list[dict[str, str]] = []

    for filename, content in files:
        kind = classify_intake_filename(filename)
        if kind == "contact":
            contact_files.append((filename, content))
        elif kind == "cv":
            cv_files.append((filename, content))
        else:
            rejected.append({"filename": filename, "reason": "Unsupported file type — use Excel, CSV, PDF, DOCX, TXT, or ZIP"})

    contact_rows: list[dict[str, str | None]] = []
    for filename, content in contact_files:
        try:
            rows = parse_contacts_csv_relaxed_from_bytes(content, filename)
            contact_rows.extend(rows)
        except ValueError as e:
            rejected.append({"filename": filename, "reason": str(e)})

    existing = list(db.execute(select(ServiceOrderRecipient).where(ServiceOrderRecipient.order_id == order.id)).scalars())
    contacts_added = 0
    if contact_rows:
        if existing:
            order = intake_contacts_merge(db, order, contact_rows)
        else:
            order = intake_contacts_csv(db, order, contact_rows)
        contacts_added = len(contact_rows)

    cv_result: dict[str, Any] = {
        "parsed_count": 0,
        "unmatched_files": [],
    }
    if cv_files:
        cv_result = intake_cv_files(db, order, cv_files)
        order = ServiceOrderService.get_order(db, order.id) or order

    final_recipients = list(
        db.execute(
            select(ServiceOrderRecipient)
            .where(ServiceOrderRecipient.order_id == order.id)
            .order_by(ServiceOrderRecipient.row_number)
        ).scalars()
    )
    recipient_payload = [recipient_intake_dict(r) for r in final_recipients]
    for u in cv_result.get("unmatched_files") or []:
        rejected.append({"filename": u.get("filename", ""), "reason": "; ".join(u.get("errors") or [])})

    if not final_recipients and not contact_files and not cv_files:
        raise ValueError("No supported files in upload")
    if not final_recipients and rejected and not contact_rows and not cv_files:
        raise ValueError(rejected[0].get("reason") or "Could not process upload")

    return {
        "order_id": order.id,
        "contact_files": len(contact_files),
        "contact_rows": contacts_added,
        "parsed_count": cv_result.get("parsed_count", 0),
        "recipient_count": len(final_recipients),
        "rejected_files": rejected,
        "recipients": recipient_payload,
        "summary": intake_summary(recipient_payload),
    }


def intake_contacts_csv(db: Session, order: ServiceOrder, rows: list[dict[str, str | None]]) -> ServiceOrder:
    _assert_interview_draft(order)
    db.execute(delete(ServiceOrderRecipient).where(ServiceOrderRecipient.order_id == order.id))
    added = 0
    for i, row in enumerate(rows, start=1):
        name = str(row.get("name") or "").strip()
        phone_raw = str(row.get("phone") or "").strip() or None
        email = str(row.get("email") or "").strip() or None
        if not name and not phone_raw:
            continue
        phone, phone_errors = _coerce_contact_phone(phone_raw)
        recipient = ServiceOrderRecipient(
            order_id=order.id,
            row_number=i,
            name=name or "Unknown",
            phone=phone,
            email=email,
            status="pending",
            cv_quality="missing",
            intake_source="csv",
            intake_errors_json=_dumps_json(phone_errors),
        )
        db.add(recipient)
        added += 1
    order.recipient_count = added
    order.updated_at = datetime.utcnow()
    db.add(order)
    db.commit()
    db.refresh(order)
    return order


def parse_contacts_csv_relaxed_from_bytes(content: bytes, filename: str) -> list[dict[str, str | None]]:
    """Like parse_recipient_file but phone is optional during intake."""
    name = str(filename or "").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        raise ValueError(
            "Legacy .xls workbooks are not supported. In Excel choose Save As → .xlsx, or upload a CSV file."
        )
    if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        raise ValueError(
            "Legacy .xls workbooks are not supported. In Excel choose Save As → .xlsx, or upload a CSV file."
        )
    if name.endswith(".xlsx") or content[:2] == b"PK":
        import io

        try:
            import openpyxl
        except ImportError as e:
            raise ValueError("Excel upload requires openpyxl on the server. Use CSV for now.") from e

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(
                "Could not read the Excel file. Use .xlsx format with columns name and phone, or upload CSV."
            ) from e
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return []
        headers = [_canonical_contact_header(x) for x in header_row]
        out: list[dict[str, str | None]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cells = [
                ServiceOrderService._excel_cell_str(row[i]) if i < len(row) and row[i] is not None else ""
                for i in range(len(headers))
            ]
            # Extend if row has more cells than headers (rare).
            if row and len(row) > len(headers):
                for i in range(len(headers), len(row)):
                    cells.append(ServiceOrderService._excel_cell_str(row[i]) if row[i] is not None else "")
            name_val, phone_val, email_val = _contact_fields_from_row(headers, cells)
            if not name_val and not phone_val:
                continue
            out.append({"name": name_val or None, "phone": phone_val or None, "email": email_val or None})
        return out
    import csv
    import io

    from app.utils.text_decoding import decode_uploaded_text

    text = decode_uploaded_text(content)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV must include a header row: name, phone, email")
    headers = [_canonical_contact_header(k) for k in reader.fieldnames]
    out: list[dict[str, str | None]] = []
    for raw in reader:
        cells = [str(raw.get(k) or "").strip() for k in reader.fieldnames]
        name_val, phone_val, email_val = _contact_fields_from_row(headers, cells)
        if not name_val and not phone_val:
            continue
        out.append({"name": name_val or None, "phone": phone_val or None, "email": email_val or None})
    return out


def intake_email_cv_for_order(
    db: Session,
    order: ServiceOrder,
    *,
    parsed: ParsedCv,
    storage_key: str,
    sender_email: str | None = None,
) -> tuple[ServiceOrder, ServiceOrderRecipient]:
    _assert_interview_email_intake(order)
    recipients = list(db.execute(select(ServiceOrderRecipient).where(ServiceOrderRecipient.order_id == order.id)).scalars())
    email_val = parsed.email or sender_email
    dup = find_duplicate_recipient(recipients, name=parsed.name, phone=parsed.phone, email=email_val)
    if dup and parsed.name and dup.name and name_similarity(dup.name, parsed.name) >= MATCH_THRESHOLD:
        _apply_parsed_cv(dup, parsed, merge=True)
        dup.cv_storage_key = storage_key
        dup.intake_source = "email"
        db.add(dup)
        db.flush()
        recipients = list(
            db.execute(select(ServiceOrderRecipient).where(ServiceOrderRecipient.order_id == order.id)).scalars()
        )
        for i, r in enumerate(recipients, start=1):
            r.row_number = i
            db.add(r)
        order.recipient_count = len(recipients)
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(order)
        db.refresh(dup)
        return order, dup

    display_name = (parsed.name or name_from_filename(parsed.filename)).strip() or "Unknown"
    recipient = ServiceOrderRecipient(
        order_id=order.id,
        row_number=len(recipients) + 1,
        name=display_name,
        phone=parsed.phone or None,
        email=email_val or None,
        status="pending",
        cv_quality=parsed.quality,
        cv_filename=parsed.filename,
        cv_text=parsed.text or None,
        cv_parsed_json=_dumps_json(parsed.to_dict()),
        intake_errors_json=_dumps_json(parsed.errors),
        intake_source="email",
        cv_storage_key=storage_key,
    )
    db.add(recipient)
    recipients.append(recipient)
    db.flush()
    for i, r in enumerate(recipients, start=1):
        r.row_number = i
        db.add(r)
    order.recipient_count = len(recipients)
    order.updated_at = datetime.utcnow()
    db.add(order)
    db.commit()
    db.refresh(order)
    db.refresh(recipient)
    return order, recipient


def _cv_bytes_by_filename(files: list[tuple[str, bytes]]) -> dict[str, bytes]:
    """Map each CV filename to raw bytes (including DOCX/PDF inside ZIP archives)."""
    from app.services.interview_cv_parse_service import iter_cv_files_from_zip

    out: dict[str, bytes] = {}
    for filename, content in files:
        if str(filename or "").lower().endswith(".zip"):
            try:
                for inner_name, inner_bytes in iter_cv_files_from_zip(content):
                    out[inner_name] = inner_bytes
            except Exception:
                continue
        else:
            out[str(filename or "upload")] = content
    return out


def intake_cv_files(db: Session, order: ServiceOrder, files: list[tuple[str, bytes]]) -> dict[str, Any]:
    _assert_interview_draft(order)
    parsed_list = parse_uploaded_cv_files(files)
    bytes_by_name = _cv_bytes_by_filename(files)
    recipients = list(db.execute(select(ServiceOrderRecipient).where(ServiceOrderRecipient.order_id == order.id)).scalars())
    unmatched: list[dict[str, Any]] = []
    created = 0

    for parsed in parsed_list:
        raw_bytes = bytes_by_name.get(parsed.filename)
        raw_file = (parsed.filename, raw_bytes) if raw_bytes is not None else None
        dup = find_duplicate_recipient(recipients, name=parsed.name, phone=parsed.phone, email=parsed.email)
        if dup and parsed.name and dup.name and name_similarity(dup.name, parsed.name) >= MATCH_THRESHOLD:
            _apply_parsed_cv(dup, parsed, merge=True)
            if raw_file:
                from app.services.career_cv_storage_service import delete_cv_file, save_cv_bytes, storage_key_for

                delete_cv_file(dup.cv_storage_key)
                key = storage_key_for(org_id=order.org_id, order_id=order.id, filename=parsed.filename)
                save_cv_bytes(storage_key=key, content=raw_file[1])
                dup.cv_storage_key = key
            db.add(dup)
            continue

        match = _find_match(recipients, parsed)
        if match:
            if raw_file:
                from app.services.career_cv_storage_service import delete_cv_file, save_cv_bytes, storage_key_for

                delete_cv_file(match.cv_storage_key)
                key = storage_key_for(org_id=order.org_id, order_id=order.id, filename=parsed.filename)
                save_cv_bytes(storage_key=key, content=raw_file[1])
                match.cv_storage_key = key
            _apply_parsed_cv(match, parsed, merge=True)
            if match.intake_source in {None, "csv"}:
                match.intake_source = "merged"
            db.add(match)
            continue

        display_name = (parsed.name or name_from_filename(parsed.filename)).strip()
        if not display_name:
            unmatched.append({"filename": parsed.filename, "errors": parsed.errors or ["Could not identify candidate name"]})
            continue

        recipient = ServiceOrderRecipient(
            order_id=order.id,
            row_number=len(recipients) + created + 1,
            name=display_name,
            phone=parsed.phone or None,
            email=parsed.email or None,
            status="pending",
            cv_quality=parsed.quality,
            cv_filename=parsed.filename,
            cv_text=parsed.text or None,
            cv_parsed_json=_dumps_json(parsed.to_dict()),
            intake_errors_json=_dumps_json(parsed.errors),
            intake_source="cv",
        )
        if raw_file:
            from app.services.career_cv_storage_service import save_cv_bytes, storage_key_for

            key = storage_key_for(org_id=order.org_id, order_id=order.id, filename=parsed.filename)
            save_cv_bytes(storage_key=key, content=raw_file[1])
            recipient.cv_storage_key = key
        db.add(recipient)
        recipients.append(recipient)
        created += 1

    db.flush()

    recipients = list(
        db.execute(
            select(ServiceOrderRecipient)
            .where(ServiceOrderRecipient.order_id == order.id)
            .order_by(ServiceOrderRecipient.row_number)
        ).scalars()
    )
    for i, r in enumerate(recipients, start=1):
        r.row_number = i
        db.add(r)

    order.recipient_count = len(recipients)
    order.updated_at = datetime.utcnow()
    db.add(order)
    db.commit()
    db.refresh(order)

    final_recipients = list(
        db.execute(
            select(ServiceOrderRecipient)
            .where(ServiceOrderRecipient.order_id == order.id)
            .order_by(ServiceOrderRecipient.row_number)
        ).scalars()
    )
    from app.services.interview_ats_service import _order_job_context

    role, _ = _order_job_context(order)
    recipient_payload = [recipient_intake_dict(r, position=role) for r in final_recipients]
    return {
        "order_id": order.id,
        "parsed_count": len(parsed_list),
        "recipient_count": len(final_recipients),
        "unmatched_files": unmatched,
        "recipients": recipient_payload,
        "summary": intake_summary(recipient_payload),
    }


def list_intake_recipients(db: Session, order: ServiceOrder) -> list[dict[str, Any]]:
    rows = list(
        db.execute(
            select(ServiceOrderRecipient)
            .where(ServiceOrderRecipient.order_id == order.id)
            .order_by(ServiceOrderRecipient.row_number)
        ).scalars()
    )
    from app.models.interview_booking_token import InterviewBookingToken
    from app.services.interview_ats_service import _order_job_context

    tokens = list(
        db.execute(
            select(InterviewBookingToken).where(InterviewBookingToken.order_id == order.id)
        ).scalars()
    )
    token_by_recipient: dict[str, InterviewBookingToken] = {}
    for token in tokens:
        existing = token_by_recipient.get(token.recipient_id)
        if existing is None or (token.updated_at or token.created_at) > (existing.updated_at or existing.created_at):
            token_by_recipient[token.recipient_id] = token

    role, _ = _order_job_context(order)
    return [recipient_intake_dict(r, position=role, booking_token=token_by_recipient.get(r.id), order=order) for r in rows]


def update_intake_recipient(
    db: Session,
    order: ServiceOrder,
    recipient: ServiceOrderRecipient,
    payload: dict[str, Any],
) -> ServiceOrderRecipient:
    _assert_interview_draft(order)
    if recipient.order_id != order.id:
        raise ValueError("Recipient does not belong to this order")
    if "name" in payload:
        from app.services.recipient_contact_validation import normalize_recipient_name

        recipient.name = normalize_recipient_name(payload.get("name"), required=True)
    if "phone" in payload:
        from app.services.recipient_contact_validation import coerce_interview_phone_e164

        raw_phone = str(payload.get("phone") or "").strip() or None
        if not raw_phone:
            recipient.phone = None
        else:
            e164, err = coerce_interview_phone_e164(raw_phone)
            if err:
                raise ValueError(err)
            recipient.phone = e164
    if "email" in payload:
        from app.services.recipient_contact_validation import normalize_recipient_email

        recipient.email = normalize_recipient_email(payload.get("email"))
    recipient.intake_errors_json = _dumps_json(compute_intake_errors(recipient))
    order.updated_at = datetime.utcnow()
    db.add(recipient)
    db.add(order)
    db.commit()
    db.refresh(recipient)
    return recipient


def delete_intake_recipient(db: Session, order: ServiceOrder, recipient: ServiceOrderRecipient) -> ServiceOrder:
    _assert_can_delete_intake_recipient(order, recipient)
    if recipient.order_id != order.id:
        raise ValueError("Recipient does not belong to this order")
    from app.models.interview_booking_token import InterviewBookingToken
    from app.services.career_cv_storage_service import delete_cv_file

    db.execute(delete(InterviewBookingToken).where(InterviewBookingToken.recipient_id == recipient.id))
    delete_cv_file(recipient.cv_storage_key)
    db.delete(recipient)
    db.flush()
    remaining = list(
        db.execute(
            select(ServiceOrderRecipient)
            .where(ServiceOrderRecipient.order_id == order.id)
            .order_by(ServiceOrderRecipient.row_number)
        ).scalars()
    )
    for i, r in enumerate(remaining, start=1):
        r.row_number = i
        db.add(r)
    order.recipient_count = len(remaining)
    order.updated_at = datetime.utcnow()
    db.add(order)
    db.commit()
    db.refresh(order)
    return order


def intake_summary(recipients: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "total": len(recipients),
        "ready": sum(1 for r in recipients if r.get("intake_ready")),
        "missing_phone": sum(1 for r in recipients if any("phone missing" in str(e).lower() for e in (r.get("intake_errors") or []))),
        "cv_good": sum(1 for r in recipients if r.get("cv_quality") == "good"),
        "cv_low_quality": sum(1 for r in recipients if r.get("cv_quality") == "low_quality"),
        "cv_missing": sum(1 for r in recipients if r.get("cv_quality") == "missing"),
    }
