"""Survey results export helpers — PDF, CSV, and Excel with KPI + option breakdowns."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

from app.services.survey_report_template import _esc, logo_data_uri
from app.services.survey_wa_open_text_service import resolve_answer_text


def question_option_buckets(block: dict[str, Any]) -> list[dict[str, Any]]:
    """Normalize a question aggregate to 2–3 option rows with count + pct."""
    total = int(block.get("total") or 0)
    responses = list(block.get("responses") or [])
    breakdown = list(block.get("breakdown") or [])

    yes_no_keys = {str(r.get("answer") or "").strip().lower() for r in responses if r.get("answer")}
    if yes_no_keys and yes_no_keys.issubset({"yes", "no"}):
        counts = {"yes": 0, "no": 0}
        for row in responses:
            key = str(row.get("answer") or "").strip().lower()
            if key in counts:
                counts[key] += int(row.get("count") or 0)
        answered = counts["yes"] + counts["no"]
        base = answered or total or 1
        return [
            {"label": "Yes", "count": counts["yes"], "pct": round((counts["yes"] / base) * 100) if answered else 0},
            {"label": "No", "count": counts["no"], "pct": round((counts["no"] / base) * 100) if answered else 0},
        ]

    if breakdown:
        by_key = {str(g.get("key") or ""): g for g in breakdown}
        mapping = (
            ("positive", "Excellent"),
            ("neutral", "Expected"),
            ("negative", "Poor"),
        )
        rows: list[dict[str, Any]] = []
        answered = sum(int(by_key.get(key, {}).get("count") or 0) for key, _ in mapping)
        base = answered or total or 1
        for key, label in mapping:
            group = by_key.get(key) or {}
            count = int(group.get("count") or 0)
            pct = int(group.get("pct") or 0) if group.get("pct") is not None else round((count / base) * 100)
            rows.append({"label": label, "count": count, "pct": pct})
        return rows

    if responses and all(re.match(r"^\d+$", str(r.get("answer") or "").strip()) for r in responses):
        excellent = expected = poor = 0
        for row in responses:
            score = int(str(row.get("answer") or "0").strip())
            count = int(row.get("count") or 0)
            if score >= 9:
                excellent += count
            elif score >= 7:
                expected += count
            else:
                poor += count
        answered = excellent + expected + poor
        base = answered or total or 1
        return [
            {"label": "Excellent", "count": excellent, "pct": round((excellent / base) * 100) if answered else 0},
            {"label": "Expected", "count": expected, "pct": round((expected / base) * 100) if answered else 0},
            {"label": "Poor", "count": poor, "pct": round((poor / base) * 100) if answered else 0},
        ]

    base = total or sum(int(r.get("count") or 0) for r in responses) or 1
    limit = 3 if len(responses) >= 3 else max(2, len(responses)) if responses else 0
    if limit == 0:
        return []
    return [
        {
            "label": str(row.get("answer") or "—"),
            "count": int(row.get("count") or 0),
            "pct": round((int(row.get("count") or 0) / base) * 100),
        }
        for row in responses[:limit]
    ]


def excellent_rate_pct(summary: dict[str, Any]) -> int:
    avg_sat5 = summary.get("average_satisfaction_5")
    if avg_sat5 is not None:
        try:
            return round((float(avg_sat5) / 5) * 100)
        except (TypeError, ValueError):
            pass
    return int(summary.get("response_rate_pct") or 0)


def poor_rate_pct(summary: dict[str, Any]) -> int:
    sentiment = summary.get("sentiment_counts") or {}
    if isinstance(sentiment, dict):
        total = sum(int(v or 0) for v in sentiment.values())
        if total > 0:
            return round((int(sentiment.get("negative") or 0) / total) * 100)
    return int(summary.get("nps_detractors_pct") or 0)


def _format_recommend_pct(summary: dict[str, Any]) -> str:
    value = summary.get("recommend_pct")
    if value is None:
        return "—"
    return f"{int(value)}%"


def _kpi_rows(payload: dict[str, Any]) -> list[tuple[str, Any]]:
    order = payload.get("order") or {}
    summary = payload.get("summary") or {}
    nps_score = summary.get("nps_score")
    if isinstance(nps_score, dict):
        nps_score = nps_score.get("score")
    return [
        ("Survey", order.get("title") or "Survey"),
        ("Organisation", order.get("organisation_name") or order.get("goal") or ""),
        ("Channel", order.get("channel") or ""),
        ("Total invited", summary.get("total_recipients", 0)),
        ("Completed responses", summary.get("completed_count", 0)),
        ("Response rate %", summary.get("response_rate_pct", 0)),
        ("Excellent rate %", excellent_rate_pct(summary)),
        ("Average satisfaction /5", summary.get("average_satisfaction_5", "")),
        ("Poor rating %", poor_rate_pct(summary)),
        ("NPS score", nps_score if nps_score is not None else ""),
        ("NPS label", summary.get("nps_label") or ""),
        ("Promoters %", summary.get("nps_promoters_pct", 0)),
        ("Passives %", summary.get("nps_passives_pct", 0)),
        ("Detractors %", summary.get("nps_detractors_pct", 0)),
        ("Would recommend %", summary.get("recommend_pct") if summary.get("recommend_pct") is not None else ""),
        ("Avg duration", summary.get("average_call_duration_label") or ""),
    ]


def build_survey_results_csv(payload: dict[str, Any], *, anonymous: bool = False) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow(["Key metrics"])
    writer.writerow(["Metric", "Value"])
    for label, value in _kpi_rows(payload):
        writer.writerow([label, value])
    writer.writerow([])

    writer.writerow(["Question breakdown"])
    writer.writerow(["Question", "Option", "Count", "Percent"])
    for block in payload.get("aggregates") or []:
        question = str(block.get("question") or "Question")
        options = question_option_buckets(block)
        if not options:
            writer.writerow([question, "No answers", 0, "0%"])
            continue
        for idx, option in enumerate(options):
            writer.writerow(
                [
                    question if idx == 0 else "",
                    option["label"],
                    option["count"],
                    f"{option['pct']}%",
                ]
            )
    writer.writerow([])

    writer.writerow(["Raw answer counts"])
    writer.writerow(["Question", "Answer", "Count"])
    for block in payload.get("aggregates") or []:
        question = block.get("question") or "Question"
        for row in block.get("responses") or []:
            writer.writerow([question, row.get("answer"), row.get("count")])
    writer.writerow([])

    writer.writerow(["Respondent summary"])
    writer.writerow(["Respondent", "Status", "Final feedback Yes/No", "Additional feedback"])
    for row in payload.get("respondents") or []:
        if not isinstance(row, dict):
            continue
        name = row.get("initials") or row.get("id") if anonymous else row.get("name") or row.get("id") or ""
        writer.writerow(
            [
                name,
                row.get("status_label") or row.get("status") or "",
                row.get("final_feedback_yes_no") or "",
                row.get("final_additional_feedback") or "",
            ]
        )
    writer.writerow([])

    writer.writerow(["Open feedback & voice answers"])
    writer.writerow(
        [
            "Respondent",
            "Question",
            "English text",
            "Original text",
            "Source",
            "Language",
            "Transcription status",
            "Translation status",
        ]
    )
    for row in payload.get("respondents") or []:
        if not isinstance(row, dict):
            continue
        label = row.get("initials") or row.get("id") if anonymous else row.get("name") or row.get("id") or ""
        for ans in row.get("open_feedback") or []:
            if not isinstance(ans, dict):
                continue
            writer.writerow(
                [
                    label,
                    ans.get("question") or "",
                    ans.get("translated_text") or ans.get("transcript") or ans.get("text") or "",
                    ans.get("original_text") or "",
                    ans.get("answer_source") or "",
                    ans.get("detected_language") or "",
                    ans.get("transcription_status") or "",
                    ans.get("translation_status") or "",
                ]
            )
        for ans in row.get("wa_answers") or []:
            if not isinstance(ans, dict) or str(ans.get("answer_source") or "") != "voice_note":
                continue
            writer.writerow(
                [
                    label,
                    ans.get("question") or "",
                    ans.get("translated_text") or resolve_answer_text(ans),
                    ans.get("original_text") or resolve_answer_text(ans),
                    ans.get("answer_source") or "voice_note",
                    ans.get("detected_language") or "",
                    ans.get("transcription_status") or "",
                    ans.get("translation_status") or "",
                ]
            )

    # UTF-8 BOM helps Excel open the file with correct encoding.
    return "\ufeff" + buf.getvalue()


def build_survey_results_xlsx(payload: dict[str, Any], *, anonymous: bool = False) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("Excel export requires openpyxl on the server.") from e

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["Metric", "Value"])
    summary_ws["A1"].font = bold
    summary_ws["B1"].font = bold
    for label, value in _kpi_rows(payload):
        summary_ws.append([label, value])

    questions_ws = wb.create_sheet("Questions")
    questions_ws.append(["Question", "Option", "Count", "Percent"])
    for cell in questions_ws[1]:
        cell.font = bold
    for block in payload.get("aggregates") or []:
        question = str(block.get("question") or "Question")
        options = question_option_buckets(block)
        if not options:
            questions_ws.append([question, "No answers", 0, "0%"])
            continue
        for idx, option in enumerate(options):
            questions_ws.append(
                [
                    question if idx == 0 else "",
                    option["label"],
                    option["count"],
                    f"{option['pct']}%",
                ]
            )

    responses_ws = wb.create_sheet("Responses")
    responses_ws.append(["Respondent", "Status", "Final feedback Yes/No", "Additional feedback"])
    for cell in responses_ws[1]:
        cell.font = bold
    for row in payload.get("respondents") or []:
        if not isinstance(row, dict):
            continue
        name = row.get("initials") or row.get("id") if anonymous else row.get("name") or row.get("id") or ""
        responses_ws.append(
            [
                name,
                row.get("status_label") or row.get("status") or "",
                row.get("final_feedback_yes_no") or "",
                row.get("final_additional_feedback") or "",
            ]
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _export_question_table_rows(aggregates: list[dict[str, Any]]) -> str:
    rows_html: list[str] = []
    for block in aggregates:
        question = _esc(block.get("question") or "Question")
        total = int(block.get("total") or 0)
        options = question_option_buckets(block)
        if not options:
            rows_html.append(
                f"<tr><td><strong>{question}</strong></td><td colspan='3'>No answers recorded</td></tr>"
            )
            continue
        for idx, option in enumerate(options):
            label = _esc(option["label"])
            count = int(option["count"])
            pct = int(option["pct"])
            q_cell = f"<strong>{question}</strong><br/><span style='color:#666;font-size:10px'>{total} responses</span>" if idx == 0 else ""
            bar = (
                f"<div style='background:#e2e2e2;border-radius:4px;height:8px;width:120px'>"
                f"<div style='background:#1a7a52;height:8px;width:{max(4, pct)}%;border-radius:4px'></div></div>"
            )
            rows_html.append(
                f"<tr><td>{q_cell}</td><td>{label}</td><td style='text-align:right'>{count}</td>"
                f"<td style='text-align:right'>{pct}%</td><td>{bar}</td></tr>"
            )
    return "".join(rows_html) if rows_html else (
        "<tr><td colspan='5' style='color:#666'>No aggregated answers yet.</td></tr>"
    )


def build_survey_results_export_html(payload: dict[str, Any], *, logo_uri: str | None = None) -> str:
    """PDF-safe HTML using tables + inline styles (works with WeasyPrint and fpdf fallback)."""
    order = payload.get("order") or {}
    summary = payload.get("summary") or {}
    aggregates = payload.get("aggregates") or []
    recommendations = payload.get("recommendations") or []

    company = _esc(order.get("organisation_name") or order.get("goal") or "Your organisation")
    title = _esc(order.get("title") or "Survey results")
    logo_uri = logo_uri if logo_uri is not None else logo_data_uri()
    logo_html = (
        f'<img src="{logo_uri}" alt="VOXBULK" style="height:28px;width:auto;max-width:180px"/>'
        if logo_uri
        else ""
    )

    completed = int(summary.get("completed_count") or 0)
    total = int(summary.get("total_recipients") or 0) or completed or 1
    response_rate = int(summary.get("response_rate_pct") or round((completed / total) * 100))
    excellent = excellent_rate_pct(summary)
    poor = poor_rate_pct(summary)
    sat5 = summary.get("average_satisfaction_5")
    sat_label = f"{sat5}/5" if sat5 is not None else "—"
    recommend = _format_recommend_pct(summary)

    nps_score = summary.get("nps_score")
    if isinstance(nps_score, dict):
        nps_score = nps_score.get("score")
    nps_label = _esc(summary.get("nps_label") or "—")
    promoters = int(summary.get("nps_promoters_pct") or 0)
    passives = int(summary.get("nps_passives_pct") or 0)
    detractors = int(summary.get("nps_detractors_pct") or 0)

    sentiment = summary.get("sentiment_counts") or {}
    sent_total = sum(int(v or 0) for v in sentiment.values()) if isinstance(sentiment, dict) else 0
    sent_rows = []
    for label, key, color in (
        ("Positive", "positive", "#1a7a52"),
        ("Neutral", "neutral", "#1854a8"),
        ("Negative", "negative", "#a32020"),
    ):
        count = int(sentiment.get(key) or 0) if isinstance(sentiment, dict) else 0
        pct = round((count / sent_total) * 100) if sent_total else 0
        sent_rows.append(
            f"<tr><td>{label}</td><td style='text-align:right'>{count}</td>"
            f"<td style='text-align:right'>{pct}%</td>"
            f"<td><div style='background:#e2e2e2;height:8px;border-radius:4px'>"
            f"<div style='width:{max(4, pct)}%;background:{color};height:8px;border-radius:4px'></div></div></td></tr>"
        )

    action_rows = []
    for rec in recommendations[:5]:
        action_rows.append(
            f"<tr><td><strong>{_esc(rec.get('title') or 'Recommendation')}</strong><br/>"
            f"{_esc(rec.get('text') or '')}</td><td>{_esc(rec.get('impact') or 'Medium')}</td></tr>"
        )
    if not action_rows:
        action_rows.append("<tr><td colspan='2'>Recommendations will appear once enough responses are analysed.</td></tr>")

    question_rows = _export_question_table_rows(aggregates)

    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8"/><title>{title}</title></head>
<body style="font-family:Helvetica,Arial,sans-serif;font-size:11px;color:#0f0f0f;margin:24px">
  <table style="width:100%;border-bottom:1px solid #e2e2e2;margin-bottom:16px"><tr>
    <td>{logo_html}</td>
    <td style="text-align:right;color:#787878;font-size:9px">Anonymous aggregate report<br/>Confidential</td>
  </tr></table>

  <div style="font-size:9px;color:#787878;text-transform:uppercase;letter-spacing:.08em">{company}</div>
  <h1 style="font-size:22px;margin:4px 0 16px">Survey results — {title}</h1>

  <table style="width:100%;margin-bottom:18px;border-collapse:collapse">
    <tr>
      <td style="padding:8px;border:1px solid #e2e2e2;text-align:center"><div style="font-size:18px;font-weight:700">{completed}</div><div style="font-size:9px;color:#787878">Responses</div></td>
      <td style="padding:8px;border:1px solid #e2e2e2;text-align:center"><div style="font-size:18px;font-weight:700">{response_rate}%</div><div style="font-size:9px;color:#787878">Response rate</div></td>
      <td style="padding:8px;border:1px solid #e2e2e2;text-align:center"><div style="font-size:18px;font-weight:700">{excellent}%</div><div style="font-size:9px;color:#787878">Excellent rate</div></td>
      <td style="padding:8px;border:1px solid #e2e2e2;text-align:center"><div style="font-size:18px;font-weight:700">{poor}%</div><div style="font-size:9px;color:#787878">Poor rating</div></td>
    </tr>
  </table>

  <h2 style="font-size:14px;margin:16px 0 8px">Key metrics</h2>
  <table style="width:100%;border-collapse:collapse;margin-bottom:16px" border="1" cellpadding="6">
    <tr style="background:#f6f5f2"><th align="left">Metric</th><th align="left">Value</th></tr>
    <tr><td>Average satisfaction</td><td>{sat_label}</td></tr>
    <tr><td>Would recommend</td><td>{recommend}</td></tr>
    <tr><td>NPS score</td><td>{nps_score if nps_score is not None else '—'} ({nps_label})</td></tr>
    <tr><td>Promoters / Passives / Detractors</td><td>{promoters}% / {passives}% / {detractors}%</td></tr>
    <tr><td>Total invited</td><td>{total}</td></tr>
    <tr><td>Avg duration</td><td>{_esc(summary.get('average_call_duration_label') or '—')}</td></tr>
  </table>

  <h2 style="font-size:14px;margin:16px 0 8px">Sentiment</h2>
  <table style="width:100%;border-collapse:collapse;margin-bottom:16px" border="1" cellpadding="6">
    <tr style="background:#f6f5f2"><th align="left">Tone</th><th align="right">Count</th><th align="right">%</th><th align="left">Distribution</th></tr>
    {''.join(sent_rows) if sent_rows else "<tr><td colspan='4'>No sentiment data yet.</td></tr>"}
  </table>

  <h2 style="font-size:14px;margin:16px 0 8px">Question breakdown</h2>
  <table style="width:100%;border-collapse:collapse;margin-bottom:16px" border="1" cellpadding="6">
    <tr style="background:#f6f5f2"><th align="left">Question</th><th align="left">Option</th><th align="right">Count</th><th align="right">%</th><th align="left">Bar</th></tr>
    {question_rows}
  </table>

  <h2 style="font-size:14px;margin:16px 0 8px">Recommended actions</h2>
  <table style="width:100%;border-collapse:collapse;margin-bottom:16px" border="1" cellpadding="6">
    <tr style="background:#f6f5f2"><th align="left">Action</th><th align="left">Impact</th></tr>
    {''.join(action_rows)}
  </table>

  <p style="font-size:9px;color:#787878;margin-top:24px">Individual names and transcripts are never included in customer-facing survey reports.</p>
</body></html>"""
