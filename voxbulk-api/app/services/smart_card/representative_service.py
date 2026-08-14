"""Smart Card QR representatives CRUD, product assignment, invites hook."""

from __future__ import annotations

import json
import logging
from datetime import datetime
from typing import Any

from sqlalchemy import delete, or_, select
from sqlalchemy.orm import Session

from app.models.smart_card import (
    SmartCardProduct,
    SmartCardRepresentative,
    SmartCardRepresentativeProduct,
)
from app.services.org_rbac import ORG_TEAM_MANAGERS, OrgRbacService, can_view_all_campaigns, effective_role
from app.services.qr_style_fields import apply_qr_style_payload, qr_style_dict
from app.services.qr_style_render import (
    normalize_corner_style,
    normalize_frame_round,
    normalize_module_style,
)
from app.services.smart_card.company_service import (
    SmartCardCompanyService,
    SmartCardEntitlementService,
    build_rep_qr_token,
)

logger = logging.getLogger(__name__)

MEMBER_EDIT_KEYS = frozenset(
    {
        "name",
        "email",
        "website",
        "mobile",
        "landline",
        "extension",
        "notes",
        "social_links",
        "extra",
        "qr_fg_color",
        "qr_bg_color",
        "qr_transparent",
        "qr_module_style",
        "qr_corner_style",
        "qr_show_arrow",
        "qr_frame_round",
        "product_ids",
    }
)


class SmartCardRepError(ValueError):
    pass


class SmartCardRepresentativeService:
    @staticmethod
    def list_for_user(
        db: Session,
        *,
        org_id: str,
        user_id: str,
        q: str | None = None,
    ) -> list[SmartCardRepresentative]:
        role = OrgRbacService.role_for(db, org_id=org_id, user_id=user_id)
        # Heal: claim any unlinked QR that matches this member's email (invite accepted earlier).
        if not can_view_all_campaigns(role):
            try:
                from app.models.user import User
                from app.services.org_invite_service import link_smart_card_reps_for_invite

                user = db.get(User, user_id)
                if user is not None:
                    n = link_smart_card_reps_for_invite(
                        db,
                        org_id=org_id,
                        user=user,
                        email=str(user.email or ""),
                    )
                    if n:
                        db.commit()
            except Exception:
                logger.exception("smart_card_auto_claim_failed org=%s user=%s", org_id, user_id)
        stmt = select(SmartCardRepresentative).where(SmartCardRepresentative.org_id == org_id)
        if not can_view_all_campaigns(role):
            stmt = stmt.where(SmartCardRepresentative.linked_user_id == user_id)
        if q and str(q).strip():
            like = f"%{str(q).strip()}%"
            stmt = stmt.where(
                or_(
                    SmartCardRepresentative.name.ilike(like),
                    SmartCardRepresentative.email.ilike(like),
                    SmartCardRepresentative.mobile.ilike(like),
                )
            )
        stmt = stmt.order_by(SmartCardRepresentative.name.asc())
        return list(db.execute(stmt).scalars().all())

    @staticmethod
    def get(db: Session, *, org_id: str, rep_id: str) -> SmartCardRepresentative | None:
        return db.execute(
            select(SmartCardRepresentative).where(
                SmartCardRepresentative.id == rep_id,
                SmartCardRepresentative.org_id == org_id,
            )
        ).scalar_one_or_none()

    @staticmethod
    def create(
        db: Session,
        *,
        org_id: str,
        user_id: str,
        payload: dict[str, Any],
    ) -> SmartCardRepresentative:
        name = str(payload.get("name") or "").strip()
        if not name:
            raise SmartCardRepError("name is required")
        seats = SmartCardEntitlementService.seat_quantity(db, org_id)
        active = SmartCardEntitlementService.active_rep_count(db, org_id)
        # Allow creating reps in preview without seats, but cap active to seats when live.
        if seats > 0 and active >= seats:
            raise SmartCardRepError(
                f"Seat limit reached ({seats}). Buy more seats or archive a representative."
            )
        company = SmartCardCompanyService.get_or_create(db, org_id)
        token = build_rep_qr_token(company_slug=company.name or org_id[:8], rep_name=name)
        while (
            db.execute(
                select(SmartCardRepresentative).where(SmartCardRepresentative.qr_token == token)
            ).scalar_one_or_none()
            is not None
        ):
            token = build_rep_qr_token(company_slug=company.name or org_id[:8], rep_name=name)

        social = payload.get("social_links")
        extra = payload.get("extra")
        qcfg = payload.get("question_config")
        qcfg_json = None
        if isinstance(qcfg, dict):
            qcfg_json = json.dumps(qcfg)
        rep = SmartCardRepresentative(
            org_id=org_id,
            name=name,
            email=(str(payload.get("email") or "").strip().lower() or None),
            website=(str(payload.get("website") or "").strip() or None),
            social_links_json=json.dumps(social) if social is not None else None,
            mobile=(str(payload.get("mobile") or "").strip() or None),
            landline=(str(payload.get("landline") or "").strip() or None),
            extension=(str(payload.get("extension") or "").strip() or None),
            notes=(str(payload.get("notes") or "").strip() or None),
            extra_json=json.dumps(extra) if extra is not None else None,
            question_config_json=qcfg_json,
            qr_token=token,
            qr_fg_color=str(payload.get("qr_fg_color") or "000000").replace("#", "")[:6],
            qr_bg_color=str(payload.get("qr_bg_color") or "ffffff").replace("#", "")[:6],
            qr_transparent=bool(payload.get("qr_transparent")),
            qr_module_style=normalize_module_style(payload.get("qr_module_style")),
            qr_corner_style=normalize_corner_style(payload.get("qr_corner_style")),
            qr_show_arrow=False,
            qr_frame_round=normalize_frame_round(payload.get("qr_frame_round")),
            status="active",
            created_by_user_id=user_id,
        )
        db.add(rep)
        db.flush()
        product_ids = payload.get("product_ids")
        if isinstance(product_ids, list):
            SmartCardRepresentativeService.set_products(
                db, org_id=org_id, representative_id=rep.id, product_ids=[str(x) for x in product_ids]
            )
        if rep.email:
            try:
                SmartCardRepresentativeService.invite_or_link_rep(
                    db, org_id=org_id, actor_user_id=user_id, rep=rep
                )
            except Exception as e:
                logger.warning("smart_card_invite_on_create_failed rep=%s err=%s", rep.id, e)
        return rep

    @staticmethod
    def invite_or_link_rep(
        db: Session,
        *,
        org_id: str,
        actor_user_id: str,
        rep: SmartCardRepresentative,
        force_resend: bool = False,
    ) -> dict[str, Any]:
        """Skip admin/owner emails; link existing members; else send member invite."""
        email = (rep.email or "").strip().lower()
        if not email or "@" not in email:
            return {"action": "skipped", "reason": "no_email"}

        from app.models.membership import OrganisationMembership
        from app.models.organisation import Organisation
        from app.models.user import User
        from app.services.org_team_service import OrgTeamService
        from app.services.smart_card.email_service import SmartCardEmailService

        actor = db.get(User, actor_user_id)
        org = db.get(Organisation, org_id)
        org_name = (org.name if org else "your organisation") or "your organisation"

        existing_user = db.execute(select(User).where(User.email == email)).scalar_one_or_none()
        if existing_user is not None:
            mem = db.execute(
                select(OrganisationMembership).where(
                    OrganisationMembership.org_id == org_id,
                    OrganisationMembership.user_id == existing_user.id,
                )
            ).scalar_one_or_none()
            if mem is not None:
                role = effective_role(mem.role)
                is_admin = role in ORG_TEAM_MANAGERS or str(existing_user.id) == str(actor_user_id)
                # Always link when already in org; never send invite to owner/manager/admin.
                if not rep.linked_user_id:
                    rep.linked_user_id = existing_user.id
                    db.add(rep)
                    db.flush()
                if is_admin:
                    return {"action": "linked_admin", "linked_user_id": existing_user.id}
                return {"action": "linked_member", "linked_user_id": existing_user.id}

        if rep.linked_user_id and not force_resend:
            return {"action": "already_linked", "linked_user_id": rep.linked_user_id}

        if actor is None:
            logger.warning("smart_card_invite_skipped rep=%s reason=no_actor", rep.id)
            return {"action": "skipped", "reason": "no_actor"}

        try:
            invite = OrgTeamService.create_invite(
                db,
                org_id=org_id,
                email=email,
                role="member",
                invited_by=actor,
                send_email=False,
            )
            rep.invite_id = invite.get("invite_id")
            db.add(rep)
            db.flush()
            sent = SmartCardEmailService.send_rep_member_invite(
                db,
                to_email=email,
                rep_name=rep.name or "there",
                org_name=org_name,
                signup_url=str(invite.get("signup_url") or ""),
            )
            return {
                "action": "invited",
                "invite_id": invite.get("invite_id"),
                "email_sent": bool(sent),
                "signup_url": invite.get("signup_url"),
            }
        except ValueError as e:
            msg = str(e)
            logger.warning("smart_card_invite_failed rep=%s err=%s", rep.id, msg)
            if "already belongs" in msg.lower():
                u = db.execute(select(User).where(User.email == email)).scalar_one_or_none()
                if u is not None:
                    rep.linked_user_id = u.id
                    db.add(rep)
                    db.flush()
                    return {"action": "linked_member", "linked_user_id": u.id}
            raise SmartCardRepError(msg) from e
        except Exception as e:
            logger.exception("smart_card_invite_unexpected rep=%s", rep.id)
            raise SmartCardRepError(f"Could not send invite: {e}") from e

    @staticmethod
    def invite_status_for(rep: SmartCardRepresentative) -> str:
        email = (rep.email or "").strip()
        if not email:
            return "needs_email"
        if rep.linked_user_id:
            return "linked"
        if rep.invite_id:
            return "pending_invite"
        return "needs_invite"

    @staticmethod
    def card_incomplete(rep: SmartCardRepresentative) -> bool:
        """Stub / empty card — invitee should complete contact details."""
        has_contact = bool((rep.mobile or "").strip() or (rep.landline or "").strip() or (rep.website or "").strip())
        return not has_contact

    @staticmethod
    def parse_invite_file(content: bytes, filename: str | None = None) -> list[dict[str, str]]:
        """Parse .xlsx/.csv with email (+ optional name) columns."""
        import csv
        import io
        import re

        name = (filename or "").strip().lower()
        rows_out: list[dict[str, str]] = []

        def _norm_header(h: str) -> str:
            return re.sub(r"[^a-z0-9]", "", (h or "").strip().lower())

        def _map_headers(headers: list[str]) -> dict[str, int]:
            mapping: dict[str, int] = {}
            for i, h in enumerate(headers):
                key = _norm_header(str(h or ""))
                if key in {"email", "e-mail", "mail", "workemail", "emailaddress"} or key.endswith("email"):
                    mapping.setdefault("email", i)
                elif key in {"name", "fullname", "fullname", "displayname", "rep", "representative"}:
                    mapping.setdefault("name", i)
            return mapping

        def _append_row(email_raw: Any, name_raw: Any) -> None:
            email = str(email_raw or "").strip().lower()
            if not email or "@" not in email or "." not in email.split("@")[-1]:
                return
            display = str(name_raw or "").strip()
            rows_out.append({"email": email, "name": display})

        is_xlsx = name.endswith(".xlsx") or (not name.endswith(".csv") and content[:2] == b"PK")
        if is_xlsx:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
            wb.close()
            if not grid:
                return []
            headers = [str(c or "") for c in grid[0]]
            mapping = _map_headers(headers)
            email_idx = mapping.get("email")
            name_idx = mapping.get("name")
            start = 1
            if email_idx is None:
                # Fallback: first column is email if it looks like one in row 1 or 2
                email_idx = 0
                name_idx = 1 if len(headers) > 1 else None
                sample = str(grid[0][0] or "")
                if "@" not in sample:
                    start = 1
                else:
                    start = 0
            for row in grid[start:]:
                if not row:
                    continue
                em = row[email_idx] if email_idx is not None and email_idx < len(row) else None
                nm = row[name_idx] if name_idx is not None and name_idx < len(row) else None
                _append_row(em, nm)
            return rows_out

        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            return []
        headers = [str(c or "") for c in all_rows[0]]
        mapping = _map_headers(headers)
        email_idx = mapping.get("email", 0)
        name_idx = mapping.get("name")
        start = 1 if mapping.get("email") is not None or "@" not in str(all_rows[0][0] if all_rows[0] else "") else 0
        for row in all_rows[start:]:
            if not row:
                continue
            em = row[email_idx] if email_idx < len(row) else None
            nm = row[name_idx] if name_idx is not None and name_idx < len(row) else None
            _append_row(em, nm)
        return rows_out

    @staticmethod
    def bulk_invite_template_xlsx() -> bytes:
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invites"
        ws.append(["email", "name"])
        ws.append(["alex@example.com", "Alex Example"])
        ws.append(["sam@example.com", ""])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def bulk_invite_from_emails(
        db: Session,
        *,
        org_id: str,
        actor_user_id: str,
        rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """Create stub representatives + send invites, capped by remaining seats."""
        seats = SmartCardEntitlementService.seat_quantity(db, org_id)
        active = SmartCardEntitlementService.active_rep_count(db, org_id)
        if seats <= 0:
            raise SmartCardRepError(
                "No Smart Card seats available. Buy seats first, then invite your team."
            )
        remaining = max(0, seats - active)

        seen: set[str] = set()
        created: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        invited = 0
        linked = 0

        existing_emails = {
            str(e or "").strip().lower()
            for e in db.execute(
                select(SmartCardRepresentative.email).where(
                    SmartCardRepresentative.org_id == org_id,
                    SmartCardRepresentative.status == "active",
                    SmartCardRepresentative.email.is_not(None),
                )
            ).scalars().all()
        }

        for raw in rows or []:
            email = str((raw or {}).get("email") or "").strip().lower()
            display_name = str((raw or {}).get("name") or "").strip()
            if not email or "@" not in email:
                skipped.append({"email": email or None, "reason": "invalid_email"})
                continue
            if email in seen:
                skipped.append({"email": email, "reason": "duplicate_in_file"})
                continue
            seen.add(email)
            if email in existing_emails:
                skipped.append({"email": email, "reason": "already_has_qr"})
                continue
            if remaining <= 0:
                skipped.append({"email": email, "reason": "seat_limit"})
                continue

            stub_name = display_name or email.split("@")[0].replace(".", " ").replace("_", " ").strip()
            if not stub_name:
                stub_name = email
            try:
                # create() also invites when email present
                rep = SmartCardRepresentativeService.create(
                    db,
                    org_id=org_id,
                    user_id=actor_user_id,
                    payload={"name": stub_name, "email": email},
                )
            except SmartCardRepError as e:
                skipped.append({"email": email, "reason": str(e)})
                continue

            remaining -= 1
            existing_emails.add(email)
            status = SmartCardRepresentativeService.invite_status_for(rep)
            action = "created"
            if rep.linked_user_id:
                linked += 1
                action = "linked"
            elif rep.invite_id:
                invited += 1
                action = "invited"
            created.append(
                {
                    "id": rep.id,
                    "email": email,
                    "name": rep.name,
                    "action": action,
                    "invite_status": status,
                }
            )

        return {
            "created_count": len(created),
            "invited_count": invited,
            "linked_count": linked,
            "skipped_count": len(skipped),
            "remaining_seats": remaining,
            "seat_quantity": seats,
            "active_reps": active + len(created),
            "created": created,
            "skipped": skipped,
        }

    @staticmethod
    def member_safe_payload(payload: dict[str, Any]) -> dict[str, Any]:
        return {k: v for k, v in (payload or {}).items() if k in MEMBER_EDIT_KEYS}

    @staticmethod
    def update(
        db: Session,
        *,
        org_id: str,
        rep_id: str,
        payload: dict[str, Any],
    ) -> SmartCardRepresentative:
        rep = SmartCardRepresentativeService.get(db, org_id=org_id, rep_id=rep_id)
        if rep is None:
            raise SmartCardRepError("Representative not found")
        for key in ("name", "email", "website", "mobile", "landline", "extension", "notes"):
            if key in payload:
                val = payload[key]
                if key == "email" and val:
                    val = str(val).strip().lower()
                elif val is not None:
                    val = str(val).strip()
                setattr(rep, key, val or None)
        if "name" in payload and not (rep.name or "").strip():
            raise SmartCardRepError("name is required")
        if "social_links" in payload:
            rep.social_links_json = json.dumps(payload.get("social_links") or {})
        if "extra" in payload:
            rep.extra_json = json.dumps(payload.get("extra") or {})
        if "qr_fg_color" in payload:
            rep.qr_fg_color = str(payload.get("qr_fg_color") or "000000").replace("#", "")[:6]
        if "qr_bg_color" in payload:
            rep.qr_bg_color = str(payload.get("qr_bg_color") or "ffffff").replace("#", "")[:6]
        if "qr_transparent" in payload:
            rep.qr_transparent = bool(payload.get("qr_transparent"))
        apply_qr_style_payload(rep, payload, allow_transparent=False)
        if "status" in payload:
            status = str(payload.get("status") or "").strip().lower()
            if status in {"active", "archived"}:
                if status == "active" and rep.status != "active":
                    seats = SmartCardEntitlementService.seat_quantity(db, org_id)
                    active = SmartCardEntitlementService.active_rep_count(db, org_id)
                    if seats > 0 and active >= seats:
                        raise SmartCardRepError("Seat limit reached — cannot restore representative")
                rep.status = status
        if "product_ids" in payload and isinstance(payload.get("product_ids"), list):
            SmartCardRepresentativeService.set_products(
                db,
                org_id=org_id,
                representative_id=rep.id,
                product_ids=[str(x) for x in payload["product_ids"]],
            )
        if "question_config" in payload:
            cfg = payload.get("question_config")
            if cfg is None:
                rep.question_config_json = None
            elif isinstance(cfg, dict):
                rep.question_config_json = json.dumps(cfg)
            else:
                raise SmartCardRepError("question_config must be an object or null")
        rep.updated_at = datetime.utcnow()
        db.add(rep)
        db.flush()
        return rep

    @staticmethod
    def set_products(
        db: Session,
        *,
        org_id: str,
        representative_id: str,
        product_ids: list[str],
    ) -> None:
        db.execute(
            delete(SmartCardRepresentativeProduct).where(
                SmartCardRepresentativeProduct.representative_id == representative_id,
                SmartCardRepresentativeProduct.org_id == org_id,
            )
        )
        seen: set[str] = set()
        for pid in product_ids:
            pid = str(pid or "").strip()
            if not pid or pid in seen:
                continue
            seen.add(pid)
            prod = db.execute(
                select(SmartCardProduct).where(
                    SmartCardProduct.id == pid,
                    SmartCardProduct.org_id == org_id,
                )
            ).scalar_one_or_none()
            if prod is None:
                continue
            db.add(
                SmartCardRepresentativeProduct(
                    org_id=org_id,
                    representative_id=representative_id,
                    product_id=pid,
                )
            )
        db.flush()

    @staticmethod
    def product_ids(db: Session, *, representative_id: str) -> list[str]:
        rows = db.execute(
            select(SmartCardRepresentativeProduct.product_id).where(
                SmartCardRepresentativeProduct.representative_id == representative_id
            )
        ).scalars().all()
        return [str(x) for x in rows]

    @staticmethod
    def serialize(db: Session, rep: SmartCardRepresentative) -> dict[str, Any]:
        social = None
        if rep.social_links_json:
            try:
                social = json.loads(rep.social_links_json)
            except Exception:
                social = None
        extra = None
        if rep.extra_json:
            try:
                extra = json.loads(rep.extra_json)
            except Exception:
                extra = None
        question_config = None
        if getattr(rep, "question_config_json", None):
            try:
                parsed = json.loads(rep.question_config_json)
                if isinstance(parsed, dict):
                    question_config = parsed
            except Exception:
                question_config = None
        web_url = SmartCardCompanyService.public_web_url(rep.qr_token)
        return {
            "id": rep.id,
            "org_id": rep.org_id,
            "name": rep.name,
            "email": rep.email,
            "website": rep.website,
            "social_links": social,
            "mobile": rep.mobile,
            "landline": rep.landline,
            "extension": rep.extension,
            "notes": rep.notes,
            "extra": extra,
            "question_config": question_config,
            "qr_token": rep.qr_token,
            **qr_style_dict(rep, include_transparent=True),
            "photo_storage_path": rep.photo_storage_path,
            "photo_url": (
                f"/smart-card/representatives/{rep.id}/photo?v={int(rep.updated_at.timestamp()) if rep.updated_at else 0}"
                if rep.photo_storage_path
                else None
            ),
            "qr_image_url": SmartCardCompanyService.qr_image_url(rep),
            "web_url": web_url,
            "status": rep.status,
            "scan_count": int(rep.scan_count or 0),
            "linked_user_id": rep.linked_user_id,
            "invite_id": rep.invite_id,
            "invite_status": SmartCardRepresentativeService.invite_status_for(rep),
            "card_incomplete": SmartCardRepresentativeService.card_incomplete(rep),
            "product_ids": SmartCardRepresentativeService.product_ids(db, representative_id=rep.id),
            "created_at": rep.created_at.isoformat() if rep.created_at else None,
            "updated_at": rep.updated_at.isoformat() if rep.updated_at else None,
        }
