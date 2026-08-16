"""Salesman (Task 8) service: reps, their customers, demo sends, offers, and commission."""

from __future__ import annotations

import csv
import io
import json
import logging
import re
from datetime import datetime, timedelta
from email.utils import parseaddr
from typing import Any

from sqlalchemy import select, update
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.models.billing_invoice import BillingInvoice
from app.models.organisation import Organisation
from app.models.promo_offer import PromoOffer
from app.models.sales_rep import SalesCommission, SalesCustomer, SalesRep
from app.models.user import User

_CODE_RE = re.compile(r"^[A-Z0-9]{4,12}$")
_EMAIL_RE = re.compile(r"^[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}$", re.I)
KIND_SALESMAN = "salesman"
KIND_PARTNER_CHANNEL = "partner_channel"
VALID_KINDS = frozenset({KIND_SALESMAN, KIND_PARTNER_CHANNEL})
DEFAULT_COMMISSION_PCT = 15.0
PARTNER_BULK_OFFER_MAX = 500
PARTNER_BULK_OFFER_MAX_BYTES = 2 * 1024 * 1024
logger = logging.getLogger(__name__)

# Fixed script for the salesman "Call & Survey" demo (matches the 'sales ai survey' agent).
# Salesmen never edit this — warm welcome, 3 questions, gentle "why" if unhappy, warm thanks.
DEMO_AI_SURVEY_SCRIPT = (
    "OPENING (warm welcome): Greet the customer warmly by first name, say this is a quick "
    "friendly survey on behalf of {organisation_name} and it only takes a minute.\n\n"
    "ASK THESE THREE QUESTIONS, ONE AT A TIME, IN ORDER:\n"
    "1. Overall, how would you rate your experience with us today — excellent, good, or poor?\n"
    "2. What did you enjoy most about your experience?\n"
    "3. Is there anything we could do to make it better next time?\n\n"
    "IF THE CUSTOMER IS UNHAPPY OR SAYS \"poor\": stay warm and empathetic, briefly acknowledge it, "
    "and gently ask why so we can improve. Never argue or get defensive.\n\n"
    "CLOSING (warm thanks): Thank the customer sincerely for their time and wish them a great day."
)


class SalesRepError(ValueError):
    pass


class SalesRepService:
    # ---- reps ----------------------------------------------------------------
    @staticmethod
    def normalize_code(raw: str) -> str:
        return re.sub(r"[^A-Za-z0-9]", "", str(raw or "")).upper()

    @staticmethod
    def normalize_kind(raw: str | None) -> str:
        kind = str(raw or KIND_SALESMAN).strip().lower()
        if kind not in VALID_KINDS:
            raise SalesRepError("kind must be 'salesman' or 'partner_channel'.")
        return kind

    @staticmethod
    def normalize_commission_pct(raw: Any) -> float:
        if raw is None or raw == "":
            return DEFAULT_COMMISSION_PCT
        try:
            pct = float(raw)
        except (TypeError, ValueError) as exc:
            raise SalesRepError("Commission % must be a number.") from exc
        if pct < 0 or pct > 100:
            raise SalesRepError("Commission % must be between 0 and 100.")
        return round(pct, 2)

    @staticmethod
    def rep_kind(rep: SalesRep) -> str:
        kind = str(getattr(rep, "kind", None) or KIND_SALESMAN).strip().lower()
        return kind if kind in VALID_KINDS else KIND_SALESMAN

    @staticmethod
    def is_partner_channel(rep: SalesRep) -> bool:
        return SalesRepService.rep_kind(rep) == KIND_PARTNER_CHANNEL

    @staticmethod
    def is_salesman(rep: SalesRep) -> bool:
        return SalesRepService.rep_kind(rep) == KIND_SALESMAN

    @staticmethod
    def commission_pct_of(rep: SalesRep) -> float:
        try:
            pct = float(getattr(rep, "commission_pct", None) or DEFAULT_COMMISSION_PCT)
        except (TypeError, ValueError):
            pct = DEFAULT_COMMISSION_PCT
        return max(0.0, min(100.0, pct))

    @staticmethod
    def apply_commission_pct(base_minor: int, pct: float) -> int:
        return max(0, int(round(int(base_minor or 0) * float(pct) / 100.0)))

    @staticmethod
    def rep_to_dict(rep: SalesRep, user: User | None = None) -> dict[str, Any]:
        from app.services.sales_hub_benefits import (
            benefit_summaries,
            commission_summary,
            currency_of_rep,
            parse_commission_mode,
            parse_commission_tiers,
            parse_one_time_bonus_minor,
            parse_partner_terms,
            parse_promo_benefits,
        )
        from app.services.sales_payout_service import SalesPayoutService

        currency = currency_of_rep(rep)
        benefits = parse_promo_benefits(rep)
        tiers = parse_commission_tiers(rep)
        partner_terms = parse_partner_terms(rep)
        partner = SalesRepService.is_partner_channel(rep)
        return {
            "id": rep.id,
            "user_id": rep.user_id,
            "name": rep.name,
            "company_name": getattr(rep, "company_name", None),
            "mobile": getattr(rep, "mobile", None),
            "kind": SalesRepService.rep_kind(rep),
            "email": user.email if user else None,
            "promo_code": rep.promo_code,
            "country": rep.country,
            "currency": currency,
            "caller_id": rep.caller_id,
            "commission_pct": SalesRepService.commission_pct_of(rep),
            "commission_type": SalesPayoutService.commission_type_of(rep),
            "commission_fixed_minor": SalesPayoutService.fixed_minor_of(rep),
            "commission_mode": parse_commission_mode(rep),
            "one_time_bonus_minor": parse_one_time_bonus_minor(rep),
            "commission_tiers": tiers,
            "commission_summary": commission_summary(
                tiers,
                currency=currency,
                partner=partner,
                partner_terms=partner_terms,
                commission_mode=parse_commission_mode(rep),
                one_time_bonus_minor=parse_one_time_bonus_minor(rep),
            ),
            "promo_benefits": benefits,
            "promo_benefit_summaries": benefit_summaries(benefits, currency=currency),
            "partner_terms": partner_terms,
            "payout": SalesPayoutService.payout_dict(rep),
            "is_active": bool(rep.is_active),
            "created_at": rep.created_at.isoformat() if rep.created_at else None,
            "has_smtp": bool(rep.smtp_host and rep.smtp_username and rep.smtp_password_enc),
            "has_imap": bool(rep.imap_host and rep.imap_username and rep.imap_password_enc),
            "smtp_host": rep.smtp_host or "",
            "smtp_port": int(rep.smtp_port or 587),
            "smtp_use_tls": bool(rep.smtp_use_tls),
            "smtp_username": rep.smtp_username or "",
            "imap_host": rep.imap_host or "",
            "imap_port": int(rep.imap_port or 993),
            "imap_use_ssl": bool(rep.imap_use_ssl),
            "imap_username": rep.imap_username or "",
            "email_signature": rep.email_signature or "",
        }

    @staticmethod
    def apply_mailbox_fields(
        rep: SalesRep,
        mailbox: dict[str, Any] | None,
        *,
        require_password_if_new: bool = False,
        default_host: str = "voxbulk.com",
    ) -> None:
        """Apply SMTP/IMAP mailbox fields onto a sales rep (encrypts passwords).

        When a username is set but neither SMTP nor IMAP password is stored yet,
        a password is required (unless require_password_if_new is False).
        A single password is applied to both SMTP and IMAP when only one is sent.
        """
        if not mailbox:
            return

        from app.core.encryption import get_encryptor

        encryptor = get_encryptor()

        if "smtp_host" in mailbox:
            rep.smtp_host = str(mailbox.get("smtp_host") or "").strip()
        if "smtp_port" in mailbox:
            rep.smtp_port = int(mailbox.get("smtp_port") or 587)
        if "smtp_use_tls" in mailbox:
            rep.smtp_use_tls = bool(mailbox.get("smtp_use_tls"))
        if "smtp_use_ssl" in mailbox:
            rep.smtp_use_ssl = bool(mailbox.get("smtp_use_ssl"))
        if "smtp_username" in mailbox:
            rep.smtp_username = str(mailbox.get("smtp_username") or "").strip()

        if "imap_host" in mailbox:
            rep.imap_host = str(mailbox.get("imap_host") or "").strip()
        if "imap_port" in mailbox:
            rep.imap_port = int(mailbox.get("imap_port") or 993)
        if "imap_use_ssl" in mailbox:
            rep.imap_use_ssl = bool(mailbox.get("imap_use_ssl"))
        if "imap_use_tls" in mailbox:
            rep.imap_use_tls = bool(mailbox.get("imap_use_tls"))
        if "imap_username" in mailbox:
            rep.imap_username = str(mailbox.get("imap_username") or "").strip()

        if "email_signature" in mailbox:
            rep.email_signature = str(mailbox.get("email_signature") or "").strip()

        smtp_pw = str(mailbox.get("smtp_password") or "").strip() if "smtp_password" in mailbox else ""
        imap_pw = str(mailbox.get("imap_password") or "").strip() if "imap_password" in mailbox else ""
        # One mailbox password usually covers both send and receive.
        if smtp_pw and not imap_pw and "imap_password" not in mailbox:
            imap_pw = smtp_pw
        if imap_pw and not smtp_pw and "smtp_password" not in mailbox:
            smtp_pw = imap_pw
        if smtp_pw and not imap_pw:
            imap_pw = smtp_pw
        if imap_pw and not smtp_pw:
            smtp_pw = imap_pw

        if smtp_pw:
            rep.smtp_password_enc = encryptor.encrypt_str(smtp_pw)
        if imap_pw:
            rep.imap_password_enc = encryptor.encrypt_str(imap_pw)

        # Fill shared host defaults when username is present but hosts were left blank.
        username = (rep.smtp_username or rep.imap_username or "").strip()
        if username:
            if not (rep.smtp_username or "").strip():
                rep.smtp_username = username
            if not (rep.imap_username or "").strip():
                rep.imap_username = username
            if not (rep.smtp_host or "").strip():
                rep.smtp_host = default_host
                rep.smtp_port = int(rep.smtp_port or 587)
                if "smtp_use_tls" not in mailbox and "smtp_use_ssl" not in mailbox:
                    rep.smtp_use_tls = True
                    rep.smtp_use_ssl = False
            if not (rep.imap_host or "").strip():
                rep.imap_host = default_host
                rep.imap_port = int(rep.imap_port or 993)
                if "imap_use_ssl" not in mailbox and "imap_use_tls" not in mailbox:
                    rep.imap_use_ssl = True
                    rep.imap_use_tls = False

        wants_mailbox = bool(username)
        if require_password_if_new and wants_mailbox:
            has_smtp_pw = bool(rep.smtp_password_enc)
            has_imap_pw = bool(rep.imap_password_enc)
            if not has_smtp_pw or not has_imap_pw:
                raise SalesRepError(
                    "Mailbox password is required to enable SMTP/IMAP for this salesman. "
                    "Enter the mailbox password and save again."
                )

    @staticmethod
    def get_rep_for_user(db: Session, *, user_id: str) -> SalesRep | None:
        return (
            db.execute(
                select(SalesRep)
                .where(SalesRep.user_id == str(user_id))
                .order_by(SalesRep.created_at.desc())
                .limit(1)
            ).scalar_one_or_none()
        )

    @staticmethod
    def partner_org_for_user(db: Session, *, user_id: str) -> Organisation | None:
        from app.models.membership import OrganisationMembership

        membership = (
            db.execute(
                select(OrganisationMembership)
                .where(OrganisationMembership.user_id == str(user_id))
                .order_by(OrganisationMembership.created_at.asc())
            )
            .scalars()
            .first()
        )
        if membership is None:
            return None
        return db.get(Organisation, membership.org_id)

    @staticmethod
    def reset_partner_org_services_to_defaults(db: Session, org: Organisation) -> None:
        """Partner workspace = normal dashboard service model (no forced all-on).

        - `allowed` inherits Admin platform grants (null override), so modules Admin turned Off stay hidden.
        - `enabled` starts as Interview + Survey (clamped to allowed), same as a typical new org.
        Partners can turn on other Admin-granted modules in Settings like any owner.
        """
        from app.services.org_enabled_services import (
            DEFAULT_ENABLED_SERVICES,
            SERVICE_KEYS,
            any_service_enabled,
            clamp_enabled_to_allowed,
            effective_services,
            serialize_enabled_services,
        )
        from app.services.platform_services_settings_service import get_platform_default_allowed

        # Drop forced all-allowed overrides so Admin Onboarding Services (platform) apply.
        org.allowed_services_json = None
        allowed = get_platform_default_allowed(db)
        enabled = clamp_enabled_to_allowed(allowed, dict(DEFAULT_ENABLED_SERVICES))
        if not any_service_enabled(effective_services(allowed, enabled)):
            enabled = {key: bool(allowed.get(key)) for key in SERVICE_KEYS}
        org.enabled_services_json = serialize_enabled_services(enabled)
        db.add(org)

    @staticmethod
    def reset_all_partner_org_services(db: Session) -> dict[str, Any]:
        """Reset every Partner Channel workspace to default active services."""
        reps = db.execute(select(SalesRep).where(SalesRep.kind == KIND_PARTNER_CHANNEL)).scalars().all()
        reset = 0
        org_ids: list[str] = []
        for rep in reps:
            org = SalesRepService.partner_org_for_user(db, user_id=rep.user_id)
            if org is None:
                continue
            SalesRepService.reset_partner_org_services_to_defaults(db, org)
            reset += 1
            org_ids.append(str(org.id))
        if reset:
            db.commit()
        return {"ok": True, "reset": reset, "org_ids": org_ids}

    @staticmethod
    def list_reps(db: Session, *, kind: str | None = None) -> list[dict[str, Any]]:
        q = select(SalesRep).order_by(SalesRep.created_at.desc())
        if kind:
            q = q.where(SalesRep.kind == SalesRepService.normalize_kind(kind))
        rows = db.execute(q).scalars().all()
        out: list[dict[str, Any]] = []
        for rep in rows:
            user = db.execute(select(User).where(User.id == rep.user_id)).scalar_one_or_none()
            d = SalesRepService.rep_to_dict(rep, user)
            stats = SalesRepService.dashboard_stats(db, rep)
            d["customers"] = stats["wallet"]["active_companies"]
            d["commission_minor"] = stats["wallet"]["commission_minor"]
            out.append(d)
        return out

    @staticmethod
    def create_rep(
        db: Session,
        *,
        email: str,
        password: str,
        name: str,
        promo_code: str,
        country: str | None = None,
        caller_id: str | None = None,
        kind: str = KIND_SALESMAN,
        commission_pct: Any = DEFAULT_COMMISSION_PCT,
        company_name: str | None = None,
        mobile: str | None = None,
        commission_type: Any = None,
        commission_fixed_minor: Any = 0,
        payout: dict[str, Any] | None = None,
        promo_benefits: Any = None,
        commission_tiers: Any = None,
        partner_terms: Any = None,
        commission_mode: Any = None,
        one_time_bonus_minor: Any = None,
        mailbox: dict[str, Any] | None = None,
        assign_existing: bool = False,
    ) -> SalesRep:
        from app.services.sales_hub_benefits import (
            default_commission_tiers,
            default_partner_terms,
            default_promo_benefits,
            set_commission_extras,
            set_commission_tiers,
            set_partner_terms,
            set_promo_benefits,
            sync_rep_currency,
        )
        from app.services.sales_payout_service import (
            COMMISSION_TYPE_FIXED,
            COMMISSION_TYPE_MONTH2,
            COMMISSION_TYPE_PERCENT,
            SalesPayoutService,
        )

        email = str(email or "").strip().lower()
        if not email or "@" not in email:
            raise SalesRepError("A valid email is required.")
        password_clean = str(password or "")
        kind_norm = SalesRepService.normalize_kind(kind)
        pct = SalesRepService.normalize_commission_pct(commission_pct)
        if commission_type is None or str(commission_type).strip() == "":
            ctype = COMMISSION_TYPE_PERCENT if kind_norm == KIND_PARTNER_CHANNEL else COMMISSION_TYPE_MONTH2
        else:
            ctype = SalesPayoutService.normalize_commission_type(commission_type)
        fixed_minor = SalesPayoutService.normalize_fixed_minor(commission_fixed_minor)
        if ctype == COMMISSION_TYPE_FIXED and fixed_minor <= 0:
            raise SalesRepError("Fixed commission amount (GBP pence) is required.")
        code = SalesRepService.normalize_code(promo_code)
        if not _CODE_RE.match(code):
            raise SalesRepError("Promo code must be 4–12 letters/numbers (e.g. UK4F2A).")
        if db.execute(select(SalesRep).where(SalesRep.promo_code == code)).scalar_one_or_none():
            raise SalesRepError(f"Promo code {code} is already in use.")
        try:
            from app.services.promo_offer_service import PromoOfferError, PromoOfferService

            PromoOfferService._assert_code_available(db, code)
        except PromoOfferError as exc:
            raise SalesRepError(str(exc)) from exc

        user = db.execute(select(User).where(User.email == email)).scalar_one_or_none()
        if assign_existing:
            if user is None:
                raise SalesRepError("No registered user found with that email. Create a new account instead.")
            if db.execute(select(SalesRep).where(SalesRep.user_id == user.id)).scalar_one_or_none():
                raise SalesRepError("This user is already a salesman or partner channel account.")
            if password_clean:
                if len(password_clean) < 6:
                    raise SalesRepError("Password must be at least 6 characters.")
                user.password_hash = hash_password(password_clean)
            user.is_active = True
        elif user is None:
            if len(password_clean) < 6:
                raise SalesRepError("Password must be at least 6 characters.")
            user = User(email=email, password_hash=hash_password(password_clean), is_active=True, is_superuser=False)
            db.add(user)
            db.flush()
        else:
            if db.execute(select(SalesRep).where(SalesRep.user_id == user.id)).scalar_one_or_none():
                raise SalesRepError("This user is already a salesman or partner channel account.")
            # Existing user row (e.g. prior signup) — always apply the admin-provided password.
            if len(password_clean) < 6:
                raise SalesRepError("Password must be at least 6 characters (or use Assign existing user).")
            user.password_hash = hash_password(password_clean)
            user.is_active = True

        # Needs an organisation membership so the dashboard login flow issues a token.
        from app.models.membership import OrganisationMembership
        from app.models.organisation import Organisation

        has_membership = db.execute(
            select(OrganisationMembership)
            .where(OrganisationMembership.user_id == user.id)
            .limit(1)
        ).scalars().first()
        display = str(company_name or name or email.split("@")[0]).strip() or email.split("@")[0]
        workspace_label = "Partner Channel" if kind_norm == KIND_PARTNER_CHANNEL else "Sales"
        partner_org: Organisation | None = None
        if not has_membership:
            org = Organisation(name=f"{display} — {workspace_label}", onboarding_state="onboarding_completed")
            db.add(org)
            db.flush()
            db.add(OrganisationMembership(org_id=org.id, user_id=user.id, role="sales"))
            db.flush()
            partner_org = org

        now = datetime.utcnow()
        rep = SalesRep(
            user_id=user.id,
            name=str(name or "").strip() or email.split("@")[0],
            company_name=(str(company_name or "").strip() or None),
            mobile=(str(mobile or "").strip() or None),
            kind=kind_norm,
            promo_code=code,
            country=(str(country or "").strip().upper()[:2] or None),
            caller_id=(str(caller_id or "").strip() or None),
            commission_pct=pct,
            commission_type=ctype,
            commission_fixed_minor=fixed_minor if ctype == COMMISSION_TYPE_FIXED else 0,
            is_active=True,
            created_at=now,
            updated_at=now,
        )
        if payout:
            SalesPayoutService.apply_payout_fields(rep, payout)
        sync_rep_currency(rep)
        if promo_benefits is not None:
            set_promo_benefits(rep, promo_benefits)
        else:
            set_promo_benefits(rep, default_promo_benefits(voucher_enabled=True))
        if commission_tiers is not None:
            set_commission_tiers(rep, commission_tiers)
        else:
            if ctype == COMMISSION_TYPE_MONTH2:
                set_commission_tiers(rep, default_commission_tiers(month2_pct=pct))
            elif ctype == COMMISSION_TYPE_FIXED:
                set_commission_tiers(
                    rep,
                    [
                        {"month": 1, "enabled": True, "kind": "fixed", "value": fixed_minor},
                        *[{"month": m, "enabled": False, "kind": "percent", "value": pct} for m in (2, 3, 4, 5, 6)],
                    ],
                )
            else:
                set_commission_tiers(
                    rep,
                    [
                        {"month": 1, "enabled": True, "kind": "percent", "value": pct},
                        *[{"month": m, "enabled": False, "kind": "percent", "value": pct} for m in (2, 3, 4, 5, 6)],
                    ],
                )
                if kind_norm == KIND_PARTNER_CHANNEL:
                    rep.commission_type = COMMISSION_TYPE_PERCENT
        set_commission_extras(rep, mode=commission_mode, one_time_bonus_minor=one_time_bonus_minor)
        if partner_terms is not None or kind_norm == KIND_PARTNER_CHANNEL:
            set_partner_terms(rep, partner_terms if partner_terms is not None else default_partner_terms())
        # Apply mailbox fields if provided (Salesman Mail)
        if mailbox:
            SalesRepService.apply_mailbox_fields(rep, mailbox, require_password_if_new=True)
        db.add(rep)

        if kind_norm == KIND_PARTNER_CHANNEL:
            try:
                target = partner_org or SalesRepService.partner_org_for_user(db, user_id=user.id)
                if target is not None:
                    SalesRepService.reset_partner_org_services_to_defaults(db, target)
            except Exception:
                logger.exception("Failed to apply partner dashboard service defaults for user %s", user.id)

        db.commit()
        db.refresh(rep)

        try:
            from app.services.promo_offer_service import PromoOfferService

            PromoOfferService.upsert_for_sales_rep(db, rep)
        except Exception:
            logger.exception("Failed to sync promo offer for sales rep %s", rep.id)

        # Demo data is no longer auto-seeded on create. New salesmen start with an empty
        # workspace; seed demo data on demand with scripts/seed_sales_demo.py (./seed-sales-demo.sh).
        return rep

    @staticmethod
    def update_rep(db: Session, *, rep: SalesRep, patch: dict[str, Any]) -> SalesRep:
        from app.services.sales_payout_service import COMMISSION_TYPE_FIXED, SalesPayoutService

        if "name" in patch:
            rep.name = str(patch["name"] or "").strip()
        if "company_name" in patch:
            rep.company_name = (str(patch["company_name"] or "").strip() or None)
        if "mobile" in patch:
            rep.mobile = (str(patch["mobile"] or "").strip() or None)
        if "commission_pct" in patch:
            rep.commission_pct = SalesRepService.normalize_commission_pct(patch["commission_pct"])
        if "commission_type" in patch:
            rep.commission_type = SalesPayoutService.normalize_commission_type(patch["commission_type"])
        if "commission_fixed_minor" in patch:
            rep.commission_fixed_minor = SalesPayoutService.normalize_fixed_minor(patch["commission_fixed_minor"])
        if SalesPayoutService.commission_type_of(rep) == COMMISSION_TYPE_FIXED and SalesPayoutService.fixed_minor_of(rep) <= 0:
            raise SalesRepError("Fixed commission amount (GBP pence) is required.")
        if "country" in patch:
            rep.country = (str(patch["country"] or "").strip().upper()[:2] or None)
            from app.services.sales_hub_benefits import sync_rep_currency

            sync_rep_currency(rep)
        if "promo_benefits" in patch:
            from app.services.sales_hub_benefits import set_promo_benefits

            set_promo_benefits(rep, patch.get("promo_benefits"))
        if "commission_tiers" in patch:
            from app.services.sales_hub_benefits import set_commission_tiers

            set_commission_tiers(rep, patch.get("commission_tiers"))
        if "commission_mode" in patch or "one_time_bonus_minor" in patch:
            from app.services.sales_hub_benefits import set_commission_extras

            set_commission_extras(
                rep,
                mode=patch.get("commission_mode") if "commission_mode" in patch else None,
                one_time_bonus_minor=patch.get("one_time_bonus_minor") if "one_time_bonus_minor" in patch else None,
            )
        if "partner_terms" in patch:
            from app.services.sales_hub_benefits import set_partner_terms

            set_partner_terms(rep, patch.get("partner_terms"))
        if "caller_id" in patch:
            rep.caller_id = (str(patch["caller_id"] or "").strip() or None)
        if "is_active" in patch:
            rep.is_active = bool(patch["is_active"])
        # Mailbox fields (Salesman Mail) — flat keys and/or nested `mailbox` dict
        mailbox_patch: dict[str, Any] = {}
        if isinstance(patch.get("mailbox"), dict):
            mailbox_patch.update(patch["mailbox"])
        for key in (
            "smtp_host",
            "smtp_port",
            "smtp_use_tls",
            "smtp_use_ssl",
            "smtp_username",
            "smtp_password",
            "imap_host",
            "imap_port",
            "imap_use_ssl",
            "imap_use_tls",
            "imap_username",
            "imap_password",
            "email_signature",
        ):
            if key in patch:
                mailbox_patch[key] = patch[key]
        if mailbox_patch:
            SalesRepService.apply_mailbox_fields(rep, mailbox_patch, require_password_if_new=True)
        payout_keys = {
            "payout_method",
            "bank_holder_name",
            "bank_name",
            "bank_sort_code",
            "bank_account_number",
            "bank_address",
            "paypal_email",
        }
        if "payout" in patch and isinstance(patch.get("payout"), dict):
            SalesPayoutService.apply_payout_fields(rep, patch["payout"])
        elif any(k in patch for k in payout_keys):
            SalesPayoutService.apply_payout_fields(rep, {k: patch[k] for k in payout_keys if k in patch})
        if "promo_code" in patch and patch["promo_code"]:
            code = SalesRepService.normalize_code(patch["promo_code"])
            if not _CODE_RE.match(code):
                raise SalesRepError("Promo code must be 4–12 letters/numbers.")
            existing = db.execute(select(SalesRep).where(SalesRep.promo_code == code)).scalar_one_or_none()
            if existing and existing.id != rep.id:
                raise SalesRepError(f"Promo code {code} is already in use.")
            try:
                from app.services.promo_offer_service import PromoOfferError, PromoOfferService

                promo = PromoOfferService._sales_rep_promo_offer(db, rep.id, prefer_code=code)
                PromoOfferService._assert_code_available(db, code, exclude_promo_id=promo.id if promo else None)
            except PromoOfferError as exc:
                raise SalesRepError(str(exc)) from exc
            rep.promo_code = code
        rep.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(rep)
        if any(k in patch for k in ("promo_code", "promo_benefits", "is_active", "name")):
            try:
                from app.services.promo_offer_service import PromoOfferService

                PromoOfferService.upsert_for_sales_rep(db, rep)
            except Exception:
                logger.exception("Failed to sync promo offer for sales rep %s", rep.id)
        return rep

    @staticmethod
    def reset_password(db: Session, *, rep: SalesRep, new_password: str) -> None:
        if len(str(new_password or "")) < 6:
            raise SalesRepError("Password must be at least 6 characters.")
        user = db.execute(select(User).where(User.id == rep.user_id)).scalar_one_or_none()
        if user is None:
            raise SalesRepError("Login user not found for this salesman.")
        user.password_hash = hash_password(new_password)
        user.is_active = True
        db.commit()

    @staticmethod
    def delete_rep(db: Session, *, rep: SalesRep) -> None:
        custs = db.execute(
            select(SalesCustomer).where(SalesCustomer.sales_rep_id == rep.id)
        ).scalars().all()
        for c in custs:
            db.delete(c)
        comms = db.execute(
            select(SalesCommission).where(SalesCommission.sales_rep_id == rep.id)
        ).scalars().all()
        for cm in comms:
            db.delete(cm)
        user = db.execute(select(User).where(User.id == rep.user_id)).scalar_one_or_none()
        db.delete(rep)
        if user is not None:
            # Block dashboard login but keep the user row to preserve any history.
            user.is_active = False
        db.commit()

    # ---- customers -----------------------------------------------------------
    @staticmethod
    def _derive_stage(c: SalesCustomer) -> str:
        """Funnel stage derived from the customer's timestamps/flags (most-advanced wins)."""
        if c.status == "won" or c.org_id:
            return "won"
        if c.interested or c.offer_sent_at:
            return "interested"
        if c.demo_wa_sent_at or c.demo_call_sent_at:
            return "demoed"
        return "lead"

    @staticmethod
    def _timeline(c: SalesCustomer) -> list[dict[str, Any]]:
        """Ordered funnel events with timestamps (None = not reached yet)."""

        def iso(dt: datetime | None) -> str | None:
            return dt.isoformat() if dt else None

        demo_at = c.demo_wa_sent_at or c.demo_call_sent_at
        return [
            {"key": "added", "label": "Added / visited", "at": iso(c.created_at)},
            {"key": "demoed", "label": "Demoed", "at": iso(demo_at)},
            {"key": "interested", "label": "Interested (offer sent)", "at": iso(c.interested_at or c.offer_sent_at)},
            {"key": "won", "label": "Signed up / won", "at": iso(c.updated_at) if (c.status == "won" or c.org_id) else None},
        ]

    @staticmethod
    def customer_to_dict(c: SalesCustomer) -> dict[str, Any]:
        return {
            "id": c.id,
            "full_name": c.full_name,
            "company_name": c.company_name,
            "address": c.address,
            "city": c.city,
            "country": c.country,
            "mobile": c.mobile,
            "email": c.email,
            "business_type": c.business_type,
            "branches": c.branches,
            "contact_person": c.contact_person,
            "org_id": c.org_id,
            "offer_details": c.offer_details,
            "offer_sent_at": c.offer_sent_at.isoformat() if c.offer_sent_at else None,
            "demo_wa_sent_at": c.demo_wa_sent_at.isoformat() if c.demo_wa_sent_at else None,
            "demo_call_sent_at": c.demo_call_sent_at.isoformat() if c.demo_call_sent_at else None,
            "interested": bool(c.interested),
            "interested_at": c.interested_at.isoformat() if c.interested_at else None,
            "status": c.status,
            "stage": SalesRepService._derive_stage(c),
            "created_at": c.created_at.isoformat() if c.created_at else None,
        }

    @staticmethod
    def get_customer_detail(db: Session, *, rep_id: str, customer_id: str) -> dict[str, Any] | None:
        cust = SalesRepService.get_customer(db, rep_id=rep_id, customer_id=customer_id)
        if cust is None:
            return None
        data = SalesRepService.customer_to_dict(cust)
        data["timeline"] = SalesRepService._timeline(cust)
        return data

    @staticmethod
    def list_customers(db: Session, *, rep_id: str) -> list[dict[str, Any]]:
        rows = (
            db.execute(
                select(SalesCustomer)
                .where(SalesCustomer.sales_rep_id == str(rep_id))
                .order_by(SalesCustomer.created_at.desc())
            )
            .scalars()
            .all()
        )
        return [SalesRepService.customer_to_dict(c) for c in rows]

    @staticmethod
    def get_customer(db: Session, *, rep_id: str, customer_id: str) -> SalesCustomer | None:
        return db.execute(
            select(SalesCustomer).where(
                SalesCustomer.id == str(customer_id), SalesCustomer.sales_rep_id == str(rep_id)
            )
        ).scalar_one_or_none()

    @staticmethod
    def upsert_customer(db: Session, *, rep_id: str, payload: dict[str, Any]) -> SalesCustomer:
        cid = str(payload.get("id") or "").strip()
        now = datetime.utcnow()
        cust = None
        if cid:
            cust = SalesRepService.get_customer(db, rep_id=rep_id, customer_id=cid)
            if cust is None:
                raise SalesRepError("Customer not found.")
        if cust is None:
            cust = SalesCustomer(sales_rep_id=str(rep_id), created_at=now)
            db.add(cust)
        for field in (
            "full_name",
            "company_name",
            "address",
            "city",
            "country",
            "mobile",
            "email",
            "business_type",
            "contact_person",
            "status",
            "offer_details",
        ):
            if field in payload:
                setattr(cust, field, (str(payload[field]).strip() if payload[field] is not None else None))
        if "branches" in payload:
            try:
                cust.branches = max(0, int(payload["branches"]))
            except (TypeError, ValueError):
                cust.branches = 1
        cust.updated_at = now
        db.commit()
        db.refresh(cust)
        # Auto-create mail contact if email is present (Salesman Mail)
        if cust.email:
            try:
                from app.models.sales_mail import SalesMailContact
                existing = db.execute(
                    select(SalesMailContact).where(
                        SalesMailContact.sales_rep_id == str(rep_id),
                        SalesMailContact.email == cust.email.strip().lower()
                    )
                ).scalar_one_or_none()
                if not existing:
                    import uuid
                    contact = SalesMailContact(
                        id=str(uuid.uuid4()),
                        sales_rep_id=str(rep_id),
                        sales_customer_id=cust.id,
                        email=cust.email.strip().lower(),
                        name=cust.full_name,
                        company=cust.company_name,
                        created_at=now,
                        updated_at=now,
                    )
                    db.add(contact)
                    db.commit()
            except Exception:
                logger.exception("Failed to auto-create mail contact for customer %s", cust.id)
        return cust

    @staticmethod
    def delete_customer(db: Session, *, rep_id: str, customer_id: str) -> None:
        cust = SalesRepService.get_customer(db, rep_id=rep_id, customer_id=customer_id)
        if cust is None:
            raise SalesRepError("Customer not found.")
        # Commissions + mail contacts FK to sales_customers — unlink before delete (MySQL 1451).
        db.execute(
            update(SalesCommission)
            .where(SalesCommission.sales_customer_id == customer_id)
            .values(sales_customer_id=None)
        )
        try:
            from app.models.sales_mail import SalesMailContact

            db.execute(
                update(SalesMailContact)
                .where(SalesMailContact.sales_customer_id == customer_id)
                .values(sales_customer_id=None)
            )
        except Exception:
            logger.exception("Failed to unlink mail contacts before deleting customer %s", customer_id)
        db.delete(cust)
        db.commit()

    # ---- demo / offer sends (best-effort, never crash the request) -----------
    @staticmethod
    def _telnyx_config(db: Session) -> dict[str, Any]:
        from app.services.telnyx_messaging_service import TelnyxMessagingService

        return TelnyxMessagingService._config(db)

    @staticmethod
    def send_offer(db: Session, *, rep: SalesRep, customer: SalesCustomer, channel: str, offer_details: str) -> dict[str, Any]:
        from app.services.promo_offer_service import PromoOfferService
        from app.services.sales_offer_send_service import SalesOfferSendService

        offer_details = str(offer_details or "").strip() or "Special VoxBulk offer"
        customer.offer_details = offer_details
        try:
            PromoOfferService.upsert_for_sales_rep(db, rep)
        except Exception:
            logger.exception("Promo sync before send_offer failed rep=%s", rep.id)
        promo = PromoOfferService.get_by_code(db, rep.promo_code)
        signup_url = PromoOfferService.signup_url(rep.promo_code)
        first_name = SalesOfferSendService._first_name(customer.full_name)
        from app.services.billing_currency import money_display
        from app.services.sales_hub_benefits import currency_of_rep, parse_promo_benefits

        benefits = parse_promo_benefits(rep)
        wv = benefits.get("wallet_voucher") or {}
        if wv.get("enabled") and int(wv.get("amount_minor") or 0) > 0:
            credit_disp = money_display(int(wv["amount_minor"]), currency_of_rep(rep))
            offer_line = f"{credit_disp} welcome credit"
            offer_summary = (
                f"{offer_details}. Includes {credit_disp} wallet credit after signup "
                "(not usable for campaign launches or Customer feedback promo sends)."
            )
        else:
            offer_line = "Special welcome offer"
            offer_summary = f"{offer_details}."
        promo_name = promo.name if promo else f"{rep.promo_code} — VoxBulk offer"
        variables = {
            "first_name": first_name,
            "offer_line": offer_line,
            "promo_name": promo_name,
            "offer_summary": offer_summary,
            "signup_url": signup_url,
            "promo_code": rep.promo_code,
        }
        log: dict[str, Any] = {}
        ok = False
        if channel == "email":
            if not customer.email:
                return {"ok": False, "message": "Customer has no email."}
            try:
                from app.services.transactional_email_service import TransactionalEmailService

                sent, err = TransactionalEmailService.send_templated_optional(
                    db,
                    template_key="sales_offer",
                    to_email=customer.email,
                    variables=variables,
                )
                ok = bool(sent)
                log = {"channel": "email", "ok": ok, "error": err}
            except Exception as e:  # noqa: BLE001
                log = {"channel": "email", "ok": False, "error": str(e)}
        elif channel == "wa":
            if not customer.mobile:
                return {"ok": False, "message": "Customer has no mobile number."}
            try:
                # Must use approved Meta/Telnyx marketing template (plain text fails outside 24h window).
                from app.services.sales_whatsapp_send_service import send_sales_whatsapp

                str_vars = {k: str(v) for k, v in variables.items()}
                plain = SalesOfferSendService._whatsapp_body(db, str_vars)
                res = send_sales_whatsapp(
                    db,
                    to_number=customer.mobile,
                    template_key="sales_offer",
                    body=plain,
                    variables=str_vars,
                )
                ok = bool(getattr(res, "ok", False))
                log = {
                    "channel": "wa",
                    "ok": ok,
                    "template_key": "sales_offer",
                    "message_id": getattr(res, "external_id", None),
                    "error": None if ok else (getattr(res, "detail", None) or getattr(res, "status", None)),
                }
            except Exception as e:  # noqa: BLE001
                log = {"channel": "wa", "ok": False, "error": str(e)}
        else:
            return {"ok": False, "message": "Unknown channel."}

        if ok:
            now = datetime.utcnow()
            customer.offer_sent_at = now
            # Sending an offer means the customer is interested.
            customer.interested = True
            if customer.interested_at is None:
                customer.interested_at = now
            if customer.status not in ("won", "interested"):
                customer.status = "interested"
        customer.offer_log_json = json.dumps(log)
        customer.updated_at = datetime.utcnow()
        db.commit()
        return {"ok": ok, "message": "Sent." if ok else f"Send failed: {log.get('error') or 'unknown error'}"}

    @staticmethod
    def normalize_offer_email(raw: Any) -> str | None:
        text = str(raw or "").strip()
        if not text:
            return None
        _name, addr = parseaddr(text)
        email = (addr or text).strip().lower()
        if not _EMAIL_RE.match(email):
            return None
        return email

    @staticmethod
    def partner_offer_template_xlsx() -> bytes:
        try:
            import openpyxl
        except ImportError as exc:
            raise SalesRepError("Excel support requires openpyxl on the server.") from exc
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(["email", "name"])
        ws.append(["alex@example.com", "Alex Example"])
        ws.append(["sam@example.com", "Sam Example"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def parse_partner_offer_recipients(content: bytes, filename: str = "") -> dict[str, Any]:
        if not content:
            raise SalesRepError("Empty file.")
        if len(content) > PARTNER_BULK_OFFER_MAX_BYTES:
            raise SalesRepError("File too large (max 2MB).")

        name = str(filename or "").lower()
        rows: list[dict[str, Any]] = []
        if name.endswith((".xlsx", ".xls", ".xlsm")) or content[:2] == b"PK":
            try:
                import openpyxl
            except ImportError as exc:
                raise SalesRepError("Excel support requires openpyxl on the server.") from exc
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            except Exception as exc:  # noqa: BLE001
                raise SalesRepError("Could not read Excel file.") from exc
            ws = wb.active
            header: list[str] | None = None
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = [("" if v is None else str(v).strip()) for v in row]
                if not any(values):
                    continue
                if header is None:
                    lowered = [v.lower() for v in values]
                    if any("email" in v or "e-mail" in v or v == "mail" for v in lowered):
                        header = lowered
                        continue
                    # No header — treat first non-empty cell as email
                    email = SalesRepService.normalize_offer_email(values[0] if values else "")
                    if email:
                        rows.append(
                            {
                                "email": email,
                                "name": values[1] if len(values) > 1 and values[1] else None,
                                "row": i,
                            }
                        )
                    continue
                email_idx = next(
                    (idx for idx, h in enumerate(header) if "email" in h or "e-mail" in h or h == "mail"),
                    0,
                )
                name_idx = next((idx for idx, h in enumerate(header) if h in ("name", "full_name", "fullname")), None)
                email = SalesRepService.normalize_offer_email(values[email_idx] if email_idx < len(values) else "")
                if not email:
                    continue
                contact_name = None
                if name_idx is not None and name_idx < len(values) and values[name_idx]:
                    contact_name = values[name_idx]
                rows.append({"email": email, "name": contact_name, "row": i})
        else:
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("latin-1", errors="replace")
            reader = csv.reader(io.StringIO(text))
            header = None
            for i, values in enumerate(reader, start=1):
                values = [str(v or "").strip() for v in values]
                if not any(values):
                    continue
                if header is None:
                    lowered = [v.lower() for v in values]
                    if any("email" in v or "e-mail" in v or v == "mail" for v in lowered):
                        header = lowered
                        continue
                    email = SalesRepService.normalize_offer_email(values[0] if values else "")
                    if email:
                        rows.append(
                            {
                                "email": email,
                                "name": values[1] if len(values) > 1 and values[1] else None,
                                "row": i,
                            }
                        )
                    continue
                email_idx = next(
                    (idx for idx, h in enumerate(header) if "email" in h or "e-mail" in h or h == "mail"),
                    0,
                )
                name_idx = next((idx for idx, h in enumerate(header) if h in ("name", "full_name", "fullname")), None)
                email = SalesRepService.normalize_offer_email(values[email_idx] if email_idx < len(values) else "")
                if not email:
                    continue
                contact_name = None
                if name_idx is not None and name_idx < len(values) and values[name_idx]:
                    contact_name = values[name_idx]
                rows.append({"email": email, "name": contact_name, "row": i})

        deduped: list[dict[str, Any]] = []
        seen: set[str] = set()
        skipped = 0
        for row in rows:
            email = row["email"]
            if email in seen:
                skipped += 1
                continue
            seen.add(email)
            deduped.append(row)

        if len(deduped) > PARTNER_BULK_OFFER_MAX:
            raise SalesRepError(f"Too many recipients (max {PARTNER_BULK_OFFER_MAX}).")

        return {
            "ok": True,
            "count": len(deduped),
            "skipped": skipped,
            "recipients": deduped,
            "message": None if deduped else "No valid emails found in this file.",
        }

    @staticmethod
    def send_bulk_partner_offers(
        db: Session,
        *,
        rep: SalesRep,
        recipients: list[dict[str, Any]],
        offer_details: str,
    ) -> dict[str, Any]:
        if not SalesRepService.is_partner_channel(rep):
            raise SalesRepError("Bulk partner offers are only available to Partner Channel accounts.")
        if not recipients:
            raise SalesRepError("No recipients to send.")
        if len(recipients) > PARTNER_BULK_OFFER_MAX:
            raise SalesRepError(f"Too many recipients (max {PARTNER_BULK_OFFER_MAX}).")

        from app.services.promo_offer_service import PromoOfferService
        from app.services.sales_offer_send_service import SalesOfferSendService
        from app.services.transactional_email_service import TransactionalEmailService

        offer_details = str(offer_details or "").strip() or "Special VoxBulk partner offer"
        try:
            PromoOfferService.upsert_for_sales_rep(db, rep)
        except Exception:
            logger.exception("Promo sync before partner bulk offer failed rep=%s", rep.id)
        promo = PromoOfferService.get_by_code(db, rep.promo_code)
        signup_url = PromoOfferService.signup_url(rep.promo_code)
        promo_name = promo.name if promo else f"{rep.promo_code} — VoxBulk offer"
        from app.services.billing_currency import money_display
        from app.services.sales_hub_benefits import currency_of_rep, parse_promo_benefits

        benefits = parse_promo_benefits(rep)
        wv = benefits.get("wallet_voucher") or {}
        if wv.get("enabled") and int(wv.get("amount_minor") or 0) > 0:
            credit_disp = money_display(int(wv["amount_minor"]), currency_of_rep(rep))
            offer_line = f"{credit_disp} welcome credit"
            credit_summary_suffix = (
                f". Includes {credit_disp} wallet credit after signup "
                "(not usable for campaign launches or Customer feedback promo sends)."
            )
        else:
            offer_line = "Special welcome offer"
            credit_summary_suffix = "."

        results: list[dict[str, Any]] = []
        sent = 0
        failed = 0
        for item in recipients:
            email = SalesRepService.normalize_offer_email(item.get("email"))
            if not email:
                failed += 1
                results.append({"email": str(item.get("email") or ""), "name": item.get("name"), "ok": False, "error": "Invalid email"})
                continue
            name = str(item.get("name") or "").strip() or None
            first_name = SalesOfferSendService._first_name(name)
            offer_summary = f"{offer_details}{credit_summary_suffix}"
            variables = {
                "first_name": first_name,
                "offer_line": offer_line,
                "promo_name": promo_name,
                "offer_summary": offer_summary,
                "signup_url": signup_url,
                "promo_code": rep.promo_code,
            }
            try:
                ok, err = TransactionalEmailService.send_templated_optional(
                    db,
                    template_key="sales_offer",
                    to_email=email,
                    variables=variables,
                )
                if ok:
                    sent += 1
                    results.append({"email": email, "name": name, "ok": True, "error": None})
                else:
                    failed += 1
                    results.append({"email": email, "name": name, "ok": False, "error": err or "Send failed"})
            except Exception as exc:  # noqa: BLE001
                failed += 1
                results.append({"email": email, "name": name, "ok": False, "error": str(exc)})

        return {
            "ok": sent > 0,
            "sent": sent,
            "failed": failed,
            "total": len(recipients),
            "results": results,
            "message": f"Sent {sent} of {len(recipients)}." if sent else "All sends failed.",
        }

    @staticmethod
    def _mark_demoed(db: Session, customer: SalesCustomer, *, channel: str) -> None:
        """Record a demo send on the customer and advance the funnel stage."""
        now = datetime.utcnow()
        if channel == "wa":
            customer.demo_wa_sent_at = now
        elif channel == "call":
            customer.demo_call_sent_at = now
        if customer.status == "lead":
            customer.status = "demoed"
        customer.updated_at = now
        db.commit()

    @staticmethod
    def send_demo_wa(db: Session, *, customer: SalesCustomer) -> dict[str, Any]:
        if not customer.mobile:
            return {"ok": False, "message": "Customer has no mobile number."}
        try:
            from app.services.telnyx_messaging_service import TelnyxMessagingService

            body = "Hi! This is a quick VoxBulk demo survey — how would you rate your last visit? Reply 1-5."
            res = TelnyxMessagingService.send_whatsapp(
                db, to_number=customer.mobile, body=body, service_code="marketing"
            )
            ok = bool(getattr(res, "ok", True))
            if ok:
                SalesRepService._mark_demoed(db, customer, channel="wa")
            return {"ok": ok, "message": "Demo WhatsApp survey sent."}
        except Exception as e:  # noqa: BLE001
            return {"ok": False, "message": f"WhatsApp not available: {e}"}

    @staticmethod
    def _rep_workspace_org_id(db: Session, rep: SalesRep) -> str | None:
        from app.models.membership import OrganisationMembership

        m = db.execute(
            select(OrganisationMembership)
            .where(OrganisationMembership.user_id == rep.user_id)
            .order_by(OrganisationMembership.created_at.asc())
            .limit(1)
        ).scalars().first()
        return str(m.org_id) if m is not None else None

    @staticmethod
    def send_demo_call(db: Session, *, rep: SalesRep, customer: SalesCustomer) -> dict[str, Any]:
        """Run a live 3-question AI survey demo to the customer's number.

        Reuses the proven outbound survey pipeline (Telnyx assistant + dispatch) so the
        transcript and answers show up in the salesman's /surveys/results. Billing/eligibility
        is bypassed (demo=True); the number must still pass the Telnyx phone allowlist.
        """
        if not customer.mobile:
            return {"ok": False, "message": "Customer has no mobile number."}

        org_id = SalesRepService._rep_workspace_org_id(db, rep)
        if not org_id:
            return {"ok": False, "message": "Salesman has no workspace organisation."}

        try:
            from app.models.agent import AgentDefinition
            from app.services.platform_catalog_service import ServiceOrderService
            from app.services.survey_call_dispatch_service import SurveyCallDispatchService

            org_name = (
                str(customer.company_name or "").strip()
                or str(customer.full_name or "").strip()
                or "VoxBulk"
            )

            config: dict[str, Any] = {
                "survey_channel": "ai_call",
                "delivery": "ai_call",
                "demo": True,
                "script_approved": True,
                "organisation_name": org_name,
                "clinic_name": org_name,
                "survey_organiser_name": org_name,
                "approved_script": DEMO_AI_SURVEY_SCRIPT.replace("{organisation_name}", org_name),
            }
            agent = db.execute(
                select(AgentDefinition).where(AgentDefinition.slug == "sales-ai-survey")
            ).scalar_one_or_none()
            if agent is not None:
                config["agent_id"] = agent.id

            order = ServiceOrderService.create_order(
                db,
                org_id=org_id,
                user_id=str(rep.user_id),
                service_code="survey",
                title=f"AI Survey Demo · {customer.company_name or customer.full_name or customer.mobile}",
                config=config,
            )
            ServiceOrderService.replace_recipients(
                db,
                order,
                [{"name": customer.full_name or customer.company_name or "Customer", "phone": customer.mobile}],
            )

            now = datetime.utcnow()
            order.status = "running"
            order.payment_status = "approved"
            order.started_at = order.started_at or now
            order.updated_at = now
            db.add(order)
            db.commit()
            db.refresh(order)

            recipients = ServiceOrderService.get_recipients(db, order.id)
            if not recipients:
                return {"ok": False, "message": "Could not prepare the demo call recipient."}

            SurveyCallDispatchService.dial_recipient(db, order, recipients[0])
            SalesRepService._mark_demoed(db, customer, channel="call")
            return {
                "ok": True,
                "message": "AI survey demo call started — watch the transcript in Surveys results.",
                "order_id": order.id,
            }
        except Exception as e:  # noqa: BLE001
            return {"ok": False, "message": f"Could not start the AI survey demo: {e}"}

    @staticmethod
    def link_customer_on_promo_redeem(
        db: Session,
        *,
        promo,
        org: Organisation,
        user_email: str | None,
    ) -> None:
        from app.models.organisation import Organisation as OrgModel

        if not isinstance(org, OrgModel):
            org = db.get(OrgModel, str(getattr(org, "id", org)))
        if org is None:
            return
        rep_id = getattr(promo, "sales_rep_id", None)
        if not rep_id:
            row = db.execute(
                select(SalesRep).where(SalesRep.promo_code == str(getattr(promo, "code", "") or ""))
            ).scalar_one_or_none()
            rep_id = row.id if row else None
        if not rep_id:
            return
        email_norm = str(user_email or org.contact_email or "").strip().lower()
        mobile = str(org.contact_phone or "").strip()
        q = select(SalesCustomer).where(SalesCustomer.sales_rep_id == rep_id)
        candidates = list(db.execute(q).scalars().all())
        matched: SalesCustomer | None = None
        for cust in candidates:
            cust_email = str(cust.email or "").strip().lower()
            cust_mobile = str(cust.mobile or "").strip()
            if email_norm and cust_email and cust_email == email_norm:
                matched = cust
                break
            if mobile and cust_mobile and cust_mobile.replace(" ", "") == mobile.replace(" ", ""):
                matched = cust
                break
        now = datetime.utcnow()
        if matched is None and email_norm:
            matched = SalesCustomer(
                sales_rep_id=rep_id,
                full_name=str(org.contact_name or org.name or "Customer"),
                company_name=org.name,
                email=email_norm or None,
                mobile=mobile or None,
                status="won",
                org_id=org.id,
                interested=True,
                interested_at=now,
                created_at=now,
                updated_at=now,
            )
            db.add(matched)
        elif matched is not None:
            matched.org_id = org.id
            matched.status = "won"
            matched.interested = True
            if matched.interested_at is None:
                matched.interested_at = now
            matched.updated_at = now
            db.add(matched)
        db.flush()

    @staticmethod
    def mark_commission_paid(db: Session, *, commission_id: str, note: str | None = None) -> SalesCommission:
        row = db.get(SalesCommission, commission_id)
        if row is None:
            raise SalesRepError("Commission not found")
        if str(row.status or "").lower() == "paid":
            return row
        row.status = "paid"
        if note:
            row.note = str(note).strip()[:255]
        row.updated_at = datetime.utcnow()
        db.add(row)
        db.commit()
        db.refresh(row)
        return row

    @staticmethod
    def mark_rep_commissions_paid(db: Session, *, rep_id: str, commission_ids: list[str] | None = None) -> dict[str, Any]:
        q = select(SalesCommission).where(
            SalesCommission.sales_rep_id == rep_id,
            SalesCommission.status == "pending",
        )
        if commission_ids:
            q = q.where(SalesCommission.id.in_(commission_ids))
        rows = list(db.execute(q).scalars().all())
        now = datetime.utcnow()
        total = 0
        for row in rows:
            row.status = "paid"
            row.updated_at = now
            total += int(row.amount_minor or 0)
            db.add(row)
        db.commit()
        return {"ok": True, "marked_paid": len(rows), "amount_minor": total}

    # ---- commission + dashboard ---------------------------------------------
    @staticmethod
    def _linked_org_ids(db: Session, rep: SalesRep) -> set[str]:
        """Orgs attributed to this rep: customers converted to orgs, or signed up via the rep promo code."""
        org_ids: set[str] = set()
        rows = db.execute(
            select(SalesCustomer.org_id).where(
                SalesCustomer.sales_rep_id == rep.id, SalesCustomer.org_id.isnot(None)
            )
        ).scalars().all()
        org_ids.update(str(x) for x in rows if x)
        try:
            from app.models.org_usage_period import OrgUsagePeriod

            via_code = db.execute(
                select(OrgUsagePeriod.org_id).where(OrgUsagePeriod.promo_code == rep.promo_code)
            ).scalars().all()
            org_ids.update(str(x) for x in via_code if x)
        except Exception:  # noqa: BLE001
            pass
        return org_ids

    @staticmethod
    def get_rep_for_org(db: Session, *, org_id: str) -> SalesRep | None:
        """Reverse of _linked_org_ids: find the salesman attributed to an org."""
        org_id = str(org_id or "")
        if not org_id:
            return None
        cust = db.execute(
            select(SalesCustomer).where(SalesCustomer.org_id == org_id).order_by(SalesCustomer.created_at.asc())
        ).scalars().first()
        if cust is not None:
            rep = db.get(SalesRep, cust.sales_rep_id)
            if rep is not None:
                return rep
        try:
            from app.models.org_usage_period import OrgUsagePeriod

            code = db.execute(
                select(OrgUsagePeriod.promo_code).where(
                    OrgUsagePeriod.org_id == org_id, OrgUsagePeriod.promo_code.isnot(None)
                )
            ).scalars().first()
            if code:
                return db.execute(
                    select(SalesRep).where(SalesRep.promo_code == SalesRepService.normalize_code(code))
                ).scalar_one_or_none()
        except Exception:  # noqa: BLE001
            pass
        return None

    @staticmethod
    def accrue_commission_for_paid_invoice(
        db: Session, invoice: BillingInvoice, *, force_subscription: bool = False
    ) -> SalesCommission | None:
        """Best-effort: accrue commission when a linked org pays a subscription invoice.

        Driven by commission_tiers_json when present:
          - Monthly: Nth paid subscription invoice → month N tier (2/3/4) if enabled
          - Yearly: month-2 tier only once (base = invoice/12)
        Legacy commission_type percent/fixed still accrues every paid invoice.

        Never raises. Pass force_subscription=True for known subscription payments whose
        invoice row is not tagged kind="subscription" (e.g. GoCardless DD webhook).
        """
        try:
            from app.services.sales_hub_benefits import currency_of_rep, parse_commission_tiers
            from app.services.sales_payout_service import (
                COMMISSION_TYPE_FIXED,
                COMMISSION_TYPE_MONTH2,
                COMMISSION_TYPE_PERCENT,
                SalesPayoutService,
            )

            if str(getattr(invoice, "status", "") or "").lower() != "paid":
                return None
            if not force_subscription and str(getattr(invoice, "kind", "") or "").lower() != "subscription":
                return None
            org_id = str(getattr(invoice, "org_id", "") or "")
            if not org_id:
                return None
            rep = SalesRepService.get_rep_for_org(db, org_id=org_id)
            if rep is None or not rep.is_active:
                return None

            amount = int(getattr(invoice, "amount_gbp_pence", 0) or 0)
            currency = str(getattr(invoice, "currency", "") or "GBP") or "GBP"
            invoice_id = getattr(invoice, "id", None)
            pct = SalesRepService.commission_pct_of(rep)
            ctype = SalesPayoutService.commission_type_of(rep)
            tiers = parse_commission_tiers(rep)
            has_tier_json = bool(getattr(rep, "commission_tiers_json", None))

            if ctype in (COMMISSION_TYPE_PERCENT, COMMISSION_TYPE_FIXED) and not (
                has_tier_json and ctype == COMMISSION_TYPE_MONTH2
            ):
                # Partner / on-payment styles
                if ctype in (COMMISSION_TYPE_PERCENT, COMMISSION_TYPE_FIXED):
                    if invoice_id:
                        already = db.execute(
                            select(SalesCommission).where(
                                SalesCommission.sales_rep_id == rep.id,
                                SalesCommission.invoice_id == invoice_id,
                            )
                        ).scalar_one_or_none()
                        if already is not None:
                            return None
                    if ctype == COMMISSION_TYPE_FIXED:
                        kind = "fixed_invoice"
                        commission_minor = SalesPayoutService.fixed_minor_of(rep)
                        note = f"Fixed commission {SalesPayoutService.format_gbp(commission_minor)} on paid invoice."
                    else:
                        kind = "percent_invoice" if not SalesRepService.is_partner_channel(rep) else "partner_invoice"
                        commission_minor = SalesRepService.apply_commission_pct(amount, pct)
                        note = f"{pct:g}% of paid subscription invoice."
                else:
                    return None
            else:
                # month2 + multi-month tiers (months 1–6)
                from app.services.sales_hub_benefits import (
                    COMMISSION_MONTHS,
                    parse_commission_mode,
                    parse_one_time_bonus_minor,
                )

                mode = parse_commission_mode(rep)
                bonus_minor = parse_one_time_bonus_minor(rep)
                interval = SalesRepService._org_plan_interval(db, org_id)

                def _already_one_time() -> bool:
                    return (
                        db.execute(
                            select(SalesCommission).where(
                                SalesCommission.sales_rep_id == rep.id,
                                SalesCommission.org_id == org_id,
                                SalesCommission.kind == "one_time_bonus",
                            )
                        ).scalar_one_or_none()
                        is not None
                    )

                def _add_one_time_if_needed(link_cust_id: str | None) -> SalesCommission | None:
                    if mode == "commission_only" or bonus_minor <= 0:
                        return None
                    if _already_one_time():
                        return None
                    if currency.upper() != currency_of_rep(rep).upper():
                        return None
                    bonus = SalesCommission(
                        sales_rep_id=rep.id,
                        sales_customer_id=link_cust_id,
                        org_id=org_id,
                        invoice_id=invoice_id,
                        amount_minor=bonus_minor,
                        currency=currency,
                        kind="one_time_bonus",
                        status="pending",
                        note=f"One-time bonus {SalesPayoutService.format_gbp(bonus_minor)}.",
                    )
                    db.add(bonus)
                    return bonus

                if mode == "one_time_only":
                    link_cust = db.execute(
                        select(SalesCustomer).where(
                            SalesCustomer.sales_rep_id == rep.id, SalesCustomer.org_id == org_id
                        )
                    ).scalars().first()
                    bonus = _add_one_time_if_needed(link_cust.id if link_cust else None)
                    if bonus is None:
                        return None
                    db.commit()
                    db.refresh(bonus)
                    return bonus

                if interval == "yearly":
                    existing = db.execute(
                        select(SalesCommission).where(
                            SalesCommission.sales_rep_id == rep.id,
                            SalesCommission.org_id == org_id,
                            SalesCommission.kind.in_(
                                ["yearly_1mo", "monthly_1st", "monthly_2nd", "monthly_3rd", "monthly_4th", "monthly_5th", "monthly_6th"]
                            ),
                        )
                    ).scalar_one_or_none()
                    if existing is not None:
                        return None
                    tier = next((t for t in tiers if int(t["month"]) == 2 and t.get("enabled")), None)
                    if tier is None:
                        # Fall back to first enabled month for yearly
                        tier = next((t for t in tiers if t.get("enabled")), None)
                    if tier is None:
                        return None
                    kind = "yearly_1mo"
                    base_minor = max(0, round(amount / 12))
                    if tier["kind"] == "fixed":
                        rep_cur = currency_of_rep(rep)
                        if currency.upper() != rep_cur.upper():
                            return None
                        commission_minor = int(round(float(tier["value"])))
                        note = f"Fixed yearly commission on one month of yearly plan."
                    else:
                        commission_minor = SalesRepService.apply_commission_pct(base_minor, float(tier["value"]))
                        note = f"{float(tier['value']):g}% of one month of a yearly plan."
                else:
                    paid_invoices = db.execute(
                        select(BillingInvoice).where(
                            BillingInvoice.org_id == org_id,
                            BillingInvoice.kind == "subscription",
                            BillingInvoice.status == "paid",
                        )
                    ).scalars().all()
                    paid_n = len(paid_invoices)
                    if paid_n not in COMMISSION_MONTHS:
                        return None
                    tier = next((t for t in tiers if int(t["month"]) == paid_n and t.get("enabled")), None)
                    if tier is None:
                        return None
                    kind_map = {
                        1: "monthly_1st",
                        2: "monthly_2nd",
                        3: "monthly_3rd",
                        4: "monthly_4th",
                        5: "monthly_5th",
                        6: "monthly_6th",
                    }
                    kind = kind_map[paid_n]
                    existing = db.execute(
                        select(SalesCommission).where(
                            SalesCommission.sales_rep_id == rep.id,
                            SalesCommission.org_id == org_id,
                            SalesCommission.kind == kind,
                        )
                    ).scalar_one_or_none()
                    if existing is not None:
                        return None
                    if invoice_id:
                        already = db.execute(
                            select(SalesCommission).where(
                                SalesCommission.sales_rep_id == rep.id,
                                SalesCommission.invoice_id == invoice_id,
                                SalesCommission.kind != "one_time_bonus",
                            )
                        ).scalar_one_or_none()
                        if already is not None:
                            return None
                    if tier["kind"] == "fixed":
                        rep_cur = currency_of_rep(rep)
                        if currency.upper() != rep_cur.upper():
                            return None
                        commission_minor = int(round(float(tier["value"])))
                        note = f"Fixed commission on month {paid_n} subscription."
                    else:
                        commission_minor = SalesRepService.apply_commission_pct(amount, float(tier["value"]))
                        note = f"{float(tier['value']):g}% of month {paid_n} subscription."
                _ = COMMISSION_TYPE_MONTH2

            if commission_minor <= 0:
                return None

            link_cust = db.execute(
                select(SalesCustomer).where(
                    SalesCustomer.sales_rep_id == rep.id, SalesCustomer.org_id == org_id
                )
            ).scalars().first()

            comm = SalesCommission(
                sales_rep_id=rep.id,
                sales_customer_id=link_cust.id if link_cust is not None else None,
                org_id=org_id,
                invoice_id=invoice_id,
                amount_minor=commission_minor,
                currency=currency,
                kind=kind,
                status="pending",
                note=note,
            )
            db.add(comm)
            # one_time_plus_commission: also grant one-time on first qualifying accrual
            try:
                from app.services.sales_hub_benefits import parse_commission_mode as _pcm

                if _pcm(rep) == "one_time_plus_commission":
                    already_ot = db.execute(
                        select(SalesCommission).where(
                            SalesCommission.sales_rep_id == rep.id,
                            SalesCommission.org_id == org_id,
                            SalesCommission.kind == "one_time_bonus",
                        )
                    ).scalar_one_or_none()
                    bonus_amt = int(getattr(rep, "one_time_bonus_minor", None) or 0)
                    if already_ot is None and bonus_amt > 0:
                        db.add(
                            SalesCommission(
                                sales_rep_id=rep.id,
                                sales_customer_id=link_cust.id if link_cust is not None else None,
                                org_id=org_id,
                                invoice_id=invoice_id,
                                amount_minor=bonus_amt,
                                currency=currency,
                                kind="one_time_bonus",
                                status="pending",
                                note=f"One-time bonus {SalesPayoutService.format_gbp(bonus_amt)}.",
                            )
                        )
            except Exception:  # noqa: BLE001
                pass
            db.commit()
            db.refresh(comm)
            return comm
        except Exception:  # noqa: BLE001 — commission accrual must never break a payment
            try:
                db.rollback()
            except Exception:  # noqa: BLE001
                pass
            return None

    @staticmethod
    def _org_plan_interval(db: Session, org_id: str) -> str:
        try:
            from app.models.plan import Plan
            from app.models.subscription import Subscription

            sub = db.execute(
                select(Subscription)
                .where(Subscription.org_id == org_id)
                .order_by(Subscription.created_at.desc())
            ).scalars().first()
            if sub is not None and sub.plan_id:
                plan = db.get(Plan, sub.plan_id)
                if plan is not None and getattr(plan, "interval", None):
                    return "yearly" if str(plan.interval).lower().startswith(("year", "annual")) else "monthly"
        except Exception:  # noqa: BLE001
            pass
        return "monthly"

    @staticmethod
    def _add_months(dt: datetime, months: int) -> datetime:
        from calendar import monthrange

        m0 = dt.month - 1 + int(months)
        year = dt.year + m0 // 12
        month = m0 % 12 + 1
        day = min(dt.day, monthrange(year, month)[1])
        return dt.replace(year=year, month=month, day=day)

    @staticmethod
    def forecast_expected_commissions(
        *,
        rep: SalesRep,
        tiers: list[dict[str, Any]],
        interval: str,
        package_amount_minor: int | None,
        currency: str,
        period_end: datetime | None,
        paid_subscription_count: int,
        existing_kinds: set[str],
    ) -> list[dict[str, Any]]:
        """Upcoming commission estimates (date + value) for salesman won-deal details."""
        from app.services.billing_currency import money_display
        from app.services.sales_hub_benefits import COMMISSION_MONTHS, parse_commission_mode, parse_one_time_bonus_minor
        from app.services.sales_payout_service import (
            COMMISSION_TYPE_FIXED,
            COMMISSION_TYPE_MONTH2,
            COMMISSION_TYPE_PERCENT,
            SalesPayoutService,
        )

        out: list[dict[str, Any]] = []
        cur = str(currency or "GBP").upper()
        amount = max(0, int(package_amount_minor or 0))
        is_yearly = str(interval or "").lower().startswith(("year", "annual"))
        ctype = SalesPayoutService.commission_type_of(rep)
        has_tier_json = bool(getattr(rep, "commission_tiers_json", None))
        base_date = period_end or (datetime.utcnow() + timedelta(days=30))
        kind_map = {
            1: "monthly_1st",
            2: "monthly_2nd",
            3: "monthly_3rd",
            4: "monthly_4th",
            5: "monthly_5th",
            6: "monthly_6th",
        }

        def _row(
            *,
            label: str,
            kind: str,
            expected_at: datetime | None,
            commission_minor: int,
            note: str,
        ) -> dict[str, Any]:
            return {
                "label": label,
                "kind": kind,
                "status": "expected",
                "expected_at": expected_at.isoformat() if expected_at else None,
                "expected_date": expected_at.date().isoformat() if expected_at else None,
                "amount_minor": int(commission_minor),
                "amount_display": money_display(int(commission_minor), cur),
                "currency": cur,
                "note": note,
            }

        mode = parse_commission_mode(rep)
        bonus_minor = parse_one_time_bonus_minor(rep)
        if mode in {"one_time_only", "one_time_plus_commission"} and bonus_minor > 0 and "one_time_bonus" not in existing_kinds:
            out.append(
                _row(
                    label="One-time bonus",
                    kind="one_time_bonus",
                    expected_at=base_date if mode == "one_time_only" else (period_end or base_date),
                    commission_minor=bonus_minor,
                    note="Expected when the customer’s first qualifying subscription payment clears.",
                )
            )
            if mode == "one_time_only":
                return out

        # Partner / every-payment styles
        if ctype in (COMMISSION_TYPE_PERCENT, COMMISSION_TYPE_FIXED) and not (
            has_tier_json and ctype == COMMISSION_TYPE_MONTH2
        ):
            if ctype == COMMISSION_TYPE_FIXED:
                commission_minor = SalesPayoutService.fixed_minor_of(rep)
                note = f"Fixed commission on the next paid subscription invoice."
                label = "Next paid invoice"
                kind = "fixed_invoice"
            else:
                commission_minor = SalesRepService.apply_commission_pct(amount, SalesRepService.commission_pct_of(rep))
                pct = SalesRepService.commission_pct_of(rep)
                note = f"{pct:g}% of the next paid subscription invoice."
                label = "Next paid invoice"
                kind = "partner_invoice" if SalesRepService.is_partner_channel(rep) else "percent_invoice"
            if commission_minor > 0:
                out.append(
                    _row(
                        label=label,
                        kind=kind,
                        expected_at=period_end or base_date,
                        commission_minor=commission_minor,
                        note=note,
                    )
                )
            return out

        # Month 1–6 tiers
        if is_yearly:
            if "yearly_1mo" in existing_kinds or any(k.startswith("monthly_") for k in existing_kinds):
                return out
            tier = next((t for t in tiers if int(t.get("month") or 0) == 2 and t.get("enabled")), None)
            if tier is None:
                tier = next((t for t in tiers if t.get("enabled")), None)
            if tier is None or amount <= 0:
                return out
            base_minor = max(0, int(round(amount / 12))) if amount else 0
            if tier.get("kind") == "fixed":
                commission_minor = int(round(float(tier["value"])))
                note = "Fixed yearly commission (one month equivalent) — expected on next yearly payment."
            else:
                commission_minor = SalesRepService.apply_commission_pct(base_minor, float(tier["value"]))
                note = f"{float(tier['value']):g}% of one month of the yearly plan — expected on next yearly payment."
            if commission_minor > 0:
                out.append(
                    _row(
                        label=f"Month {int(tier.get('month') or 2)} (yearly)",
                        kind="yearly_1mo",
                        expected_at=period_end or base_date,
                        commission_minor=commission_minor,
                        note=note,
                    )
                )
            return out

        paid_n = max(0, int(paid_subscription_count or 0))
        remaining = [
            t
            for t in tiers
            if t.get("enabled")
            and int(t.get("month") or 0) in COMMISSION_MONTHS
            and int(t.get("month") or 0) > paid_n
            and kind_map[int(t["month"])] not in existing_kinds
        ]
        remaining.sort(key=lambda t: int(t["month"]))
        for idx, tier in enumerate(remaining):
            month = int(tier["month"])
            kind = kind_map[month]
            if tier.get("kind") == "fixed":
                commission_minor = int(round(float(tier["value"])))
                note = f"Fixed commission on subscription month {month}."
            else:
                commission_minor = SalesRepService.apply_commission_pct(amount, float(tier["value"]))
                note = f"{float(tier['value']):g}% of subscription month {month}."
            if commission_minor <= 0:
                continue
            # Next unpaid month aligns to current period end; later months step forward.
            steps = max(0, month - paid_n - 1)
            expected_at = SalesRepService._add_months(period_end or base_date, steps) if (period_end or base_date) else None
            out.append(
                _row(
                    label=f"Month {month} commission",
                    kind=kind,
                    expected_at=expected_at,
                    commission_minor=commission_minor,
                    note=note + (f" Estimated around renewal #{idx + 1}." if expected_at else ""),
                )
            )
        return out

    @staticmethod
    def dashboard_stats(db: Session, rep: SalesRep) -> dict[str, Any]:
        customers = db.execute(
            select(SalesCustomer).where(SalesCustomer.sales_rep_id == rep.id)
        ).scalars().all()
        org_ids = SalesRepService._linked_org_ids(db, rep)

        paid_invoices: list[BillingInvoice] = []
        if org_ids:
            paid_invoices = db.execute(
                select(BillingInvoice).where(
                    BillingInvoice.org_id.in_(list(org_ids)),
                    BillingInvoice.status == "paid",
                )
            ).scalars().all()
        total_paid_minor = sum(int(getattr(inv, "amount_gbp_pence", 0) or 0) for inv in paid_invoices)

        from app.services.sales_payout_service import SalesPayoutService

        commissions = db.execute(
            select(SalesCommission)
            .where(SalesCommission.sales_rep_id == rep.id)
            .order_by(SalesCommission.created_at.desc())
        ).scalars().all()
        wallet_totals = SalesPayoutService.wallet_totals(db, rep_id=rep.id)
        payout_invoices = SalesPayoutService.list_invoices(db, rep_id=rep.id)

        won = [c for c in customers if c.status == "won" or c.org_id]
        codes_used = len(org_ids) if SalesRepService.is_partner_channel(rep) else len(
            [c for c in customers if c.offer_sent_at]
        )
        org_names: dict[str, str] = {}
        org_packages: dict[str, dict[str, Any]] = {}
        commissions_by_org: dict[str, list[Any]] = {}
        paid_sub_count_by_org: dict[str, int] = {}
        if org_ids:
            for org in db.execute(select(Organisation).where(Organisation.id.in_(list(org_ids)))).scalars().all():
                org_names[str(org.id)] = str(org.name or org.id)
            for inv in paid_invoices:
                if str(getattr(inv, "kind", "") or "").lower() != "subscription":
                    continue
                oid = str(getattr(inv, "org_id", "") or "")
                if oid:
                    paid_sub_count_by_org[oid] = paid_sub_count_by_org.get(oid, 0) + 1
            try:
                from app.models.plan import Plan
                from app.models.subscription import Subscription
                from app.services.billing_currency import money_display
                from app.services.plan_price_service import PlanPriceService

                for oid in org_ids:
                    sub = (
                        db.execute(
                            select(Subscription)
                            .where(Subscription.org_id == oid)
                            .order_by(Subscription.created_at.desc())
                        )
                        .scalars()
                        .first()
                    )
                    if sub is None or not sub.plan_id:
                        continue
                    plan = db.get(Plan, sub.plan_id)
                    if plan is None:
                        continue
                    cur = str(getattr(sub, "billing_currency", None) or getattr(sub, "currency", None) or "GBP")
                    price = PlanPriceService.get_price(db, plan.id, cur)
                    interval = str(getattr(sub, "billing_interval", None) or getattr(plan, "interval", "") or "monthly")
                    amount = None
                    if price is not None:
                        if str(interval).lower().startswith(("year", "annual")):
                            amount = getattr(price, "yearly_price_minor", None)
                        else:
                            amount = getattr(price, "monthly_price_minor", None)
                    org_packages[str(oid)] = {
                        "plan_name": getattr(plan, "name", None),
                        "plan_code": getattr(plan, "code", None),
                        "service_kind": getattr(plan, "service_kind", None),
                        "billing_interval": interval,
                        "amount_minor": amount,
                        "currency": cur,
                        "amount_display": money_display(amount, cur) if amount is not None else None,
                        "subscription_status": getattr(sub, "status", None),
                        "current_period_end": sub.current_period_end.isoformat() if getattr(sub, "current_period_end", None) else None,
                        "paid_subscription_count": paid_sub_count_by_org.get(str(oid), 0),
                    }
            except Exception:  # noqa: BLE001
                pass

        for row in commissions:
            oid = str(getattr(row, "org_id", None) or "")
            if not oid:
                continue
            commissions_by_org.setdefault(oid, []).append(row)

        def _iso(dt: Any) -> str | None:
            return dt.isoformat() if dt is not None else None

        from app.services.billing_currency import money_display as _money_display
        from app.services.sales_hub_benefits import parse_commission_tiers as _parse_tiers_early

        tiers_for_forecast = _parse_tiers_early(rep)

        won_companies: list[dict[str, Any]] = []
        for c in won:
            oid = str(c.org_id) if c.org_id else None
            pkg = dict(org_packages.get(oid, {})) if oid else {}
            org_comms = commissions_by_org.get(oid or "", []) if oid else []
            total_c = sum(int(x.amount_minor or 0) for x in org_comms)
            pending_c = sum(int(x.amount_minor or 0) for x in org_comms if str(x.status or "") == "pending")
            paid_c = sum(int(x.amount_minor or 0) for x in org_comms if str(x.status or "") == "paid")
            requested_c = sum(int(x.amount_minor or 0) for x in org_comms if str(x.status or "") == "requested")
            cur = str(pkg.get("currency") or (org_comms[0].currency if org_comms else None) or getattr(rep, "currency", None) or "GBP")
            won_at = c.offer_sent_at or c.interested_at or c.updated_at or c.created_at
            existing_kinds = {str(x.kind or "") for x in org_comms}
            period_end_raw = pkg.get("current_period_end")
            period_end_dt = None
            if period_end_raw:
                try:
                    period_end_dt = datetime.fromisoformat(str(period_end_raw).replace("Z", ""))
                except ValueError:
                    period_end_dt = None
            expected = (
                SalesRepService.forecast_expected_commissions(
                    rep=rep,
                    tiers=tiers_for_forecast,
                    interval=str(pkg.get("billing_interval") or "monthly"),
                    package_amount_minor=int(pkg["amount_minor"]) if pkg.get("amount_minor") is not None else None,
                    currency=cur,
                    period_end=period_end_dt,
                    paid_subscription_count=int(pkg.get("paid_subscription_count") or 0),
                    existing_kinds=existing_kinds,
                )
                if oid
                else []
            )
            next_expected = expected[0] if expected else None
            won_companies.append(
                {
                    "id": c.id,
                    "name": c.company_name or c.full_name,
                    "contact_person": c.contact_person or c.full_name,
                    "email": c.email,
                    "mobile": c.mobile,
                    "org_id": c.org_id,
                    "status": "Converted" if c.org_id else (c.status or "Pending"),
                    "created_at": _iso(c.created_at),
                    "won_at": _iso(won_at),
                    "offer_sent_at": _iso(c.offer_sent_at),
                    **pkg,
                    "commission_total_minor": total_c,
                    "commission_pending_minor": pending_c,
                    "commission_paid_minor": paid_c,
                    "commission_requested_minor": requested_c,
                    "commission_total_display": _money_display(total_c, cur) if total_c else None,
                    "commission_pending_display": _money_display(pending_c, cur) if pending_c else None,
                    "commission_paid_display": _money_display(paid_c, cur) if paid_c else None,
                    "expected_commissions": expected,
                    "next_expected_commission": next_expected,
                    "commissions": [
                        {
                            "id": x.id,
                            "amount_minor": int(x.amount_minor or 0),
                            "amount_display": _money_display(int(x.amount_minor or 0), x.currency or cur),
                            "currency": x.currency or cur,
                            "kind": x.kind,
                            "status": x.status,
                            "note": x.note,
                            "created_at": _iso(x.created_at),
                            "invoice_id": x.invoice_id,
                        }
                        for x in org_comms
                    ],
                }
            )

        won_companies.sort(key=lambda row: str(row.get("won_at") or row.get("created_at") or ""), reverse=True)

        from app.services.sales_hub_benefits import (
            benefit_summaries,
            commission_summary,
            currency_of_rep,
            packages_for_currency,
            parse_commission_mode,
            parse_commission_tiers,
            parse_one_time_bonus_minor,
            parse_partner_terms,
            parse_promo_benefits,
        )

        currency = currency_of_rep(rep)
        benefits = parse_promo_benefits(rep)
        tiers = parse_commission_tiers(rep)
        partner_terms = parse_partner_terms(rep)
        mode = parse_commission_mode(rep)
        bonus_minor = parse_one_time_bonus_minor(rep)

        return {
            "kind": SalesRepService.rep_kind(rep),
            "commission_pct": SalesRepService.commission_pct_of(rep),
            "commission_type": SalesPayoutService.commission_type_of(rep),
            "commission_fixed_minor": SalesPayoutService.fixed_minor_of(rep),
            "commission_mode": mode,
            "one_time_bonus_minor": bonus_minor,
            "commission_tiers": tiers,
            "commission_summary": commission_summary(
                tiers,
                currency=currency,
                partner=SalesRepService.is_partner_channel(rep),
                partner_terms=partner_terms,
                commission_mode=mode,
                one_time_bonus_minor=bonus_minor,
            ),
            "promo_code": rep.promo_code,
            "promo_benefits": benefits,
            "promo_benefit_summaries": benefit_summaries(benefits, currency=currency),
            "partner_terms": partner_terms,
            "currency": currency,
            "country": rep.country,
            "packages": packages_for_currency(db, currency),
            "payout": SalesPayoutService.payout_dict(rep),
            "won_deals": {
                "count": len(won_companies),
                "total_value_minor": total_paid_minor,
                "companies": won_companies,
            },
            "wallet": {
                "active_companies": len(org_ids),
                "codes_used": codes_used,
                "revenue_minor": total_paid_minor,
                "commission_minor": wallet_totals["total_minor"],
                "commission_paid_minor": wallet_totals["paid_minor"],
                "commission_pending_minor": wallet_totals["available_minor"] + wallet_totals["requested_minor"],
                "commission_available_minor": wallet_totals["available_minor"],
                "commission_requested_minor": wallet_totals["requested_minor"],
            },
            "commissions": [
                {
                    "id": c.id,
                    "org_id": c.org_id,
                    "org_name": org_names.get(str(c.org_id or ""), c.org_id),
                    "invoice_id": c.invoice_id,
                    "payout_invoice_id": getattr(c, "payout_invoice_id", None),
                    "amount_minor": int(c.amount_minor or 0),
                    "currency": c.currency or "GBP",
                    "kind": c.kind,
                    "status": c.status,
                    "note": c.note,
                    "created_at": c.created_at.isoformat() if c.created_at else None,
                }
                for c in commissions
            ],
            "payout_invoices": payout_invoices,
            "visited_count": len(customers),
        }
