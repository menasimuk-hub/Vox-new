from __future__ import annotations

import csv
import io
import json
import logging
import re
from datetime import datetime
from typing import Any

from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.platform_service import PlatformService
from app.models.service_order import ServiceOrder, ServiceOrderRecipient
from app.services.billing_call_minutes import billable_call_minutes, call_outcome_label

logger = logging.getLogger(__name__)

SURVEY_CHANNEL_ALIASES = {"call": "ai_call"}
SURVEY_DELIVERY_CHANNELS = frozenset({"ai_call", "whatsapp"})


class PlatformCatalogService:
    # Service catalog (no pricing rules — all VoxBulk pricing lives in plan_prices /
    # pricing_currency_settings, resolved per-org currency by PlanPriceService).
    DEFAULT_SERVICES = [
        {
            "code": "survey",
            "name": "Survey",
            "description": "AI phone and WhatsApp survey campaigns with smart reporting.",
            "service_kind": "order",
            "sort_order": 10,
        },
        {
            "code": "interview",
            "name": "Interview",
            "description": "AI phone or web meeting interview screening campaigns.",
            "service_kind": "order",
            "sort_order": 20,
        },
        {
            "code": "interview_ats",
            "name": "Interview ATS",
            "description": "AI CV screening score per candidate (DeepSeek ATS).",
            "service_kind": "addon",
            "sort_order": 25,
        },
        {
            "code": "appointments",
            "name": "Appointment Manager",
            "description": "CRM-synced appointments with WhatsApp and AI call confirmation.",
            "service_kind": "order",
            "sort_order": 30,
        },
    ]

    @staticmethod
    def normalize_survey_channel(raw: str | None) -> str:
        channel = str(raw or "").strip().lower()
        return SURVEY_CHANNEL_ALIASES.get(channel, channel)

    @staticmethod
    def interview_delivery_options(db: Session) -> list[str]:
        return ["ai_call", "ai_meeting"]

    @staticmethod
    def interview_platform_capabilities(db: Session) -> dict[str, Any]:
        options = PlatformCatalogService.interview_delivery_options(db)
        return {
            "interview_meeting_enabled": "ai_meeting" in options,
            "interview_delivery_options": options,
        }

    @staticmethod
    def normalize_interview_delivery(db: Session, raw: str | None) -> str:
        delivery = PlatformCatalogService.normalize_survey_channel(str(raw or "ai_call"))
        if delivery not in {"ai_call", "ai_meeting"}:
            raise ValueError("Interview delivery must be ai_call or ai_meeting")
        if delivery not in PlatformCatalogService.interview_delivery_options(db):
            raise ValueError("Interview delivery mode is not available")
        return delivery

    @staticmethod
    def resolve_survey_channel(options: dict[str, Any] | None) -> str:
        options = options or {}

        explicit = options.get("survey_channel") or options.get("delivery")
        if explicit:
            channel = PlatformCatalogService.normalize_survey_channel(str(explicit))
            if channel not in SURVEY_DELIVERY_CHANNELS:
                raise ValueError("Survey channel must be ai_call or whatsapp")
            return channel

        channels = options.get("channels")
        if isinstance(channels, list) and channels:
            normalized = [PlatformCatalogService.normalize_survey_channel(str(c)) for c in channels if c]
            normalized = [c for c in normalized if c in SURVEY_DELIVERY_CHANNELS]
            if len(normalized) > 1:
                raise ValueError("Mixed survey channels are not supported — choose ai_call or whatsapp")
            if len(normalized) == 1:
                return normalized[0]

        contact_method = str(options.get("contact_method") or "").strip().lower()
        if "both" in contact_method:
            raise ValueError("Mixed survey channels are not supported — choose ai_call or whatsapp")
        if "whatsapp" in contact_method:
            return "whatsapp"
        if "ai" in contact_method or "phone" in contact_method or "call" in contact_method:
            return "ai_call"

        return "ai_call"

    @staticmethod
    def ensure_defaults(db: Session) -> None:
        for svc in PlatformCatalogService.DEFAULT_SERVICES:
            code = str(svc["code"])
            rows = list(
                db.execute(
                    select(PlatformService)
                    .where(PlatformService.code == code)
                    .order_by(PlatformService.created_at.asc(), PlatformService.id.asc())
                ).scalars()
            )
            if len(rows) > 1:
                logger.error(
                    "duplicate PlatformService code=%s ids=%s — deactivating extras, keeping id=%s",
                    code,
                    [row.id for row in rows],
                    rows[0].id,
                )
                row = rows[0]
                for dup in rows[1:]:
                    dup.is_active = False
                    db.add(dup)
            elif len(rows) == 1:
                row = rows[0]
            else:
                row = PlatformService(
                    code=code,
                    name=svc["name"],
                    description=svc["description"],
                    service_kind=svc["service_kind"],
                    sort_order=int(svc["sort_order"]),
                    is_active=True,
                )
                db.add(row)
                db.flush()

        db.commit()

    @staticmethod
    def list_services(db: Session, *, active_only: bool = True) -> list[PlatformService]:
        PlatformCatalogService.ensure_defaults(db)
        stmt = select(PlatformService).order_by(PlatformService.sort_order.asc(), PlatformService.name.asc())
        if active_only:
            stmt = stmt.where(PlatformService.is_active.is_(True))
        return list(db.execute(stmt).scalars())

    @staticmethod
    def get_service_by_code(db: Session, code: str) -> PlatformService | None:
        PlatformCatalogService.ensure_defaults(db)
        return (
            db.execute(
                select(PlatformService)
                .where(PlatformService.code == code)
                .order_by(PlatformService.updated_at.desc())
                .limit(1)
            )
            .scalars()
            .first()
        )

    @staticmethod
    def _money(pence: int) -> str:
        return f"£{int(pence or 0) / 100:.2f}"

    @staticmethod
    def calculate_quote(
        db: Session,
        *,
        service_code: str,
        recipient_count: int,
        options: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Estimate a campaign cost from VoxBulk per-currency pricing (plan_prices)."""
        from app.models.organisation import Organisation
        from app.services.billing_currency import money_display
        from app.services.plan_price_service import PlanPriceService

        options = options or {}
        svc = PlatformCatalogService.get_service_by_code(db, service_code)
        if svc is None:
            raise ValueError(f"Unknown service: {service_code}")

        count = max(int(recipient_count or 0), 0)
        org_id = str(options.get("org_id") or "").strip() or None
        org = db.get(Organisation, org_id) if org_id else None
        from app.services.gocardless_service import BillingService

        plan = BillingService.resolve_active_plan(db, org_id) if org_id else None
        rates = PlanPriceService.rates_for_org(db, org, plan=plan)
        currency = str(rates["currency"])

        lines: list[dict[str, Any]] = []
        total = 0
        payload: dict[str, Any] = {
            "service_code": service_code,
            "recipient_count": count,
            "currency": currency,
        }

        if service_code == "survey":
            channel = PlatformCatalogService.resolve_survey_channel(options)
            payload["survey_channel"] = channel
            if channel == "whatsapp":
                unit = int(rates["wa_extra_minor"] or 0)
                total = count * unit
                lines.append(
                    {
                        "kind": "per_recipient",
                        "channel": channel,
                        "label": "WhatsApp survey recipients",
                        "amount_pence": total,
                        "detail": f"WhatsApp survey: {count} × {money_display(unit, currency)} = {money_display(total, currency)}",
                        "unit_price_pence": unit,
                    }
                )
            else:
                duration = max(int(options.get("duration_minutes") or options.get("estimated_duration_min") or 12), 1)
                per_min = int(rates["interview_per_min_minor"] or 0)
                conn = int(rates["connection_fee_minor"] or 0)
                per_call = conn + per_min * duration
                total = per_call * count
                if conn > 0:
                    lines.append(
                        {
                            "kind": "connection_fee",
                            "channel": channel,
                            "label": "Connection fee",
                            "amount_pence": conn * count,
                            "detail": f"Connection fee: {count} × {money_display(conn, currency)}",
                        }
                    )
                lines.append(
                    {
                        "kind": "per_minute",
                        "channel": channel,
                        "label": "AI call minutes",
                        "amount_pence": per_min * duration * count,
                        "detail": f"Call minutes: {count} × {duration} min × {money_display(per_min, currency)}/min",
                        "duration_minutes": duration,
                        "per_min_pence": per_min,
                    }
                )
        elif service_code == "interview":
            delivery = PlatformCatalogService.normalize_interview_delivery(
                db, str(options.get("delivery") or "ai_call")
            )
            payload["delivery"] = delivery
            duration = max(int(options.get("duration_minutes") or 12), 1)
            per_min = int(rates["interview_per_min_minor"] or 0)
            conn = int(rates["connection_fee_minor"] or 0) if delivery in {"ai_call", "ai_meeting"} else 0
            per_call = conn + per_min * duration
            total = per_call * count
            if conn > 0:
                lines.append(
                    {
                        "kind": "connection_fee",
                        "channel": delivery,
                        "label": "Connection fee",
                        "amount_pence": conn * count,
                        "detail": f"Connection fee: {count} × {money_display(conn, currency)}",
                    }
                )
            lines.append(
                {
                    "kind": "per_minute",
                    "channel": delivery,
                    "label": "Interview AI call — per minute",
                    "amount_pence": per_min * duration * count,
                    "detail": f"Call minutes: {count} × {duration} min × {money_display(per_min, currency)}/min",
                    "duration_minutes": duration,
                    "per_min_pence": per_min,
                }
            )
        elif service_code == "interview_ats":
            unit = int(rates["cv_scan_fee_minor"] or 0)
            total = count * unit
            lines.append(
                {
                    "kind": "per_person",
                    "channel": "base",
                    "label": "ATS scan per CV",
                    "amount_pence": total,
                    "detail": f"ATS scans: {count} × {money_display(unit, currency)} = {money_display(total, currency)}",
                    "unit_price_pence": unit,
                }
            )
        else:
            raise ValueError(f"No pricing defined for service: {service_code}")

        payload.update(
            {
                "total_pence": total,
                "total_gbp": money_display(total, currency),
                "total_display": money_display(total, currency),
                "lines": lines,
            }
        )
        return payload


class ServiceOrderService:
    RECIPIENT_TEMPLATE_HEADERS = ["name", "phone", "email"]
    SURVEY_RECIPIENT_TEMPLATE_HEADERS = ["name", "phone", "language"]

    @staticmethod
    def recipient_template_csv(*, for_survey: bool = False) -> str:
        headers = (
            ServiceOrderService.SURVEY_RECIPIENT_TEMPLATE_HEADERS
            if for_survey
            else ServiceOrderService.RECIPIENT_TEMPLATE_HEADERS
        )
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        if for_survey:
            writer.writerow(["Sarah Ahmed", "+447700900123", "en"])
            writer.writerow(["James Lee", "+447700900456", ""])
        else:
            writer.writerow(["Sarah Ahmed", "+447700900123", "sarah@example.com"])
            writer.writerow(["James Lee", "+447700900456", ""])
        return buf.getvalue()

    @staticmethod
    def _norm_header(h: str) -> str:
        return re.sub(r"[^a-z0-9]", "", str(h or "").strip().lower())

    @staticmethod
    def _excel_cell_str(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            text = format(value, "f").rstrip("0").rstrip(".")
            return text.strip()
        return str(value).strip()

    @staticmethod
    def _spreadsheet_kind(content: bytes, filename: str) -> str:
        if content[:2] == b"PK":
            return "xlsx"
        if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            return "xls"
        name = str(filename or "").lower()
        if name.endswith(".xlsx"):
            return "xlsx"
        if name.endswith(".xls"):
            return "xls"
        return "csv"

    @staticmethod
    def _row_field_values(data: dict[str, str]) -> tuple[str, str, str]:
        name = (
            data.get("name")
            or data.get("fullname")
            or data.get("contactname")
            or data.get("customername")
            or data.get("respondentname")
            or ""
        )
        if not name:
            first = data.get("firstname") or data.get("first") or data.get("givenname") or ""
            last = data.get("lastname") or data.get("last") or data.get("surname") or data.get("familyname") or ""
            name = f"{first} {last}".strip()
        phone = (
            data.get("phone")
            or data.get("mobile")
            or data.get("telephone")
            or data.get("phonenumber")
            or data.get("mobilenumber")
            or data.get("cellphone")
            or data.get("cell")
            or data.get("tel")
            or data.get("contactnumber")
            or data.get("contactphone")
            or data.get("whatsapp")
            or data.get("whatsappnumber")
            or data.get("msisdn")
            or ""
        )
        email = data.get("email") or data.get("emailaddress") or data.get("mail") or data.get("emailid") or ""
        language = data.get("language") or data.get("locale") or data.get("lang") or ""
        return str(name or "").strip(), str(phone or "").strip(), str(email or "").strip(), str(language or "").strip()

    @staticmethod
    def _parse_recipient_spreadsheet(content: bytes, *, filename: str, kind: str) -> list[dict[str, str]]:
        if kind == "xls":
            raise ValueError(
                "Legacy .xls workbooks are not supported. In Excel choose Save As → .xlsx, or upload a CSV file."
            )
        try:
            import openpyxl
        except ImportError as e:
            raise ValueError("Excel upload requires openpyxl on the server. Use CSV for now.") from e
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(
                "Could not read the Excel file. Use .xlsx format with columns name and phone, or upload CSV."
            ) from e
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("Excel file is empty — add a header row (name, phone) and at least one contact.")
        headers = [ServiceOrderService._norm_header(x) for x in header_row]
        if not any(h for h in headers):
            raise ValueError("Excel header row must include name and phone columns.")
        rows: list[dict[str, str]] = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data: dict[str, str] = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                cell = row[i] if i < len(row) else None
                data[header] = ServiceOrderService._excel_cell_str(cell)
            parsed = ServiceOrderService._row_from_dict(data, idx)
            if parsed:
                rows.append(parsed)
        return rows

    @staticmethod
    def parse_recipient_file(content: bytes, filename: str) -> list[dict[str, str]]:
        kind = ServiceOrderService._spreadsheet_kind(content, filename)
        if kind in {"xlsx", "xls"}:
            return ServiceOrderService._parse_recipient_spreadsheet(content, filename=filename, kind=kind)
        from app.utils.text_decoding import decode_uploaded_text

        text = decode_uploaded_text(content)
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV must include a header row: name, phone, email")
        rows: list[dict[str, str]] = []
        for idx, raw in enumerate(reader, start=2):
            data = {ServiceOrderService._norm_header(k): str(v or "").strip() for k, v in raw.items()}
            parsed = ServiceOrderService._row_from_dict(data, idx)
            if parsed:
                rows.append(parsed)
        return rows

    @staticmethod
    def _row_from_dict(data: dict[str, str], row_number: int) -> dict[str, str] | None:
        name, phone, email, language = ServiceOrderService._row_field_values(data)
        if not name and not phone:
            return None
        if not name or not phone:
            raise ValueError(f"Row {row_number}: name and phone are required")
        row: dict[str, str] = {"name": name, "phone": phone, "email": email or None}
        if language:
            row["language"] = language
        return row

    @staticmethod
    def order_to_dict(
        order: ServiceOrder,
        *,
        include_recipients: bool = False,
        recipients: list[ServiceOrderRecipient] | None = None,
        db: Session | None = None,
    ) -> dict[str, Any]:
        config = {}
        breakdown = []
        quote_meta: dict[str, Any] = {}
        report = None
        try:
            if order.config_json:
                config = json.loads(order.config_json)
        except Exception:
            config = {}
        if order.service_code == "survey" and isinstance(config, dict):
            from app.services.survey_builder_flow_service import (
                normalize_survey_config_step_labels,
                survey_step_labels_from_config,
            )

            config = normalize_survey_config_step_labels(config, campaign_title=str(order.title or ""))
        try:
            if order.quote_breakdown_json:
                parsed = json.loads(order.quote_breakdown_json)
                if isinstance(parsed, dict):
                    breakdown = parsed.get("lines") or []
                    quote_meta = {
                        "survey_channel": parsed.get("survey_channel"),
                        "selected_package_id": parsed.get("selected_package_id"),
                    }
                else:
                    breakdown = parsed
        except Exception:
            breakdown = []
        try:
            if order.report_json:
                report = json.loads(order.report_json)
        except Exception:
            report = None
        launch_billing: dict[str, Any] = {}
        try:
            if order.launch_billing_json:
                parsed_launch = json.loads(order.launch_billing_json)
                if isinstance(parsed_launch, dict):
                    launch_billing = parsed_launch
        except Exception:
            launch_billing = {}
        out = {
            "id": order.id,
            "org_id": order.org_id,
            "user_id": order.user_id,
            "service_code": order.service_code,
            "title": order.title,
            "survey_name": str(config.get("survey_name") or order.title or ""),
            "survey_id": str(order.campaign_id or config.get("survey_id") or ""),
            "reference_id": order.reference_id,
            "campaign_id": order.campaign_id,
            "status": order.status,
            "payment_status": order.payment_status,
            "recipient_count": order.recipient_count,
            "quote_total_pence": order.quote_total_pence,
            "quote_total_gbp": f"£{int(order.quote_total_pence or 0) / 100:.2f}",
            "quote_breakdown": breakdown,
            "quote_survey_channel": quote_meta.get("survey_channel"),
            "quote_selected_package_id": quote_meta.get("selected_package_id"),
            "config": config,
            "run_mode": order.run_mode,
            "scheduled_start_at": order.scheduled_start_at.isoformat() if order.scheduled_start_at else None,
            "scheduled_end_at": order.scheduled_end_at.isoformat() if order.scheduled_end_at else None,
            "started_at": order.started_at.isoformat() if order.started_at else None,
            "completed_at": order.completed_at.isoformat() if order.completed_at else None,
            "payment_method": order.payment_method,
            "payment_note": order.payment_note,
            "admin_decision_note": order.admin_decision_note,
            "report": report,
            "launch_billing": launch_billing,
            "billing_settlement": launch_billing.get("settlement") if launch_billing else None,
            "billing_phase": launch_billing.get("billing_phase") if launch_billing else None,
            "created_at": order.created_at.isoformat() if order.created_at else None,
            "updated_at": order.updated_at.isoformat() if order.updated_at else None,
        }
        from app.services.service_order_workflow_service import ServiceOrderWorkflowService

        out.update(ServiceOrderWorkflowService.visible_state(order))
        if include_recipients:
            if order.service_code == "interview":
                from app.services.interview_activity_service import InterviewActivityService

                enriched: list[dict[str, Any]] = []
                for r in recipients or []:
                    row = ServiceOrderService.recipient_to_dict(r)
                    row["activity_status"] = InterviewActivityService.activity_status(r)
                    enriched.append(row)
                out["recipients"] = enriched
            else:
                out["recipients"] = [ServiceOrderService.recipient_to_dict(r) for r in (recipients or [])]

            total_secs = 0
            billable_mins = 0
            connected = 0
            for row in out.get("recipients") or []:
                raw_secs = row.get("duration_seconds")
                try:
                    secs = int(raw_secs) if raw_secs is not None else 0
                except (TypeError, ValueError):
                    secs = 0
                if secs <= 0:
                    continue
                total_secs += secs
                connected += 1
                billable_mins += billable_call_minutes(secs)
            out["call_usage"] = {
                "total_duration_seconds": total_secs,
                "billable_minutes_actual": billable_mins,
                "connected_calls": connected,
            }
        out["is_archived"] = ServiceOrderService.is_archived_order(order)
        if order.service_code == "survey":
            from app.services.survey_builder_flow_service import survey_step_labels_from_config

            step_labels = survey_step_labels_from_config(
                config,
                campaign_title=str(config.get("survey_name") or order.title or ""),
                campaign_goal=str(config.get("goal") or ""),
                db=db,
            )
            out["step_labels"] = step_labels
            out["first_step_name"] = step_labels[0] if step_labels else ""
            out["next_action"] = ServiceOrderService.survey_next_action(order)
            out["status_label"] = ServiceOrderService.survey_status_label(order)
            out["is_live"] = ServiceOrderService.is_live_survey(order)
            out["is_finished"] = ServiceOrderService.is_finished_survey(order)
            out["audit_timeline"] = ServiceOrderService.order_audit_timeline(order)
        if order.service_code == "interview":
            recs = recipients if recipients is not None else None
            out["is_live"] = ServiceOrderService.is_live_interview(order, recipients=recs)
            out["is_finished"] = ServiceOrderService.is_finished_interview(order, recipients=recs)
            out["status_label"] = ServiceOrderService.interview_status_label(order)
            from app.services.interview_cv_email_service import interview_cv_phase_payload

            out["cv_collection"] = interview_cv_phase_payload(order)
        return out

    @staticmethod
    def recipient_to_dict(recipient: ServiceOrderRecipient) -> dict[str, Any]:
        result: dict[str, Any] = {}
        try:
            if recipient.result_json:
                parsed = json.loads(recipient.result_json)
                if isinstance(parsed, dict):
                    result = parsed
        except Exception:
            result = {}
        analysis = result.get("analysis") if isinstance(result.get("analysis"), dict) else {}
        hangup = str(result.get("hangup_cause") or "")
        voicemail = bool(result.get("voicemail") or result.get("answering_machine"))
        duration_seconds = result.get("duration_seconds")
        try:
            duration_seconds = int(duration_seconds) if duration_seconds is not None else None
        except (TypeError, ValueError):
            duration_seconds = None
        bm = result.get("billable_minutes")
        try:
            billable_minutes_val = int(bm) if bm is not None else billable_call_minutes(duration_seconds)
        except (TypeError, ValueError):
            billable_minutes_val = billable_call_minutes(duration_seconds)
        out = {
            "id": recipient.id,
            "row_number": recipient.row_number,
            "name": recipient.name,
            "phone": recipient.phone,
            "email": recipient.email,
            "status": recipient.status,
            "duration_seconds": duration_seconds,
            "billable_minutes": billable_minutes_val,
            "call_type": call_outcome_label(
                status=str(recipient.status or ""),
                hangup_cause=hangup,
                voicemail=voicemail,
            ),
            "hangup_cause": hangup or None,
            "call_channel": result.get("channel") or "ai_call",
            "transport": result.get("transport"),
            "call_control_id": result.get("call_control_id"),
            "telnyx_conversation_id": result.get("telnyx_conversation_id") or result.get("conversation_id"),
            "call_session_id": result.get("call_session_id") or result.get("telnyx_session_id"),
            "call_summary": result.get("call_summary"),
            "sentiment": analysis.get("sentiment") or result.get("sentiment"),
            "short_summary": analysis.get("short_summary") or result.get("short_summary"),
            "created_at": recipient.created_at.isoformat() if recipient.created_at else None,
        }
        if recipient.cv_quality is not None or recipient.cv_filename or recipient.intake_source or recipient.cv_text:
            from app.services.interview_intake_service import compute_intake_errors
            from app.services.interview_ats_service import ats_display_for_recipient

            out.update(
                {
                    "cv_quality": recipient.cv_quality or "missing",
                    "cv_filename": recipient.cv_filename,
                    "intake_source": recipient.intake_source,
                    "intake_errors": compute_intake_errors(recipient),
                    "intake_ready": bool(str(recipient.name or "").strip() and str(recipient.phone or "").strip()),
                    "has_cv_file": bool(recipient.cv_storage_key or (recipient.cv_text or "").strip()),
                }
            )
            out.update(ats_display_for_recipient(recipient, position=""))
        return out

    @staticmethod
    def order_audit_timeline(order: ServiceOrder) -> list[dict[str, Any]]:
        events: list[dict[str, Any]] = []

        def add(at, kind, label, detail=""):
            if not at:
                return
            events.append(
                {
                    "at": at.isoformat() if hasattr(at, "isoformat") else str(at),
                    "kind": kind,
                    "label": label,
                    "detail": detail or None,
                }
            )

        label_created = "Interview created" if order.service_code == "interview" else "Survey created"
        label_started = "Interview started" if order.service_code == "interview" else "Survey started"
        label_finished = "Interview finished" if order.service_code == "interview" else "Survey finished"
        label_paused = "Interview paused" if order.service_code == "interview" else "Survey paused"
        label_cancelled = "Interview cancelled" if order.service_code == "interview" else "Survey cancelled"

        add(order.created_at, "created", label_created)
        add(order.updated_at, "updated", "Last updated")
        if order.payment_status == "pending_approval":
            add(order.updated_at, "payment", "Cash payment submitted", order.payment_note)
        if order.payment_status == "approved":
            add(order.updated_at, "payment", "Payment approved", order.admin_decision_note or order.payment_note)
        if order.payment_status == "rejected":
            add(order.updated_at, "payment", "Payment rejected", order.admin_decision_note or order.payment_note)
        add(order.scheduled_start_at, "schedule", "Scheduled start")
        add(order.scheduled_end_at, "schedule", "Scheduled end")
        add(order.started_at, "status", label_started)
        add(order.completed_at, "status", label_finished)
        if order.status == "paused":
            add(order.updated_at, "status", label_paused)
        if order.status == "cancelled":
            add(order.completed_at or order.updated_at, "status", label_cancelled, order.admin_decision_note)
        events.sort(key=lambda e: e["at"] or "")
        return events

    @staticmethod
    def update_recipient(
        db: Session,
        order: ServiceOrder,
        recipient: ServiceOrderRecipient,
        payload: dict[str, Any],
    ) -> ServiceOrderRecipient:
        if recipient.order_id != order.id:
            raise ValueError("Recipient does not belong to this order")
        if order.status == "completed":
            raise ValueError("Cannot edit contacts on a completed survey")
        if "name" in payload:
            name = str(payload.get("name") or "").strip()
            if not name:
                raise ValueError("Name is required")
            recipient.name = name
        if "phone" in payload:
            phone = str(payload.get("phone") or "").strip()
            if not phone:
                raise ValueError("Phone is required")
            recipient.phone = phone
        if "email" in payload:
            email = str(payload.get("email") or "").strip()
            recipient.email = email or None
        if "status" in payload and str(payload.get("status") or "").strip():
            recipient.status = str(payload.get("status")).strip()
        db.add(recipient)
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(recipient)
        return recipient

    @staticmethod
    def survey_operations_overview(db: Session) -> dict[str, Any]:
        rows = list(
            db.execute(select(ServiceOrder).where(ServiceOrder.service_code == "survey")).scalars()
        )
        running = sum(1 for r in rows if r.status == "running")
        paused = sum(1 for r in rows if r.status == "paused")
        completed = sum(1 for r in rows if r.status == "completed")
        failed_pay = sum(1 for r in rows if r.payment_status == "rejected")
        live = sum(1 for r in rows if ServiceOrderService.is_live_survey(r))
        scheduled = sum(1 for r in rows if r.status == "scheduled")
        pending = sum(1 for r in rows if r.payment_status == "pending_approval")
        return {
            "total": len(rows),
            "live": live,
            "running": running,
            "paused": paused,
            "scheduled": scheduled,
            "completed": completed,
            "failed_payments": failed_pay,
            "pending_payment_approval": pending,
        }

    @staticmethod
    def interview_operations_overview(db: Session) -> dict[str, Any]:
        rows = list(
            db.execute(select(ServiceOrder).where(ServiceOrder.service_code == "interview")).scalars()
        )
        running = sum(1 for r in rows if r.status == "running")
        paused = sum(1 for r in rows if r.status == "paused")
        completed = sum(1 for r in rows if r.status == "completed")
        failed_pay = sum(1 for r in rows if r.payment_status == "rejected")
        live = sum(1 for r in rows if ServiceOrderService.is_live_interview(r))
        scheduled = sum(1 for r in rows if r.status == "scheduled")
        pending = sum(1 for r in rows if r.payment_status == "pending_approval")
        drafts = sum(1 for r in rows if r.status == "draft")
        return {
            "total": len(rows),
            "live": live,
            "running": running,
            "paused": paused,
            "scheduled": scheduled,
            "completed": completed,
            "failed_payments": failed_pay,
            "pending_payment_approval": pending,
            "drafts": drafts,
        }

    @staticmethod
    def is_archived_order(order: ServiceOrder) -> bool:
        return str(order.status or "") == "archived"

    @staticmethod
    def is_finished_interview(
        order: ServiceOrder,
        recipients: list[ServiceOrderRecipient] | None = None,
    ) -> bool:
        if order.service_code != "interview":
            return False
        if ServiceOrderService.is_archived_order(order):
            return False
        if order.status == "cancelled":
            return True
        if order.status != "completed":
            return False
        if recipients:
            for recipient in recipients:
                parsed: dict[str, Any] = {}
                try:
                    raw = json.loads(recipient.result_json or "{}")
                    if isinstance(raw, dict):
                        parsed = raw
                except Exception:
                    parsed = {}
                status = str(recipient.status or "").lower()
                if parsed.get("analysis_saved_at") or parsed.get("call_completed_at"):
                    continue
                if status in {"completed", "done"}:
                    continue
                if status in {"", "pending", "queued", "sent", "ringing", "calling", "in_progress"}:
                    return False
                if parsed.get("booked_start_at") and status not in {
                    "completed",
                    "no_answer",
                    "failed",
                    "busy",
                    "skipped",
                    "cancelled",
                    "opted_out",
                    "done",
                }:
                    return False
        return True

    @staticmethod
    def is_live_interview(
        order: ServiceOrder,
        recipients: list[ServiceOrderRecipient] | None = None,
    ) -> bool:
        return (
            order.service_code == "interview"
            and not ServiceOrderService.is_finished_interview(order, recipients=recipients)
            and not ServiceOrderService.is_archived_order(order)
        )

    @staticmethod
    def interview_status_label(order: ServiceOrder) -> str:
        if order.status == "running":
            return "Live"
        if order.status == "completed":
            return "Completed"
        if order.status == "cancelled":
            return "Cancelled"
        if order.status == "archived":
            return "Archived"
        if order.payment_status == "pending_approval":
            return "Awaiting payment"
        if order.status in {"quoted", "awaiting_payment"}:
            return "Quoted"
        if order.status == "scheduled":
            return "Scheduled"
        if order.status == "draft":
            return "Draft"
        return order.status or "—"

    @staticmethod
    def is_finished_survey(order: ServiceOrder) -> bool:
        return (
            order.service_code == "survey"
            and str(order.status or "") in {"completed", "cancelled"}
            and not ServiceOrderService.is_archived_order(order)
        )

    @staticmethod
    def is_live_survey(order: ServiceOrder) -> bool:
        return (
            order.service_code == "survey"
            and not ServiceOrderService.is_finished_survey(order)
            and not ServiceOrderService.is_archived_order(order)
        )

    @staticmethod
    def survey_status_label(order: ServiceOrder) -> str:
        ps = str(order.payment_status or "")
        st = str(order.status or "")
        if ps == "pending_approval":
            return "Pending payment"
        if ps == "rejected":
            return "Payment failed"
        if st == "draft":
            return "Draft"
        if st == "quoted":
            return "Quoted"
        if st == "awaiting_payment":
            return "Awaiting payment"
        if st == "scheduled":
            return "Scheduled"
        if st == "running":
            return "Running"
        if st == "paused":
            return "Paused"
        if st == "paid":
            return "Paid — ready"
        if st == "completed":
            return "Finished"
        if st == "cancelled":
            return "Cancelled"
        if st == "archived":
            return "Archived"
        return st.replace("_", " ").title()

    @staticmethod
    def survey_next_action(order: ServiceOrder) -> dict[str, Any]:
        config: dict[str, Any] = {}
        try:
            config = json.loads(order.config_json or "{}")
        except Exception:
            config = {}
        ps = str(order.payment_status or "")
        st = str(order.status or "")

        if st == "completed":
            return {"action": "view_report", "label": "View report", "hint": "Open anonymous aggregate results."}
        if st == "cancelled":
            return {"action": "reopen", "label": "Reopen survey", "hint": "Duplicate this survey to launch again."}
        if ps == "rejected":
            reason = (order.admin_decision_note or order.payment_note or "Payment was rejected.").strip()
            return {"action": "pay", "label": "Retry payment", "hint": reason, "reason": reason}
        if ps == "pending_approval":
            return {
                "action": "wait",
                "label": "Waiting for approval",
                "hint": "Admin must approve your cash payment before the survey can start.",
            }
        if ps == "unpaid" and st in {"draft", "quoted", "awaiting_payment"}:
            if not config.get("script_approved"):
                return {"action": "approve_prompt", "label": "Approve prompt", "hint": "Approve the AI script before paying."}
            if not order.scheduled_start_at:
                return {"action": "set_schedule", "label": "Set date and time", "hint": "Choose when calls should start and end."}
            if order.recipient_count <= 0:
                return {"action": "upload_contacts", "label": "Upload contacts", "hint": "Add a contact list to continue."}
            return {
                "action": "pay",
                "label": "Pay now",
                "hint": f"Total due: £{int(order.quote_total_pence or 0) / 100:.2f}",
            }
        if ps == "approved" and st in {"paid", "scheduled"}:
            if order.run_mode == "scheduled" and order.scheduled_start_at:
                return {"action": "wait", "label": "Scheduled", "hint": "Survey will start automatically at the scheduled time."}
            return {"action": "start", "label": "Start survey", "hint": "Payment approved — start outbound calls."}
        if st == "running":
            return {"action": "pause", "label": "Pause survey", "hint": "Pause outbound calls. You can resume later."}
        if st == "paused":
            return {"action": "resume", "label": "Resume survey", "hint": "Continue outbound calls."}
        return {"action": "edit", "label": "Review survey", "hint": "Open details to review settings."}

    @staticmethod
    def order_to_admin_dict(
        db: Session,
        order: ServiceOrder,
        *,
        include_recipients: bool = False,
        recipients: list[ServiceOrderRecipient] | None = None,
    ) -> dict[str, Any]:
        from app.models.organisation import Organisation
        from app.models.user import User

        out = ServiceOrderService.order_to_dict(order, include_recipients=include_recipients, recipients=recipients)
        org = db.get(Organisation, order.org_id)
        user = db.get(User, order.user_id)
        out["org_name"] = org.name if org else None
        out["org_phone"] = getattr(org, "contact_phone", None) if org else None
        out["owner_email"] = user.email if user else None
        return out

    @staticmethod
    def duplicate_order(
        db: Session,
        order: ServiceOrder,
        *,
        org_id: str,
        user_id: str,
    ) -> ServiceOrder:
        if order.org_id != org_id:
            raise ValueError("Order not found")
        try:
            config = json.loads(order.config_json or "{}")
        except Exception:
            config = {}
        if not isinstance(config, dict):
            config = {}
        base_title = str(order.title or config.get("survey_name") or "Survey").strip() or "Survey"
        copy_title = f"Copy of {base_title}"
        config = dict(config)
        config["survey_name"] = copy_title
        new_order = ServiceOrderService.create_order(
            db,
            org_id=org_id,
            user_id=user_id,
            service_code=order.service_code,
            title=copy_title,
            config=config,
        )
        recipients = ServiceOrderService.get_recipients(db, order.id)
        if recipients:
            rows = [
                {
                    "name": str(r.name or "").strip() or "Contact",
                    "phone": str(r.phone or "").strip(),
                    "email": str(r.email or "").strip() if r.email else None,
                }
                for r in recipients
            ]
            new_order = ServiceOrderService.replace_recipients(db, new_order, rows)
        return new_order

    @staticmethod
    def _order_config_dict(order: ServiceOrder) -> dict[str, Any]:
        try:
            data = json.loads(order.config_json or "{}")
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    @staticmethod
    def _purge_recipient_dependencies(db: Session, recipient_ids: list[str]) -> None:
        if not recipient_ids:
            return
        from app.models.survey_voice_note_job import SurveyVoiceNoteJob
        from app.models.survey_session import SurveySession

        db.execute(delete(SurveyVoiceNoteJob).where(SurveyVoiceNoteJob.recipient_id.in_(recipient_ids)))
        db.execute(delete(SurveySession).where(SurveySession.recipient_id.in_(recipient_ids)))

    @staticmethod
    def delete_order(db: Session, order: ServiceOrder, *, confirm_running_delete: bool = False) -> None:
        """Soft-delete: archive order and keep recipients/results for reference."""
        if order.status in {"running", "paused", "scheduled"}:
            if not confirm_running_delete:
                raise ValueError("Stop the survey before deleting")
            ServiceOrderService.stop_order(db, order, reason="Deleted by user")
            db.refresh(order)
        if order.payment_status == "approved" and order.status not in {"completed", "cancelled", "draft", "quoted"}:
            raise ValueError("Cannot delete a paid survey that has started")
        if order.service_code == "interview":
            try:
                from app.services.interview_booking_service import InterviewBookingService, campaign_invites_were_sent

                if campaign_invites_were_sent(order):
                    InterviewBookingService.notify_campaign_closed(
                        db,
                        order,
                        reason="This interview campaign was removed.",
                    )
            except Exception:
                import logging

                logging.getLogger(__name__).exception(
                    "interview_campaign_cancel_notify_failed order_id=%s", order.id
                )
        cfg = ServiceOrderService._order_config_dict(order)
        cfg["deleted_at"] = datetime.utcnow().isoformat()
        cfg["user_deleted"] = True
        order.config_json = json.dumps(cfg, ensure_ascii=False)
        order.status = "archived"
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()

    @staticmethod
    def pause_order(db: Session, order: ServiceOrder) -> ServiceOrder:
        if order.status != "running":
            raise ValueError("Only running surveys can be paused")
        order.status = "paused"
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(order)
        return order

    @staticmethod
    def resume_order(db: Session, order: ServiceOrder) -> ServiceOrder:
        if order.status != "paused":
            raise ValueError("Survey is not paused")
        if order.payment_status != "approved":
            raise ValueError("Payment must be approved before resuming")
        order.status = "running"
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(order)
        if order.service_code == "survey":
            from app.services.survey_call_dispatch_service import SurveyCallDispatchService, is_ai_call_survey_order

            if is_ai_call_survey_order(order):
                SurveyCallDispatchService.dial_next_recipient(db, order)
        return order

    @staticmethod
    def stop_order(db: Session, order: ServiceOrder, *, reason: str | None = None) -> ServiceOrder:
        if order.service_code == "interview":
            if str(order.status or "") in {"cancelled", "completed", "archived"}:
                if str(order.status or "") == "cancelled":
                    return order
                raise ValueError("This interview campaign is already stopped")
            if order.status not in {"running", "paused", "scheduled", "paid"}:
                raise ValueError(f"Cannot stop interview campaign while status is '{order.status}'")
        elif order.status not in {"running", "paused", "scheduled", "paid"}:
            raise ValueError("Survey is not active")
        now = datetime.utcnow()
        note = (reason or "").strip()
        if note:
            order.admin_decision_note = note
        try:
            cfg = json.loads(order.config_json or "{}")
            if not isinstance(cfg, dict):
                cfg = {}
        except Exception:
            cfg = {}
        cfg["booking_closed_at"] = now.isoformat()
        if note:
            cfg["booking_closed_reason"] = note
        order.config_json = json.dumps(cfg, ensure_ascii=False)
        order.status = "cancelled"
        order.completed_at = now
        order.updated_at = now
        db.add(order)
        db.commit()
        db.refresh(order)
        try:
            from app.services.billing_reconciliation_service import BillingReconciliationService

            BillingReconciliationService.on_order_terminal(db, order, trigger="cancellation")
        except Exception:
            import logging

            logging.getLogger(__name__).exception("billing_reconciliation_cancel_failed order_id=%s", order.id)
        recipients = ServiceOrderService.get_recipients(db, order.id)
        if order.service_code == "interview":
            try:
                from app.services.interview_booking_service import (
                    InterviewBookingService,
                    campaign_invites_were_sent,
                    order_has_booking_outreach_candidates,
                )

                notify_reason = note or "This interview campaign was cancelled by the employer."
                had_prior_outreach = campaign_invites_were_sent(order) or order_has_booking_outreach_candidates(
                    db, order
                )
                close_result = InterviewBookingService.notify_campaign_closed(
                    db,
                    order,
                    reason=notify_reason,
                    include_uninvited=True,
                    notify_all_with_email=True,
                )
                cfg_stop = {}
                try:
                    cfg_stop = json.loads(order.config_json or "{}")
                    if not isinstance(cfg_stop, dict):
                        cfg_stop = {}
                except Exception:
                    cfg_stop = {}
                cfg_stop["last_campaign_close_dispatch"] = close_result
                order.config_json = json.dumps(cfg_stop, ensure_ascii=False)
                db.add(order)
                db.commit()
                db.refresh(order)
            except Exception:
                import logging

                logging.getLogger(__name__).exception(
                    "interview_campaign_cancel_notify_failed order_id=%s", order.id
                )
            for recipient in recipients:
                if str(recipient.status or "pending").lower() in {
                    "pending",
                    "scheduled",
                    "sent",
                    "calling",
                }:
                    recipient.status = "cancelled"
                    db.add(recipient)
            db.commit()
            db.refresh(order)
            return order
        for recipient in recipients:
            if str(recipient.status or "pending").lower() in {"pending", "calling"}:
                recipient.status = "cancelled"
                db.add(recipient)
        db.commit()
        db.refresh(order)
        return order

    @staticmethod
    def complete_order(db: Session, order: ServiceOrder) -> ServiceOrder:
        if order.status not in {"running", "paused"}:
            raise ValueError("Only running surveys can be finished")
        now = datetime.utcnow()
        order.status = "completed"
        order.completed_at = now
        order.updated_at = now
        db.add(order)
        db.commit()
        db.refresh(order)
        try:
            from app.services.billing_reconciliation_service import BillingReconciliationService

            BillingReconciliationService.on_order_terminal(db, order, trigger="completion")
        except Exception:
            import logging

            logging.getLogger(__name__).exception("billing_reconciliation_complete_failed order_id=%s", order.id)
        return order

    @staticmethod
    def create_order(
        db: Session,
        *,
        org_id: str,
        user_id: str,
        service_code: str,
        title: str,
        config: dict[str, Any] | None = None,
    ) -> ServiceOrder:
        import logging

        logger = logging.getLogger(__name__)
        code = str(service_code or "").strip().lower()
        logger.info("create_order entry org=%s user=%s service_code=%s", org_id, user_id, code)
        if code not in {"survey", "interview"}:
            raise ValueError("service_code must be survey or interview")
        from app.services.uk_compliance_service import UkComplianceService

        base_config = dict(config or {})
        if code == "survey":
            cfg_name = str(base_config.get("survey_name") or title or "").strip() or "Survey draft"
            base_config["survey_name"] = cfg_name
            order_title = cfg_name
        else:
            order_title = title.strip() or "Interview order"
        order = ServiceOrder(
            org_id=org_id,
            user_id=user_id,
            service_code=code,
            title=order_title,
            status="draft",
            payment_status="unpaid",
            config_json=json.dumps(base_config, ensure_ascii=False),
        )
        db.add(order)
        db.commit()
        db.refresh(order)
        logger.info("create_order persisted order_id=%s", order.id)
        order = UkComplianceService.seed_order_compliance_config(db, order, commit=True)
        from app.services.interview_campaign_service import ensure_campaign_id

        if code in {"interview", "survey"}:
            order = ensure_campaign_id(db, order)
            if code == "survey" and order.campaign_id:
                try:
                    cfg = json.loads(order.config_json or "{}")
                    if isinstance(cfg, dict):
                        cfg["survey_id"] = str(order.campaign_id)
                        order.config_json = json.dumps(cfg, ensure_ascii=False)
                        db.add(order)
                        db.commit()
                        db.refresh(order)
                except Exception:
                    pass
        if code == "interview":
            from app.services.interview_reference_service import ensure_order_reference_id

            order = ensure_order_reference_id(db, order)
        logger.info("create_order ok order_id=%s service_code=%s", order.id, order.service_code)
        return order

    @staticmethod
    def replace_recipients(db: Session, order: ServiceOrder, rows: list[dict[str, str]]) -> ServiceOrder:
        if order.payment_status == "approved":
            raise ValueError("Cannot change recipients after payment is approved")
        old_ids = list(
            db.execute(
                select(ServiceOrderRecipient.id).where(ServiceOrderRecipient.order_id == order.id)
            ).scalars()
        )
        if old_ids:
            ServiceOrderService._purge_recipient_dependencies(db, old_ids)
        db.execute(delete(ServiceOrderRecipient).where(ServiceOrderRecipient.order_id == order.id))
        for i, row in enumerate(rows, start=1):
            result_json = None
            language = str(row.get("language") or "").strip()
            if language:
                result_json = json.dumps({"language": language}, ensure_ascii=False)
            db.add(
                ServiceOrderRecipient(
                    order_id=order.id,
                    row_number=i,
                    name=row["name"],
                    phone=row["phone"],
                    email=row.get("email"),
                    result_json=result_json,
                    status="pending",
                )
            )
        order.recipient_count = len(rows)
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(order)
        return order

    @staticmethod
    def update_order(db: Session, order: ServiceOrder, payload: dict[str, Any]) -> ServiceOrder:
        if order.status == "completed":
            raise ValueError("Order can no longer be edited")
        running_like = order.payment_status == "approved" and order.status in {"running", "paused", "scheduled"}
        if running_like:
            allowed = {"scheduled_start_at", "scheduled_end_at", "title", "config"}
            payload = {k: v for k, v in payload.items() if k in allowed}
            if not payload:
                raise ValueError("This interview can only update schedule times or title while running")
        config = {}
        try:
            config = json.loads(order.config_json or "{}")
        except Exception:
            config = {}
        if "config" in payload and isinstance(payload["config"], dict):
            patch = dict(payload["config"])
            if patch.get("builder_template_ids") or patch.get("builder_step_sequence") or patch.get("builder_runtime"):
                from app.services.survey_builder_flow_service import is_builder_bound_flow
                from app.services.survey_builder_runtime_service import (
                    attach_builder_runtime_to_config,
                    load_builder_runtime,
                )

                for stale_key in (
                    "flow_snapshot",
                    "flow_snapshot_json",
                    "flow_definition_id",
                    "flow_branches",
                    "order_config_flow",
                ):
                    config.pop(stale_key, None)
                merged = {**config, **patch}
                runtime = patch.get("builder_runtime")
                if isinstance(runtime, dict) and runtime.get("step_sequence"):
                    merged = attach_builder_runtime_to_config(merged, runtime)
                else:
                    seq = merged.get("builder_step_sequence") or []
                    ids = merged.get("builder_template_ids") or []
                    legacy_runtime = load_builder_runtime(merged)
                    if isinstance(legacy_runtime, dict):
                        merged = attach_builder_runtime_to_config(merged, legacy_runtime)
                    elif isinstance(seq, list) and isinstance(ids, list) and seq and ids:
                        from app.services.survey_builder_flow_service import builder_generation_config

                        builder_core = builder_generation_config(
                            builder_step_sequence=seq,
                            builder_template_ids=[int(x) for x in ids],
                        )
                        merged = {**merged, **builder_core}
                        rebuilt = load_builder_runtime(merged)
                        if isinstance(rebuilt, dict):
                            merged = attach_builder_runtime_to_config(merged, rebuilt)
                    elif is_builder_bound_flow(merged):
                        from app.services.survey_builder_flow_service import sanitize_builder_config

                        merged = sanitize_builder_config(merged)
                config = merged
            else:
                config.update(patch)
            order.config_json = json.dumps(config, ensure_ascii=False)
        from app.services.uk_compliance_service import UkComplianceService

        order = UkComplianceService.seed_order_compliance_config(db, order, commit=False)
        if payload.get("title") and order.service_code != "survey":
            order.title = str(payload["title"]).strip()
        if order.service_code == "survey" and isinstance(config, dict):
            cfg_name = str(config.get("survey_name") or payload.get("title") or "").strip()
            if cfg_name:
                config["survey_name"] = cfg_name
                order.title = cfg_name
                order.config_json = json.dumps(config, ensure_ascii=False)
        if payload.get("run_mode") in {"manual", "scheduled"}:
            order.run_mode = str(payload["run_mode"])
        if payload.get("scheduled_start_at"):
            order.scheduled_start_at = datetime.fromisoformat(str(payload["scheduled_start_at"]).replace("Z", "+00:00")).replace(tzinfo=None)
        if payload.get("scheduled_end_at"):
            order.scheduled_end_at = datetime.fromisoformat(str(payload["scheduled_end_at"]).replace("Z", "+00:00")).replace(tzinfo=None)
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(order)
        return order

    @staticmethod
    def quote_order(db: Session, order: ServiceOrder) -> ServiceOrder:
        if order.recipient_count <= 0:
            cfg: dict[str, Any] = {}
            try:
                cfg = json.loads(order.config_json or "{}")
            except Exception:
                cfg = {}
            if isinstance(cfg, dict) and cfg.get("cv_email_enabled"):
                raise ValueError(
                    "No CVs received yet — share your job reference and careers@voxbulk.com with applicants, "
                    "or upload candidates manually"
                )
            raise ValueError("Upload a contact list before requesting a quote")
        if order.service_code == "interview":
            from app.services.interview_cv_email_service import assert_cv_collection_complete

            assert_cv_collection_complete(order)
        try:
            config = json.loads(order.config_json or "{}")
        except json.JSONDecodeError:
            config = {}
        if not isinstance(config, dict):
            config = {}
        config = dict(config)
        config["org_id"] = order.org_id
        duration = config.get("expected_duration_minutes")
        if duration is not None and config.get("duration_minutes") is None:
            config["duration_minutes"] = duration
        quote = PlatformCatalogService.calculate_quote(
            db,
            service_code=order.service_code,
            recipient_count=order.recipient_count,
            options=config,
        )
        order.quote_total_pence = int(quote["total_pence"])
        order.quote_breakdown_json = json.dumps(quote, ensure_ascii=False)
        order.status = "quoted"
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(order)
        return order

    @staticmethod
    def submit_cash_payment(db: Session, order: ServiceOrder, note: str | None = None) -> ServiceOrder:
        if order.status not in {"quoted", "draft"} or order.quote_total_pence <= 0:
            raise ValueError("Generate a quote before marking payment")
        order.payment_method = "cash"
        order.payment_status = "pending_approval"
        order.payment_note = (note or "").strip() or None
        order.status = "awaiting_payment"
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(order)
        return order

    @staticmethod
    def admin_approve_payment(db: Session, order: ServiceOrder, note: str | None = None) -> ServiceOrder:
        if order.payment_status != "pending_approval":
            raise ValueError("Order is not awaiting payment approval")
        order.payment_status = "approved"
        order.status = "paid"
        order.admin_decision_note = (note or "").strip() or None
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(order)
        from app.services.service_order_payment_workflow_service import (
            ServiceOrderPaymentWorkflowError,
            ServiceOrderPaymentWorkflowService,
        )

        try:
            ServiceOrderPaymentWorkflowService.confirm_payment_and_issue_invoice(db, order)
        except ServiceOrderPaymentWorkflowError as e:
            raise ValueError(str(e)) from e
        except Exception:
            import logging

            logging.getLogger(__name__).exception("admin_approve_payment_invoice_failed order_id=%s", order.id)
            raise ValueError("Payment approved but invoice could not be issued — check billing email") from None
        db.refresh(order)
        return order

    @staticmethod
    def admin_reject_payment(db: Session, order: ServiceOrder, note: str | None = None) -> ServiceOrder:
        if order.payment_status != "pending_approval":
            raise ValueError("Order is not awaiting payment approval")
        order.payment_status = "rejected"
        order.status = "quoted"
        order.admin_decision_note = (note or "").strip() or None
        order.updated_at = datetime.utcnow()
        db.add(order)
        db.commit()
        db.refresh(order)
        return order

    @staticmethod
    def schedule_order(db: Session, order: ServiceOrder) -> ServiceOrder:
        """Mark a paid order as scheduled/ready without dispatching calls."""
        if order.payment_status != "approved":
            raise ValueError("Payment must be approved before scheduling")
        if order.service_code == "survey":
            from app.models.organisation import Organisation
            from app.services.survey_launch_eligibility_service import (
                SurveyLaunchEligibilityError,
                SurveyLaunchEligibilityService,
            )

            org = db.get(Organisation, order.org_id)
            if org is not None:
                try:
                    SurveyLaunchEligibilityService.assert_can_launch(db, order, org)
                except SurveyLaunchEligibilityError as e:
                    raise ValueError(str(e)) from e
        if order.status in {"running", "completed"}:
            raise ValueError("Order is already running or completed")
        now = datetime.utcnow()
        if order.run_mode == "scheduled" and order.scheduled_start_at and now < order.scheduled_start_at:
            order.status = "scheduled"
        else:
            order.status = "paid"
        order.updated_at = now
        db.add(order)
        db.commit()
        db.refresh(order)
        return order

    @staticmethod
    def start_order(db: Session, order: ServiceOrder) -> ServiceOrder:
        if order.payment_status != "approved":
            raise ValueError("Payment must be approved by admin before starting")
        from app.services.service_order_payment_workflow_service import (
            ServiceOrderPaymentWorkflowError,
            ServiceOrderPaymentWorkflowService,
        )

        try:
            ServiceOrderPaymentWorkflowService.assert_launch_ready(db, order)
        except ServiceOrderPaymentWorkflowError as e:
            raise ValueError(str(e)) from e
        if order.service_code == "survey":
            from app.models.organisation import Organisation
            from app.services.survey_launch_eligibility_service import (
                SurveyLaunchEligibilityError,
                SurveyLaunchEligibilityService,
            )

            org = db.get(Organisation, order.org_id)
            if org is not None:
                try:
                    SurveyLaunchEligibilityService.assert_can_launch(db, order, org)
                except SurveyLaunchEligibilityError as e:
                    raise ValueError(str(e)) from e
        from app.services.uk_compliance_service import UkComplianceService
        from app.services.uk_compliance_audit_service import UkComplianceAuditService

        UkComplianceService.assert_order_launch_allowed(db, order)
        if order.status == "paused":
            raise ValueError("Survey is paused — use resume instead")
        if order.status in {"running", "completed"}:
            raise ValueError("Order is already running or completed")
        if order.service_code == "interview":
            from app.services.interview_cv_email_service import (
                assert_ai_call_window_after_cv_collection,
                assert_cv_collection_complete,
            )

            assert_cv_collection_complete(order)
            assert_ai_call_window_after_cv_collection(order)
            config_iv = {}
            try:
                config_iv = json.loads(order.config_json or "{}")
            except Exception:
                config_iv = {}
            delivery = PlatformCatalogService.normalize_interview_delivery(
                db, str(config_iv.get("delivery") or "ai_call")
            )
            if delivery == "ai_meeting":
                pass
        now = datetime.utcnow()
        if order.run_mode == "scheduled":
            if order.scheduled_start_at and now < order.scheduled_start_at:
                order.status = "scheduled"
                order.updated_at = now
                db.add(order)
                db.commit()
                db.refresh(order)
                return order
        if order.service_code == "survey":
            config = {}
            try:
                config = json.loads(order.config_json or "{}")
            except Exception:
                config = {}
            channel = PlatformCatalogService.resolve_survey_channel(config)
            if channel == "ai_call":
                from app.services.survey_call_dispatch_service import SurveyCallDispatchService, is_ai_call_survey_order

                if is_ai_call_survey_order(order):
                    if not SurveyCallDispatchService.start_campaign(db, order):
                        raise ValueError(
                            "Could not start AI calls — check payment, approved script, voice agent, and calling window."
                        )
                    db.refresh(order)
                    return order
            elif channel != "ai_call":
                from app.services.survey_dispatch_service import SurveyDispatchService

                order.status = "running"
                order.started_at = now
                order.updated_at = now
                db.add(order)
                db.commit()
                db.refresh(order)
                SurveyDispatchService.dispatch_survey_order(db, order)
                db.refresh(order)
                return order
        if order.service_code == "interview":
            from app.services.interview_call_dispatch_service import (
                InterviewCallDispatchService,
            )

            cfg = {}
            try:
                cfg = json.loads(order.config_json or "{}")
            except Exception:
                cfg = {}
            if cfg.get("require_booking", True) is not False:
                # Booking-based interview (candidate chooses phone or web):
                # launch_after_payment sends invitation emails and creates booking
                # tokens for EVERY interview order, not just ai_call — this closes the
                # gap where an ai_meeting order launched via /start became "running"
                # with no invite email sent.
                from app.services.interview_launch_service import InterviewLaunchService

                InterviewLaunchService.launch_after_payment(db, order)
                db.refresh(order)
                return order
            if not InterviewCallDispatchService.start_campaign(db, order):
                raise ValueError(
                    "Could not start AI interviews — check payment, approved script, voice agent, and calling window."
                )
            db.refresh(order)
            return order
        order.status = "running"
        order.started_at = now
        order.updated_at = now
        db.add(order)
        db.commit()
        db.refresh(order)
        UkComplianceAuditService.record(
            db,
            event_type="workflow.launch",
            org_id=order.org_id,
            order_id=order.id,
            detail={"service_code": order.service_code, "status": order.status},
        )
        return order

    @staticmethod
    def list_orders(db: Session, *, org_id: str | None = None, service_code: str | None = None, limit: int = 100) -> list[ServiceOrder]:
        stmt = select(ServiceOrder).order_by(ServiceOrder.created_at.desc()).limit(limit)
        if org_id:
            stmt = stmt.where(ServiceOrder.org_id == org_id)
        if service_code:
            stmt = stmt.where(ServiceOrder.service_code == service_code)
        return list(db.execute(stmt).scalars())

    @staticmethod
    def get_order(db: Session, order_id: str, *, org_id: str | None = None) -> ServiceOrder | None:
        stmt = select(ServiceOrder).where(ServiceOrder.id == order_id)
        if org_id:
            stmt = stmt.where(ServiceOrder.org_id == org_id)
        return db.execute(stmt).scalar_one_or_none()

    @staticmethod
    def resolve_order_ref(db: Session, ref: str, *, org_id: str | None = None) -> ServiceOrder | None:
        """Match order UUID, campaign_id (VB-CMP-…), or reference_id (VB-INT-…)."""
        key = str(ref or "").strip()
        if not key:
            return None
        order = ServiceOrderService.get_order(db, key, org_id=org_id)
        if order is not None:
            return order
        upper = key.upper()
        stmt = select(ServiceOrder).where(
            (ServiceOrder.campaign_id == upper) | (ServiceOrder.reference_id == upper)
        )
        if org_id:
            stmt = stmt.where(ServiceOrder.org_id == org_id)
        return db.execute(stmt.limit(1)).scalar_one_or_none()

    @staticmethod
    def get_recipients(db: Session, order_id: str) -> list[ServiceOrderRecipient]:
        return list(
            db.execute(
                select(ServiceOrderRecipient)
                .where(ServiceOrderRecipient.order_id == order_id)
                .order_by(ServiceOrderRecipient.row_number.asc())
            ).scalars()
        )

    @staticmethod
    def get_recipient(db: Session, order_id: str, recipient_id: str) -> ServiceOrderRecipient | None:
        return db.execute(
            select(ServiceOrderRecipient).where(
                ServiceOrderRecipient.order_id == order_id,
                ServiceOrderRecipient.id == recipient_id,
            )
        ).scalar_one_or_none()
